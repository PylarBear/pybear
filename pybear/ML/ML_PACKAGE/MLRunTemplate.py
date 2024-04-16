import sys, inspect, time, datetime
from copy import deepcopy
import pandas as pd, numpy as np

pd.set_option("display.max_rows", None, "display.max_columns", None)
from openpyxl import Workbook
from ML_PACKAGE.openpyxl_shorthand import openpyxl_write as ow
from general_sound import winlinsound as wls
from data_validation import validate_user_input as vui
from debug import get_module_name as gmn
from general_data_ops import get_shape as gs
from general_list_ops import list_select as ls, manual_num_seq_list_fill as mnslf

from ML_PACKAGE.GENERIC_PRINT import DictMenuPrint as dmp, general_ml_setup_dump as gmsd, print_vectors as pv, \
    general_dev_results_dump as gdrd, ObjectsExcelDump as oed, kfold_results_print as krp, show_time as st
from MLObjects import MLTrainDevTestSplit as mltdts
from MLObjects.SupportObjects import master_support_object_dict as msod
from MLObjects import MLObject as mlo, MLRowColumnOperations as mlrco
from read_write_file.generate_full_filename import base_path_select as bps, filename_enter as fe
from ML_PACKAGE import TestPerturber as tp



# X = module-specific
# MODULE-SPECIFIC MAIN MENU DECLARATIONS ############################################################################
#Xmodule_specific_main_menu_cmds()       # returns module-specific top-level menu options, overwritten in child
#Xmodule_specific_main_menu_operations() # returns module-specific execution code for post-run cmds, overwritten in child
# END MODULE-SPECIFIC MAIN MENU DECLARATIONS ############################################################################

# GENERIC FUNCTIONS ####################################################################################################################
#Xhyperparameter_display_module()        # print hyperparameter settings to screen
# row_column_display_select()            # user selects rows and columns for display on screen and filedump
# filedump_path()                        # hard-coded directory, user-entered filename
# filedump_general_ml_setup_module()     # module for filedump of general setup for all ML packages, used for train, dev, & test filedump
#Xfiledump_package_specific_setup_module()  # module for filedump of setup for specific package, used for train, dev, & test filedump, overwritten in child
# dev_or_test_draw_params()              # module that returns the sample size to pull from train data for dev or test data sets
# random_dev_or_test_draw()              # randomly select user-specified quantity of examples for dev & test sets from (remaining) examples in TRAIN_SWNL
# partition_dev_or_test_draw()           # partition (remaining) examples in TRAIN_SWNL for selection into dev or test
# category_dev_or_test_draw()            # select examples for dev & test sets using categories in TRAIN_SWNL
#Xgeneric_ml_core_calc_algorithm()       # return module for package-specific core output algorithm, returns final output vector, overwritten in child
#Xtrain_data_calc_algorithm()            # module for passing train data thru package-specific core algorithm, returns final output vector
#Xdev_data_calc_algorithm()              # module for passing dev data thru package-specific core algorithm, returns final output vector
#Xtest_data_calc_algorithm()             # module for passing test data thru package-specific core algorithm, returns final output vector
#Xgeneric_ml_core_error_algorithm()      # return module for package-specific core error algorithm, returns total error, overwritten in child
#Xtrain_data_error_algorithm()           # module for passing train data thru package-specific error algorithm, returns total error
#Xdev_data_error_algorithm()             # module for passing dev data thru package-specific error algorithm, returns total error
#Xtest_data_error_algorithm()            # module for passing test data thru package-specific error algorithm, returns total error
# output_calc()                          # module for running user-selected test data into trained ML algorithm and calculating and sorting results for all ML packages
# _len()                                 # return len of list-type or dict object
# END GENERIC FUNCTIONS ####################################################################################################################

# TRAIN BUILD DEFINITIONS ################################################################################################
# reset_train_data()                     # reset train data to ORIGINAL DATA
# END TRAIN BUILD DEFINITIONS ################################################################################################

# DEV BUILD DEFINITIONS ##################################################################################################
# base_dev_build_module()                # module with code for building dev objects for all ML packages
#Xsub_dev_build_module()                 # return module with package-specific dev objects build code, overwritten in child
#Xsub_dev_build_cmds()                   # return list with package-specific dev objects build prompts, overwritten in child
# dev_build()                            # module for building dev objects, train objects starts as original objects, then dev objects are extracted from train objects
# END DEV BUILD DEFINITIONS ##################################################################################################

# TEST BUILD DEFINITIONS ##################################################################################################
# base_test_build_module()               # module with code for building test objects for all ML packages
#Xsub_test_build_module()                # return module with package-specific test objects build code, overwritten in child
#Xsub_test_build_cmds()                  # return list with package-specific test objects build prompts, overwritten in child
# test_build()                           # module for building test objects, train objects starts as original objects, then test objects are extracted from train objects
# END TEST BUILD DEFINITIONS ##################################################################################################

# TRAIN CALC DEFINITIONS ###################################################################################################
#Xcore_training_code()                   # return unique core training algorithm for particular ML package
#Xkfold_core_training_code()             # return unique kfold core training algorithm for particular ML package
# kfold_config()                         # configures number of regularization trials, factors, and early stopping
# END TRAIN CALC DEFINITIONS ###################################################################################################

# DEV CALC DEFINITIONS ##################################################################################################
# rglztn_partition_iterator()            # only used to save space in base_dev_calc_module()
# base_dev_calc_module()                 # module for performing dev calculations for all ML packages
#Xsub_dev_calc_module()                  # return module with package-specific commands for performing dev calculations
#Xsub_dev_calc_cmds()                    # return list with package-specific dev calc prompts
# END DEV CALC DEFINITIONS ##################################################################################################

# TEST CALC DEFINITIONS ##################################################################################################
# base_test_calc_module()                # module for performing test calculations for all ML packages
#Xsub_test_calc_module()                 # return module with package-specific commands for performing test calculations
#Xsub_test_calc_cmds()                   # return list with package-specific test calc prompts
# END TEST CALC DEFINITIONS ##################################################################################################

# TRAIN DATA DISPLAY ##############################################################################################################
#Xtrain_summary_statistics_module()      # returns module for printing summary statistics of train data for particular ML package
#Xprint_train_results_module()           # returns module for printing train results to screen for particular ML package
#Xtrain_filedump_module()                # returns module for filedump of train results for particular ML package
# END TRAIN DATA DISPLAY ##############################################################################################################

# DEV DATA DISPLAY ###########################################################################################################
# dev_summary_statistics_module()        # returns module for printing summary statistics of dev data for all ML packages
# print_dev_results_module()             # returns module for printing dev results to screen for all ML packages
# dev_filedump_module()                  # returns module for filedump of dev results for all ML packages
# END DEV DATA DISPLAY ############################################################################################A##############

# CALC DATA DISPLAY ###########################################################################################################
#Xtest_summary_statistics_module()       # returns module for printing summary statistics of test data for particular ML package
#Xprint_test_results_module()            # returns module for printing test results to screen for particular ML package
#Xtest_filedump_module()                 # returns module for filedump of test results for particular ML package
# END CALC DATA DISPLAY ###########################################################################################################

# base_return_fxn()                      # specify what to return for all ML packages
#Xreturn_fxn()                           # return self.base_return_fxn()
# run()                                  # MLRunTemplate loop





class MLRunTemplate:

    def __init__(self,
                    standard_config,
                    sub_config,
                    SUPER_RAW_NUMPY_LIST,
                    RAW_SUPOBJS,
                    SUPER_WORKING_NUMPY_LIST,
                    WORKING_SUPOBJS,
                    data_run_orientation,
                    target_run_orientation,
                    refvecs_run_orientation,
                    WORKING_CONTEXT,
                    WORKING_KEEP,
                    TRAIN_SWNL,
                    DEV_SWNL,
                    TEST_SWNL,
                    split_method,
                    LABEL_RULES,
                    number_of_labels,
                    event_value,
                    negative_value,
                    conv_kill,
                    pct_change,
                    conv_end_method,
                    rglztn_type,
                    rglztn_fctr,
                    bypass_validation,
                    module):

        self.this_module = gmn.get_module_name(str(sys.modules[module]))
        fxn = '__init__'

        self.bypass_validation = bypass_validation

        self.standard_config = standard_config
        self.sub_config = sub_config

        self.SUPER_RAW_NUMPY_LIST = SUPER_RAW_NUMPY_LIST
        self.RAW_SUPOBJS = RAW_SUPOBJS
        self.SUPER_WORKING_NUMPY_LIST = SUPER_WORKING_NUMPY_LIST
        self.WORKING_SUPOBJS = WORKING_SUPOBJS

        self.data_run_orientation = data_run_orientation
        self.target_run_orientation = target_run_orientation
        self.refvecs_run_orientation = refvecs_run_orientation

        self.WORKING_CONTEXT = WORKING_CONTEXT
        self.WORKING_KEEP = WORKING_KEEP
        self.TRAIN_SWNL = TRAIN_SWNL
        self.DEV_SWNL = DEV_SWNL
        self.TEST_SWNL = TEST_SWNL
        self.split_method = split_method
        self.LABEL_RULES = LABEL_RULES
        self.number_of_labels = number_of_labels
        self.event_value = event_value
        self.negative_value = negative_value
        self.conv_kill = conv_kill
        self.pct_change = pct_change
        self.conv_end_method = conv_end_method
        self.rglztn_type = rglztn_type
        self.rglztn_fctr = rglztn_fctr
        self.early_stop_interval = 1e12
        # ONLY FOR SVM & MLR... FOR MANAGING K-FOLD KERNEL REBUILDS
        self.rebuild_kernel = True

        self.SUPER_NUMPY_DICT = dict({0: 'DATA', 1: 'TARGET', 2: 'REFERENCE'})

        # BASE "MAIN MENU" DECLARATIONS ########################################################################################################
        self.BASE_MAIN_MENU_CMDS = {
                                    'o': '\nprint hyperparameter settings',
                                    'l': 'display objects and columns',
                                    'x': 'print frequencies',
                                    'n': 'dump WORKING DATA objects to file',
                                    'i': '\n\ngenerate \ reset train objects',
                                    'b': 'run train without reconfig',
                                    'v': 'print train target & output vectors',
                                    's': 'print train summary stats',
                                    'p': 'print train results to screen',
                                    'f': 'dump train results to file',
                                    'h': '\n\ngenerate \ reset dev objects',
                                    'g': 'run dev',
                                    'w': 'run k-fold cross validation',
                                    'j': 'print dev results to screen',
                                    'k': 'dump dev results to file',
                                    't': '\n\ngenerate \ reset test objects',
                                    'r': 'calc test data',
                                    'e': 'print test target & output vectors',
                                    'u': 'print test summary stats',
                                    'c': 'print test results to screen',
                                    'd': 'dump test results to file',
                                    'z': 'dump all results to file',
                                    'a': '\n\naccept, return to config-run menu',
                                    'q': 'quit'
        }

        if len(self.BASE_MAIN_MENU_CMDS | self.module_specific_main_menu_cmds()) != \
                            len(self.BASE_MAIN_MENU_CMDS) + len(self.module_specific_main_menu_cmds()):
            raise ValueError(f'\n*** {self.this_module} >>> BASE_MAIN_MENU_CMDS AND module_specific_main_menu_cmds HAVE A DUPLICATE KEY. ***\n')

        self.base_disallowed = ''
        self.post_run_select = ''
        # END BASE "MAIN MENU" DECLARATIONS ########################################################################################################

        # "TRAIN BUILD" BASE MENU DECLARATIONS #########################################################################################################
            # TRAIN MUST ALWAYS BE BUILT FROM INCOMING DATA, THEREFORE NO MENU

        # "END TRAIN BUILD" BASE MENU DECLARATIONS #########################################################################################################

        # "DEV BUILD" BASE MENU DECLARATIONS #########################################################################################################
        self.BASE_DEV_BUILD_CMDS = {
                                    'r': 'random draw from training DATA',
                                    's': 'sub-partition from training DATA',
                                    'd': 'use original data',
                                    'f': 'load from file',
                                    't': 'use dev matrices currently loaded',
                                    'u': 'select block from train data',
                                    'v': 'draw based on categories in a column',
                                    'b': 'load standard config or manual build',
                                    'n': 'none',
                                    'a': 'accept & exit dev objects build'
        }

        if len(self.BASE_DEV_BUILD_CMDS | self.sub_dev_build_cmds()) != \
                            len(self.BASE_DEV_BUILD_CMDS) + len(self.sub_dev_build_cmds()):
            raise ValueError(f'\n*** {self.this_module} >>> BASE_DEV_BUILD_CMDS AND sub_dev_build_cmds HAVE A DUPLICATE KEY. ***\n')

        self.dev_build_select = ''
        # END "DEV BUILD" BASE MENU DECLARATIONS #########################################################################################################

        # "TEST BUILD" BASE MENU DECLARATIONS #########################################################################################################
        self.BASE_TEST_BUILD_CMDS = {
                                    'd': 'use original DATA',
                                    'r': 'random draw from training data',
                                    's': 'sub-partition from training DATA',
                                    'p': 'generate all permutations',
                                    'f': 'load from file',
                                    't': 'use test matrices currently loaded',
                                    'u': 'User select block from train data',
                                    'v': 'draw based on categories in a column',
                                    'b': 'load standard config or manual build',
                                    'o': 'overwrite TEST_SWNL',
                                    'n': 'none',
                                    'a': 'accept & exit test objects build'
        }

        if len(self.BASE_TEST_BUILD_CMDS | self.sub_test_build_cmds()) != \
                            len(self.BASE_TEST_BUILD_CMDS) + len(self.sub_test_build_cmds()):
            raise ValueError(f'\n*** {self.this_module} >>> BASE_TEST_BUILD_CMDS AND sub_test_build_cmds HAVE A DUPLICATE KEY. ***\n')


        self.test_build_select = ''
        # END "TEST BUILD" BASE MENU DECLARATIONS #########################################################################################################

        # "TRAIN CALC" BASE MENU DECLARATIONS ######################################################################################
            # NO MENU, "TRAIN CALC" IS core_training_code ONLY
        # END "TRAIN CALC" BASE MENU DECLARATIONS ######################################################################################

        # "DEV CALC" BASE MENU DECLARATIONS ######################################################################################
        self.BASE_DEV_CALC_CMDS = {
                                    'd': 'run truncated k-fold dev loop on one partition',
                                    's': 'run standard dev loop on current dev objects',
                                    'p': 'print dev results to screen',
                                    'a': 'accept & exit dev calc'
        }

        if len(self.BASE_DEV_CALC_CMDS | self.sub_dev_calc_cmds()) != len(self.BASE_DEV_CALC_CMDS) + len(self.sub_dev_calc_cmds()):
            raise ValueError(f'\n*** {self.this_module} >>> BASE_DEV_CALC_CMDS AND sub_dev_calc_cmds HAVE A DUPLICATE KEY. ***\n')

        self.dev_calc_select = ''
        # END "DEV CALC" BASE MENU DECLARATIONS ######################################################################################


        # "TEST CALC" BASE MENU DECLARATIONS ######################################################################################
        self.BASE_TEST_CALC_CMDS = {
                                    'n': 'run one-off pass of test data',
                                    'p': 'run test data perturber',
                                    'a': 'accept & exit test calc'
        }

        if len(self.BASE_TEST_CALC_CMDS | self.sub_test_calc_cmds()) != len(self.BASE_TEST_CALC_CMDS) + len(self.sub_test_calc_cmds()):
            raise ValueError(f'\n*** {self.this_module} >>> BASE_TEST_CALC_CMDS AND sub_test_calc_cmds HAVE A DUPLICATE KEY. ***\n')

        self.test_calc_select = ''
        # END "TEST CALC" BASE MENU DECLARATIONS ######################################################################################


        # CALCULATION & DISPLAY PLACEHOLDERS ##############################################################################

        self.TRAIN_CSUTM_DF = []
        self.TRAIN_OUTPUT_VECTOR = []
        self.train_error = 0
        self.train_has_been_run = False

        self.RGLZTN_FACTORS = []
        self.DEV_ERROR = [[]]
        self.DEV_CSUTM_DF = []
        self.DEV_OUTPUT_VECTOR = []
        self.dev_error = 0

        self.CSUTM_DF = pd.DataFrame({})
        self.PERTURBER_RESULTS = pd.DataFrame({})
        self.TEST_OUTPUT_VECTOR = []
        self.test_error = 0

        self.display_select = ''
        self.display_rows = ''
        self.DISPLAY_COLUMNS = []

        self.wb = Workbook()
        self.full_path = ''
        # END CALCULATION & DISPLAY DECLARATIONS ##########################################################################


    # END __init__ ########################################################################################################
    #######################################################################################################################


    # MODULE-SPECIFIC MAIN MENU DECLARATIONS ##############################################################################
    def module_specific_main_menu_cmds(self):
        # module-specific top-level menu options
        return {}  # SPECIFIED IN CHILDREN   # CANT USE 'ABCDEFGHIJKLNOPQRSTUVWXZ'

    def module_specific_main_menu_operations(self):
        # execution code for post-run cmds
        pass    # CANT USE 'ABCDEFGHIJKLNOPQRSTUVWXZ'
    # END MODULE-SPECIFIC MAIN MENU DECLARATIONS ############################################################################

    # GENERIC FUNCTIONS ####################################################################################################################
    def hyperparameter_display_module(self):
        # print hyperparameter settings to screen
        # OVERWRITTEN BY CHILD
        pass


    def row_column_display_select(self):
        while True:

            # SELECT ROWS TO DISPLAY FROM RESULTS
            if not self.CSUTM_DF.equals(pd.DataFrame({})):
                print(f"\nTEST RESULTS HAS {len(self.CSUTM_DF)} ROWS.\n")
                self.display_select = vui.validate_user_str(
                    f'Show head(h), tail(t), both(b), or all(a)? > ', 'HTBA')

                if self.display_select == 'A':
                    if vui.validate_user_str(f"TEST MATRIX has {len(self.CSUTM_DF)} examples.  " + \
                                             f"Accept display of all rows? (y/n) > ", 'YN') == 'N': continue
                    self.display_rows = len(self.CSUTM_DF)
                else: self.display_rows = vui.validate_user_int(f'Display how many rows? > ', min=1)

                self.DISPLAY_COLUMNS = \
                    ls.list_custom_select([_ for _ in self.CSUTM_DF], 'idx')
            else:
                print(f"\n*** TEST RESULTS HAVE NOT BEEN GENERATED YET ***\n")

            if vui.validate_user_str(f'Accept display setup? (y/n) > ', 'YN') == 'Y':
                break


    def filedump_path(self):
        # hard-coded directory, user-entered filename
        base_path = bps.base_path_select()
        file_name = fe.filename_wo_extension()
        self.full_path = base_path + file_name + '.xlsx'
        self.wb.save(self.full_path)


    def filedump_general_ml_setup_module(self):
        # returns module for filedump of general setup for all ML packages, used for run & test filedump
        self.wb = gmsd.general_ml_setup_dump(self.wb, self.standard_config, *self.WORKING_SUPOBJS, self.WORKING_CONTEXT,
            self.WORKING_KEEP, self.split_method, self.LABEL_RULES, self.number_of_labels, self.event_value, self.negative_value)


    def filedump_package_specific_setup_module(self):
        # returns module for filedump of setup for specific package, used for run & test filedump
        # overwritten in child
        pass


    def dev_or_test_draw_params(self):
        # select method and return number of rows to pull
        train_rows = gs.get_shape('TRAIN_DATA', self.TRAIN_SWNL[0], self.data_run_orientation)[0]

        percent_or_number_method = \
            vui.validate_user_str(f'Select a percent(p) or number(n) to select from (remaining) data > ', 'NP')
        if percent_or_number_method == 'P':
            percent_or_number = vui.validate_user_float(
                f'Enter percent to randomly select > ', min=0, max=100 * (train_rows - 1) / (train_rows))
            size = int(np.floor(train_rows * percent_or_number / 100))
            del percent_or_number
        elif percent_or_number_method == 'N':
            size = vui.validate_user_int(f'Enter number to select > ', min=0, max=train_rows - 1)

        del train_rows, percent_or_number_method

        return size


    def random_dev_or_test_draw(self, size=None):
        '''randomly select user-specified quantity of examples from (remaining) examples in TRAIN_SWNL'''

        # DEV_OR_TEST_SWNL COULD BE DEV_SWNL OR TEST_SWNL, BASED ON WHAT THIS FXN'S RETURN IS ASSIGNED TO

        if size is None: size = self.dev_or_test_draw_params()

        print(f'\nSelecting DEV/TEST objects from TRAIN objects and deleting selected from TRAIN objects...')

        RandomClass = mltdts.MLTrainDevTestSplitNEWSUPOBJS(DATA=self.TRAIN_SWNL[0],
                                                            TARGET=self.TRAIN_SWNL[1],
                                                            REFVECS=self.TRAIN_SWNL[2],
                                                            data_given_orientation=self.data_run_orientation,
                                                            target_given_orientation=self.target_run_orientation,
                                                            refvecs_given_orientation=self.refvecs_run_orientation,
                                                            bypass_validation=self.bypass_validation
        )

        RandomClass.random(dev_count=None, dev_percent=None, test_count=size, test_percent=None)

        self.TRAIN_SWNL = list(RandomClass.TRAIN)
        RETURN_SWNL = list(RandomClass.TEST)
        del RandomClass

        return RETURN_SWNL


    def partition_dev_or_test_draw(self, number_of_partitions, partition_number):
        '''partition (remaining) examples in TRAIN_SWNL for selection into dev or test'''

        # DEV_OR_TEST_SWNL COULD BE DEV_SWNL OR TEST_SWNL, BASED ON WHAT THIS FXN'S RETURN IS ASSIGNED TO

        PartitionClass = mltdts.MLTrainDevTestSplitNEWSUPOBJS(DATA=self.TRAIN_SWNL[0],
                                                            TARGET=self.TRAIN_SWNL[1],
                                                            REFVECS=self.TRAIN_SWNL[2],
                                                            data_given_orientation=self.data_run_orientation,
                                                            target_given_orientation=self.target_run_orientation,
                                                            refvecs_given_orientation=self.refvecs_run_orientation,
                                                            bypass_validation=self.bypass_validation
        )

        PartitionClass.partition(number_of_partitions=number_of_partitions, dev_partition_number=None, test_partition_number=partition_number)

        self.TRAIN_SWNL = list(PartitionClass.TRAIN)
        RETURN_SWNL = list(PartitionClass.TEST)
        del PartitionClass

        return RETURN_SWNL


    def category_dev_or_test_draw(self):
        '''select examples for dev & test sets using categories in TRAIN_SWNL'''
        # DEV_OR_TEST_SWNL COULD BE DEV_SWNL OR TEST_SWNL, BASED ON WHAT THIS FXN'S RETURN IS ASSIGNED TO

        PartitionClass = mltdts.MLTrainDevTestSplitNEWSUPOBJS(DATA=self.TRAIN_SWNL[0],
                                                            TARGET=self.TRAIN_SWNL[1],
                                                            REFVECS=self.TRAIN_SWNL[2],
                                                            data_given_orientation=self.data_run_orientation,
                                                            target_given_orientation=self.target_run_orientation,
                                                            refvecs_given_orientation=self.refvecs_run_orientation,
                                                            bypass_validation=self.bypass_validation
        )

        PartitionClass.category(object_name_for_dev=None,
                                 dev_column_name_or_idx=None,
                                 DEV_SPLIT_ON_VALUES_AS_LIST=None,
                                 object_name_for_test=None,
                                 test_column_name_or_idx=None,
                                 TEST_SPLIT_ON_VALUES_AS_LIST=None,
                                 DATA_FULL_SUPOBJ=self.WORKING_SUPOBJS[0],
                                 TARGET_FULL_SUPOBJ=self.WORKING_SUPOBJS[1],
                                 REFVECS_FULL_SUPOBJ=self.WORKING_SUPOBJS[2])


        self.TRAIN_SWNL[0], self.TRAIN_SWNL[1], self.TRAIN_SWNL[2] = PartitionClass.TRAIN
        RETURN_SWNL = list(PartitionClass.TEST)
        del PartitionClass
        return RETURN_SWNL


    def generic_ml_core_calc_algorithm(self):
        # return module for package-specific core output algorithm, returns final output vector, overwritten in child
        # not in use
        pass


    def train_data_calc_algorithm(self):
        # module for passing train data thru package-specific core algorithm, returns final output vector
        pass


    def dev_data_calc_algorithm(self):
        # module for passing dev data thru package-specific core algorithm, returns final output vector
        pass


    def test_data_calc_algorithm(self):
        # module for passing test data thru package-specific core algorithm, returns final output vector
        pass


    def generic_ml_core_error_algorithm(self):
        # module for package-specific core error algorithm, returns total error
        # not in use
        pass


    def train_data_error_algorithm(self):
        # module for passing train data thru package-specific error algorithm, returns total error
        pass


    def dev_data_error_algorithm(self):
        # module for passing dev data thru package-specific error algorithm, returns total error
        pass


    def test_data_error_algorithm(self):
        # module for passing test data thru package-specific error algorithm, returns total error
        pass


    def output_calc(self, calc_algorithm, CALC_OBJECT, object_name):
        # module for running user-selected test data into trained ML algorithm and calculating and sorting results for all ML packages
        st.show_start_time(f'Calculate {object_name} CASES')
        print(f'\nCalculating {object_name} CASES...')

        CALC_DATA_WIP = np.array(CALC_OBJECT[0], dtype=object)
        CALC_DATA_HEADER_WIP = np.array(CALC_OBJECT[1], dtype=str)
        CALC_TARGET_WIP = np.array(CALC_OBJECT[2], dtype=float)
        CALC_TARGET_HEADER_WIP = np.array(CALC_OBJECT[3], dtype=str)
        CALC_REFERENCE_VECTORS_WIP = np.array(CALC_OBJECT[4], dtype=object)
        CALC_REFERENCE_VECTORS_HEADER_WIP = np.array(CALC_OBJECT[5], dtype=str)
        del CALC_OBJECT

        # CALCULATE SCORE VECTOR#############################################################################################
        CALC_OUTPUT_VECTOR = calc_algorithm    # OUTPUT_VECTOR produced by calc_algorithm MUST BE [[]]

        if len(CALC_OUTPUT_VECTOR) > 1:
            # 2-15-22 GENERALIZING THIS TO ACCOMODATE SOFTMAX.  FOR EACH EXAMPLE, FIND THE HIGHEST PROBABILITY ACROSS ALL THE CLASSES,
            # THEN USE THIS NUMBER AS SCORE FOR SCORING PURPOSES.
            print(f'\n*** TARGET IS MULTICLASS. CREATING A SCORE VECTOR THAT RETURNS THE HIGHEST PROBABILITY FOR EACH EXAMPLE ***')
            SCORE_VECTOR = np.max(np.array(CALC_OUTPUT_VECTOR), axis=0)
        else:
            SCORE_VECTOR = np.array(CALC_OUTPUT_VECTOR[0], dtype=np.float64)

        # SORTING############################################################################################################
        #  CREATE AN ARGSORT KEY  USED TO SORT CALC_DATA_WIP, TARGET, REF_VECS AND SCORE VECTORS ############################

        OBJECT_DICT = dict((("T", "TARGET"), ("O", "OUTPUT"), ("N", "None")))
        ORDER_DICT = dict((("A", "ASCENDING"), ("D", "DESCENDING")))

        while True:

            is_sorting = {'Y': True, 'N': False}[vui.validate_user_str(f'\nSort results? (y/n) > ', 'YN')]
            if is_sorting:
                first_sort_column = vui.validate_user_str(f'First sort by TARGET(t) or OUTPUT(o) > ', 'TO')
                first_column_order = vui.validate_user_str(
                    f'Sort {OBJECT_DICT[first_sort_column]} ascending(a), descending(d) > ', 'AD')
                first_column_text = f'{OBJECT_DICT[first_sort_column]} {ORDER_DICT[first_column_order]}'

                second_sort_column = vui.validate_user_str(
                    f'Then sort by {["OUTPUT(o)" if first_sort_column == "T" else "TARGET(t)"][0]} or None(n) > ',
                    'ONT')
                second_column_order = None
                if second_sort_column != 'N':
                    second_column_order = vui.validate_user_str(
                        f'Sort {OBJECT_DICT[second_sort_column]} ascending(a), descending(d) > ', 'AD')
                    second_column_text = f'{OBJECT_DICT[second_sort_column]} {ORDER_DICT[second_column_order]}'
                elif second_sort_column == 'N':
                    second_column_text = ''
            else:
                first_column_text = f'not to sort'

            print(f'\nUser selected {["to sort" if is_sorting else "not to sort"][0]}' + \
                  str([f' {OBJECT_DICT[first_sort_column]} {ORDER_DICT[first_column_order]}' if is_sorting else ''][
                          0]) + \
                  str([f' then by {OBJECT_DICT[second_sort_column]} {ORDER_DICT[second_column_order]}.'
                       if is_sorting and second_sort_column != 'N' else '.'][0])
                  )

            _ = vui.validate_user_str(
                [f'Accept(a), start over(s), no sort(n) > ' if is_sorting else 'Accept(a), start over(s) > '][0], 'ASN')
            if _ == 'A':
                break
            elif _ == 'S':
                continue
            elif _ == 'N':
                is_sorting = False
                break

        #   CREATE A DF W TARGET & SCORE FOR EASY SORTING
        TARGET_AND_SCORE_HOLDER = np.empty((0, len(SCORE_VECTOR)), dtype=np.float64)
        TARGET_AND_SCORE_HOLDER = np.vstack((TARGET_AND_SCORE_HOLDER, CALC_TARGET_WIP.copy(), SCORE_VECTOR.copy()))
        TARGET_AND_SCORE_HOLDER = pd.DataFrame(data=TARGET_AND_SCORE_HOLDER.transpose(), columns=['TARGET', 'OUTPUT'],
                                              dtype=float)

        if is_sorting:
            SORT_BY = [OBJECT_DICT[first_sort_column]]
            SORT_ORDER = [True if first_column_order == "A" else False]
            if second_sort_column == 'N':
                pass
            elif second_sort_column != 'N':  # "O" IF #1 WAS "T" & VICE VERSA
                SORT_BY.append(OBJECT_DICT[second_sort_column])
                SORT_ORDER.append([True if second_column_order == "A" else False][0])

            TARGET_AND_SCORE_HOLDER.sort_values(by=SORT_BY, ascending=SORT_ORDER, inplace=True)

        elif not is_sorting:
            pass  # NO CHANGE TO DF index

        SORT_KEY = TARGET_AND_SCORE_HOLDER.index.to_numpy(dtype=np.int32)

        del TARGET_AND_SCORE_HOLDER

        # END CREATE AN ARGSORT KEY ##########################################################################################

        # SORT BY SORT KEY          # DATA IS SORTED BELOW DURING DECONSTRUCTION
        SORTED_SCORE_VECTOR = SCORE_VECTOR[SORT_KEY]

        SORTED_CALC_OUTPUT_VECTOR = CALC_OUTPUT_VECTOR[..., SORT_KEY]
        SORTED_CALC_TARGET_WIP = CALC_TARGET_WIP[..., SORT_KEY]
        SORTED_CALC_REFERENCE_VECTORS_WIP = CALC_REFERENCE_VECTORS_WIP[..., SORT_KEY]

        SORTED_CALC_TARGET_HEADER_WIP = CALC_TARGET_HEADER_WIP
        SORTED_CALC_REFERENCE_VECTORS_HEADER_WIP = CALC_REFERENCE_VECTORS_HEADER_WIP

        del SCORE_VECTOR, CALC_TARGET_WIP, CALC_TARGET_HEADER_WIP, CALC_REFERENCE_VECTORS_WIP, \
            CALC_REFERENCE_VECTORS_HEADER_WIP

        ####################################################################################################################
        # SORT DATA BY SORT_KEY ###############################################################################################

        # 9/24/22 THINK ABOUT CHOPPING THESE OBJECTS HERE BASED ON DISPLAY SELECTIONS.....
        # 9/25/22 ... DISPLAY SELECTION ARE MADE AFTER CSUTM_DF IS BUILT :(

        SORTED_CALC_DATA_WIP = CALC_DATA_WIP.transpose()[SORT_KEY].transpose()
        # COLUMN_SORT = np.indices((len(CALC_DATA_WIP[0]), len(CALC_DATA_WIP)))[1]  # NO SORT, JUST ORIGINAL IDXs
        # ROW_SORT = np.tile(SORT_KEY, (len(CALC_DATA_WIP), 1)).transpose()
        # SORTED_CALC_DATA_WIP = CALC_DATA_WIP[COLUMN_SORT, ROW_SORT]
        # del COLUMN_SORT, ROW_SORT

        SORTED_CALC_DATA_HEADER_WIP = CALC_DATA_HEADER_WIP
        del SORT_KEY, CALC_DATA_WIP, CALC_DATA_HEADER_WIP

        # RECONSTRUCTING ORIGINAL RAW DATA FROM THE (STILL EXPANDED) CALC DATA MATRIX ##########################################
        # CONSOLIDATED_SORTED_CALC_DATA_WIP #############################################################################################
        # MODIFY SORTED_CALC_DATA_WIP IN-PLACE

        dum_indicator = ' - '
        intx_indicator = '_x_'
        # lag_indicator = ???

        working_feature_idx = 0
        while True:

            feature = SORTED_CALC_DATA_HEADER_WIP[0][working_feature_idx]

            if self.WORKING_DATA_SUPOBJS[msod.QUICK_POSN_DICT()['MODIFIEDDATATYPES']][working_feature_idx] in ['FLOAT', 'INT']:
                # IF IS FLOAT OR INT, JUST CARRY ALONG
                working_feature_idx += 1
            elif intx_indicator in feature:  # MUST BE BIN, WOULD HAVE BEEN PICKED UP ABOVE
                # IF intx_indicator IS IN feature THEN IS A (NON-FLOAT/NON-INT) INTX, SO CARRY THRU W/O CONSOLIDATION
                working_feature_idx += 1
            elif self.WORKING_DATA_SUPOBJS['MODIFIEDDATATYPES'][working_feature_idx] == 'BIN':  # IF IS BINARY, CONSOLIDATE
                # 9-24-22 THIS IS A COPOUT, JUST SPLITTING ON " - " FOR NOW TO GET FEATURES AND CATEGORIES.  IN THE FUTURE
                # HAVE TO FIGURE OUT A WAY TO DO THIS ROBUSTLY FOR LAG, INTERACTIONS, ETC.

                # AT THIS POINT IT IS implicit THAT "and intx_indicator not in feature:"

                raw_feature = feature[:feature.index(dum_indicator)]
                # REMEMBER THAT "DUMMY" WAS PUT AT THE END DURING EXPANSION!
                try:
                    category_label = \
                        feature[len(raw_feature)+len(dum_indicator):SORTED_CALC_DATA_HEADER_WIP[0][working_feature_idx].index(' DUMMY')]
                except:
                    category_label = feature[len(raw_feature) + len(dum_indicator):]
                # IF raw_feature NOT IN SORTED_CALC_DATA_HEADER_WIP, THEN NEW RAW FEATURE
                if raw_feature not in SORTED_CALC_DATA_HEADER_WIP[0]:
                    # CHANGE HEADER TO FEATURE NAME
                    SORTED_CALC_DATA_HEADER_WIP[0][working_feature_idx] = raw_feature
                    # CHANGE DUM 1s TO CATEGORY NAME
                    SORTED_CALC_DATA_WIP[working_feature_idx] = \
                        np.where(SORTED_CALC_DATA_WIP[working_feature_idx]==1, category_label, SORTED_CALC_DATA_WIP[working_feature_idx])
                    working_feature_idx += 1
                # IF raw_feature IS IN SORTED_CALC_DATA_HEADER_WIP, THEN FIND OWNING COLUMN (SOMEWHERE TO THE LEFT) & UPDATE
                elif raw_feature in SORTED_CALC_DATA_HEADER_WIP[0]:
                    owning_column = np.argwhere(SORTED_CALC_DATA_HEADER_WIP[0]==raw_feature)[0][0]
                    SORTED_CALC_DATA_WIP[owning_column] = \
                        np.where(SORTED_CALC_DATA_WIP[working_feature_idx]==1, category_label, SORTED_CALC_DATA_WIP[owning_column])
                    SORTED_CALC_DATA_WIP = np.delete(SORTED_CALC_DATA_WIP, working_feature_idx, axis=0)
                    SORTED_CALC_DATA_HEADER_WIP = np.delete(SORTED_CALC_DATA_HEADER_WIP, working_feature_idx, axis=1)
                    # DONT INCREMENT working_feature_idx BECAUSE OF THE COLUMN DELETE!!!

            else:
                raise TypeError(f'UNKNOWN MODIFIED DATATYPE IN {self.this_module}.output_calc() WHEN CONSOLIDATING WORKING DATA FOR DISPLAY.')

            if working_feature_idx == len(SORTED_CALC_DATA_WIP): break

        # ADD OTHER COLUMNS TO SORTED_CALC_DATA_WIP
        SORTED_CALC_DATA_WIP = np.vstack((SORTED_CALC_REFERENCE_VECTORS_WIP[0],
                                     SORTED_SCORE_VECTOR,
                                     SORTED_CALC_TARGET_WIP,
                                     SORTED_CALC_DATA_WIP,
                                     SORTED_CALC_REFERENCE_VECTORS_WIP[1:])).transpose()  # TRANSPOSE BACK TO [ [] = ROWS ]

        # CREATE FEATURE HEADER FOR CONSOLIDATED_SORTED_CALC_DATA_WIP W/ SCORE COLUMN
        for header in reversed(['INDEX', 'SCORE', *SORTED_CALC_TARGET_HEADER_WIP[0]]):
            SORTED_CALC_DATA_HEADER_WIP = np.insert(SORTED_CALC_DATA_HEADER_WIP, 0, header, axis=1)

        SORTED_CALC_DATA_HEADER_WIP = np.hstack((SORTED_CALC_DATA_HEADER_WIP, SORTED_CALC_REFERENCE_VECTORS_HEADER_WIP[..., 1:]))

        self.CSUTM_DF = pd.DataFrame(data=SORTED_CALC_DATA_WIP, columns=SORTED_CALC_DATA_HEADER_WIP[0], dtype=object)

        del SORTED_SCORE_VECTOR, SORTED_CALC_DATA_WIP, SORTED_CALC_DATA_HEADER_WIP, SORTED_CALC_TARGET_WIP, \
            SORTED_CALC_TARGET_HEADER_WIP, SORTED_CALC_REFERENCE_VECTORS_WIP, SORTED_CALC_REFERENCE_VECTORS_HEADER_WIP

        # BUILDING OF RESULTS OBJECTS AND DISPLAY OBJECTS COMPLETE##########################################################
        print(f'\nCalculate {object_name} CASES complete.')
        st.show_end_time(f'Calculate {object_name} CASES\n')

        return self.CSUTM_DF, CALC_OUTPUT_VECTOR
    # END GENERIC FUNCTIONS ####################################################################################################################

    # TRAIN BUILD DEFINITIONS ################################################################################################
    def reset_train_data(self):     # reset train data to ORIGINAL DATA

        if vui.validate_user_str(f'\nReset TRAIN objects as full copy of original data? (y/n) > ', 'YN') == 'Y':
            self.TRAIN_SWNL = np.array(
                [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in self.SUPER_WORKING_NUMPY_LIST], dtype=object)

            self.train_has_been_run = False

            # 4-8-22 TO ACCOMODATE self.TRAIN_RESULTS in GMLR and MI, JUST pass IF ANY OTHER MODULE
            try: self.TRAIN_RESULTS = pd.DataFrame({})
            except: pass

            if vui.validate_user_str(f'\nReset DEV and TEST objects to empty? (y/n) > ', 'YN') == 'Y':
                self.DEV_SWNL = []
                self.TEST_SWNL = []
                self.CSUTM_DF = pd.DataFrame({})
            if vui.validate_user_str(f'\nReset DEV results object to empty? (y/n) > ', 'YN') == 'Y':
                self.DEV_ERROR = [[]]
    # END TRAIN BUILD DEFINITIONS ################################################################################################

    # DEV BUILD DEFINITIONS ##################################################################################################
    def base_dev_build_module(self):
        # return module with commands for building dev matrices for all ML packages

        fxn = inspect.stack()[0][3]

        if self.dev_build_select == 'D':  # 'use original data(d)'
            self.DEV_SWNL = np.array(
                [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in self.SUPER_WORKING_NUMPY_LIST], dtype=object)
            print(f'\nDEV MATRICES SUCCESSFULLY LOADED AS ENTIRE ORIGINAL DATA.\n')
            print(f'DEV DATA has {gs.get_shape("DEV_DATA", self.DEV_SWNL[0], self.data_run_orientation)[0]} examples.')

        elif self.dev_build_select == 'R':  # 'randomly draw a subset from training DATA(r)'
            # IF USER SENT TimeoutError BACK TO HERE, IT WAS BECAUSE INVERT OF XTX FOR MLR & GMLR FAILED AND CHOSE ABORT
            # MODIFICATIONS TO self.TRAIN_SWNL ARE IMPLICIT
            try:
                size = self.dev_or_test_draw_params()
                print(f'\nDEV MATRICES BEING CREATED BY RANDOM DRAW OF TRAINING DATA...\n')
                self.DEV_SWNL = self.random_dev_or_test_draw(size=size)
                print(f'\nDEV MATRICES SUCCESSFULLY LOADED BY RANDOM DRAW OF TRAINING DATA.\n')
                print(f'DEV DATA has {gs.get_shape("DEV_DATA", self.DEV_SWNL[0], self.data_run_orientation)[0]} examples.')
            except TimeoutError: pass
            except: raise

        elif self.dev_build_select == 'S':  #  'draw a sub-partition from training DATA(s)'
            # IF USER SENT TimeoutError BACK TO HERE, IT WAS BECAUSE INVERT OF XTX FOR MLR & GMLR FAILED AND CHOSE ABORT
            # MODIFICATIONS TO self.TRAIN_SWNL ARE IMPLICIT
            number_of_partitions = vui.validate_user_int(f'Enter number of partitions > ', min=2,
                                                 max=gs.get_shape("TRAIN_DATA", self.TRAIN_SWNL[0], self.data_run_orientation)
            )

            partition_number = vui.validate_user_int(f'Select partition number (1 to {number_of_partitions}) > ', min=1,
                                                         max=number_of_partitions)
            try:
                print(f'\nDEV MATRICES BEING CREATED FROM PARTITION OF TRAINING DATA...\n')
                self.DEV_SWNL = self.partition_dev_or_test_draw(number_of_partitions, partition_number-1)
                print(f'\nDEV MATRICES SUCCESSFULLY LOADED FROM PARTITION OF TRAINING DATA.\n')
                print(f'DEV DATA has {gs.get_shape("DEV_DATA", self.DEV_SWNL[0], self.data_run_orientation)[0]} examples.')
            except TimeoutError: pass
            except: raise

        elif self.dev_build_select == 'U':    #  'select block from train data(u)'
            train_data_rows = gs.get_shape("TRAIN_DATA", self.TRAIN_SWNL[0], self.data_run_orientation)[0]
            while True:
                start_row = vui.validate_user_int(f'Select start row (zero-indexed, {train_data_rows} rows) > ', min=0, max=train_data_rows-1)
                end_row = vui.validate_user_int(f'Select end row (inclusive, zero-indexed, {train_data_rows} rows)) > ', min=start_row, max=train_data_rows)
                if vui.validate_user_str(f'User selected row range {start_row} to and including {end_row}, accept? (y/n) > ', 'YN') == 'Y': break
            _ROWS = list(range(start_row, end_row+1)); del train_data_rows, start_row, end_row

            self.DEV_SWNL = [None for _ in range(3)]

            NAMES = ('DATA', 'TARGET', 'REFVECS')
            ORIENTS = (self.data_run_orientation, self.target_run_orientation, self.refvecs_run_orientation)
            for obj_idx, (_name, _orient) in enumerate(zip(NAMES, ORIENTS)):
                TrainDevClass = mlo.MLObject(deepcopy(self.TRAIN_SWNL[obj_idx]) if isinstance(self.TRAIN_SWNL[obj_idx], dict) else self.TRAIN_SWNL[obj_idx].copy(),
                                             _orient,
                                             name=_name,
                                             return_orientation='AS_GIVEN',
                                             return_format='AS_GIVEN',
                                             bypass_validation=self.bypass_validation,
                                             calling_module=self.this_module,
                                             calling_fxn=fxn
                )
    
                self.DEV_SWNL[obj_idx] = TrainDevClass.return_rows(_ROWS, return_orientation='AS_GIVEN', return_format='AS_GIVEN')
                self.TRAIN_SWNL[obj_idx] = TrainDevClass.delete_rows(_ROWS)

            del TrainDevClass, NAMES, ORIENTS


        elif self.dev_build_select == 'V':    #  'draw based on categories in a column(v)'
            # IF USER SENT TimeoutError BACK TO HERE, IT WAS BECAUSE INVERT OF XTX FOR MLR & GMLR FAILED AND CHOSE ABORT
            # MODIFICATIONS TO self.TRAIN_SWNL ARE IMPLICIT
            try:
                print(f'\nDEV MATRICES BEING CREATED BY DRAW FROM TRAINING DATA BY CATEGORY...\n')
                self.DEV_SWNL = self.category_dev_or_test_draw()
                print(f'DEV DATA has {gs.get_shape("DEV_DATA", self.DEV_SWNL[0], self.data_run_orientation)[0]} examples.')
                print(f'\nDEV MATRICES SUCCESSFULLY LOADED BY CATEGORY OF TRAINING DATA.\n')
            except TimeoutError: pass
            except: raise

        elif self.dev_build_select == 'F':  # 'load from file(f)'
            # self.DEV_SWNL = '2-20-22 BEAR FIX some_fxn_that_reads_from_file()'
            print(f'\n*** DEV MATRICES FROM FILE NOT LOADED.  CODE NOT FINISHED YET. :( ***\n')
            # print(f'\nDEV MATRICES SUCCESSFULLY LOADED FROM FILE.\n')
            # print(f'DEV DATA has {gs.get_shape("DATA_DEV", self.DEV_SWNL[0], self.data_run_orientation)[0]} examples.')

        elif self.dev_build_select == 'T':  # 'use dev matrices currently loaded(t)'
            pass
            print(f'\nCURRENT DEV MATRICES BEING RETAINED.\n')
            print(f'DEV DATA has {gs.get_shape("DEV_DATA", self.DEV_SWNL[0], self.data_run_orientation)[0]} examples.')

        elif self.dev_build_select == 'B':  # 'load standard config or manual build(b)'
            # from ML_PACKAGE.TEST_CONFIG_RUN import dev_matrix_config_build as dmcb
            # self.DEV_SWNL = \
            #     'BEAR FIX dmcb.dev_matrix_config_build(dev_matrix_config, DATA_DF, BIG_MATRIX, BASE_BIG_MATRIX, \
            #     TEST_MATRIX, intercept, CONTEXT, standard_config)'
            print(
                f'\n*** DEV MATRICES FROM STANDARD CONFIG / MANUAL BUILD NOT LOADED.  CODE NOT FINISHED YET. :( ***\n')
            # print(f'\nDEV MATRICES SUCCESSFULLY LOADED FROM STANDARD CONFIG / MANUAL BUILD.\n')
            # print(f'DEV DATA has {gs.get_shape("DEV_DATA", self.DEV_SWNL[0], self.data_run_orientation)[0]} examples.')

        elif self.dev_build_select == 'N':  # 'None'
            self.DEV_SWNL = []
            # self.DEV_ERROR = [[]]
            print(f'\n*** DEV OBJECTS AND PREVIOUS DEV RESULTS SUCCESSFULLY ERASED. ***\n')

        #    'accept & dev objects build(a)'  IS EXTERNAL TO THIS BLOCK


    def sub_dev_build_module(self):
        # return module with package-specific test matrix build commands, overwritten in child
        # overwritten in child
        pass


    def sub_dev_build_cmds(self):
        # return list with package-specific test matrix build commands
        return {}  # SPECIFIED IN CHILDREN   CANT USE 'RSDFTUVBNA'


    def dev_build(self):

        print(f'\n*** BUILD DEV OBJECTS ***')
        print(f'*** IF PLANNING TO RUN K-FOLD CROSS VALIDATION OR TRUNCATED K-FOLD DEV LOOP DO NOT NEED TO BUILD DEV OBJECTS HERE ***')
        print(f'*** DID YOU NEED TO RESET TRAIN OR TEST DATA BEFORE BUILDING DEV OBJECTS? ***')

        while True:
            print()
            self.dev_build_select = dmp.DictMenuPrint(self.BASE_DEV_BUILD_CMDS | self.sub_dev_build_cmds(),
                                                                disp_len=140).select(f'Select dev data build option')

            if self.dev_build_select == 'A':
                break

            self.base_dev_build_module()
            self.sub_dev_build_module()

    # END DEV BUILD DEFINITIONS ##################################################################################################

    # TEST BUILD DEFINITIONS ##################################################################################################
    def base_test_build_module(self):
        # return module with commands for building test matrices for all ML packages

        fxn = inspect.stack()[0][3]

        if self.test_build_select == 'D':    # 'use original DATA(d)'
            self.TEST_SWNL = np.array(
                [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in self.SUPER_WORKING_NUMPY_LIST], dtype=object)
            print(f'\nORIGINAL DATA SUCCESSFULLY LOADED AS TEST MATRICES.\n')
            print(f'TEST DATA has {gs.get_shape("TEST_DATA", self.TEST_SWNL[0], self.data_run_orientation)[0]} examples.')

        elif self.test_build_select == 'R':    # 'random draw from training DATA(r)'
            # IF USER SENT TimeoutError BACK TO HERE, IT WAS BECAUSE INVERT OF XTX FOR MLR & GMLR FAILED AND CHOSE ABORT
            # MODIFICATIONS TO self.TRAIN_SWNL ARE IMPLICIT
            try:
                size = self.dev_or_test_draw_params()
                print(f'\nTEST MATRICES BEING CREATED BY RANDOM DRAW OF TRAINING DATA...\n')
                self.TEST_SWNL = self.random_dev_or_test_draw(size=size)
                print(f'\nTEST MATRICES SUCCESSFULLY LOADED BY RANDOM DRAW OF TRAINING DATA.\n')
                print(f'TEST DATA has {gs.get_shape("TEST_DATA", self.TEST_SWNL[0], self.data_run_orientation)[0]} examples.')
            except TimeoutError: pass
            except: raise

        elif self.test_build_select == 'S':  # 'draw a sub-partition from training DATA(s)'
            # IF USER SENT TimeoutError BACK TO HERE, IT WAS BECAUSE INVERT OF XTX FOR MLR & GMLR FAILED AND CHOSE ABORT
            # MODIFICATIONS TO self.TRAIN_SWNL ARE IMPLICIT
            number_of_partitions = vui.validate_user_int(f'Enter number of partitions > ', min=2,
                                        max=gs.get_shape("TRAIN_DATA", self.TRAIN_SWNL[0], self.data_run_orientation)[0]
            )
            partition_number = vui.validate_user_int(f'Select partition number (1 to {number_of_partitions}) > ', min=1,
                                                         max=number_of_partitions)
            try:
                print(f'\nTEST MATRICES BEING CREATED FROM PARTITION OF TRAINING DATA...\n')
                self.TEST_SWNL = self.partition_dev_or_test_draw(number_of_partitions, partition_number-1)
                print(f'\nTEST MATRICES SUCCESSFULLY LOADED FROM PARTITION OF TRAINING DATA.\n')
                print(f'TEST DATA has {gs.get_shape("TEST_DATA", self.TEST_SWNL[0], self.data_run_orientation)[0]} examples.')
            except TimeoutError: pass
            except: raise

        elif self.test_build_select == 'U':    #'User select block from train data(u)'
            train_data_rows = gs.get_shape("TRAIN_DATA", self.TRAIN_SWNL[0], self.data_run_orientation)[0]
            while True:
                start_row = vui.validate_user_int(f'Select start row (zero-indexed, {train_data_rows} rows) > ', min=0, max=train_data_rows - 1)
                end_row = vui.validate_user_int(f'Select end row (inclusive, zero-indexed, {train_data_rows} rows)) > ', min=start_row, max=train_data_rows)
                if vui.validate_user_str( f'User selected row range {start_row} to and including {end_row}, accept? (y/n) > ', 'YN') == 'Y': break
            _ROWS = list(range(start_row, end_row + 1)); del train_data_rows, start_row, end_row

            self.TEST_SWNL = [None for _ in range(3)]

            NAMES = ('DATA', 'TARGET', 'REFVECS')
            ORIENTS = (self.data_run_orientation, self.target_run_orientation, self.refvecs_run_orientation)
            for obj_idx, (_name, _orient) in enumerate(zip(NAMES, ORIENTS)):
                TrainTestClass = mlo.MLObject(
                                                deepcopy(self.TRAIN_SWNL[obj_idx]) if isinstance(self.TRAIN_SWNL[obj_idx], dict) else
                                                self.TRAIN_SWNL[obj_idx].copy(),
                                                _orient,
                                                name=_name,
                                                return_orientation='AS_GIVEN',
                                                return_format='AS_GIVEN',
                                                bypass_validation=self.bypass_validation,
                                                calling_module=self.this_module,
                                                calling_fxn=fxn
                )

                self.TEST_SWNL[obj_idx] = TrainTestClass.return_rows(_ROWS, return_orientation='AS_GIVEN',
                                                                   return_format='AS_GIVEN')
                self.TRAIN_SWNL[obj_idx] = TrainTestClass.delete_rows(_ROWS)

            del TrainTestClass, NAMES, ORIENTS

        elif self.test_build_select == 'V':    #  'draw based on categories in a column(v)'
            # IF USER SENT TimeoutError BACK TO HERE, IT WAS BECAUSE INVERT OF XTX FOR MLR & GMLR FAILED AND CHOSE ABORT
            # MODIFICATIONS TO self.TRAIN_SWNL ARE IMPLICIT
            try:
                print(f'\nTEST MATRICES BEING CREATED BY DRAW FROM TRAINING DATA BY CATEGORY...\n')
                self.TEST_SWNL = self.category_dev_or_test_draw()
                print(f'TEST DATA has {gs.get_shape("TEST_DATA", self.TEST_SWNL[0], self.data_run_orientation)[0]} examples.')
                print(f'\nTEST MATRICES SUCCESSFULLY LOADED FROM TRAINING DATA BY CATEGORY.\n')
            except TimeoutError: pass
            except: raise

        elif self.test_build_select == 'P':    # 'generate all permutations(p)'
            if 'FLOAT' in self.WORKING_SUPOBJS[msod.QUICK_POSN_DICT()['MODIFIEDDATATYPES']]:
                print(f'\n*** DATA CONTAINS AT LEAST ONE NON-CATEGORICAL DATA SERIES.  CANNOT GENERATE PERMUTATION MATRIX. ***\n')
            else:
                # self.TEST_SWNL = '2-20-22 BEAR FIX some_fxn_that_loads_all_permutations()'
                # WONT BE ABLE TO INCLUDE REFERENCE VECTORS HERE
                print(f'\n*** ALL PERMUTATIONS NOT LOADED.  CODE NOT FINISHED YET. ***\n')
                # print(f'\nALL PERMUTATIONS TEST DATA SUCCESSFULLY LOADED.\n')
                # print(f'TEST DATA has {gs.get_shape("TEST_DATA", self.TEST_SWNL[0], self.data_run_orientation)[0]} examples.')

        elif self.test_build_select == 'F':    # 'load from file(f)'
            # self.TEST_SWNL = '2-20-22 BEAR FIX some_fxn_that_reads_from_file()'
            print(f'\n*** TEST MATRICES FROM FILE NOT LOADED.  CODE NOT FINISHED YET. :( ***\n')
            # print(f'\nTEST MATRICES SUCCESSFULLY LOADED FROM FILE.\n')
            # print(f'TEST DATA has {gs.get_shape("TEST_DATA", self.TEST_SWNL[0], self.data_run_orientation)[0]} examples.')

        elif self.test_build_select == 'T':    # 'use test matrices currently loaded(t)'
            pass
            print(f'\nCURRENT TEST MATRICES BEING RETAINED.\n')
            print(f'TEST DATA has {gs.get_shape("TEST_DATA", self.TEST_SWNL[0], self.data_run_orientation)[0]} examples.')

        elif self.test_build_select == 'B':     # 'load standard config or manual build(b)'
            # BEAR FIX
            # self.TEST_SWNL = tmcb.test_matrix_config_build(test_matrix_config, DATA_DF, BIG_MATRIX, BASE_BIG_MATRIX, \
            #     TEST_MATRIX, intercept, CONTEXT, standard_config)
            print(f'\n*** TEST MATRICES FROM STANDARD CONFIG / MANUAL BUILD NOT LOADED.  CODE NOT FINISHED YET. ***\n')
            # print(f'\nTEST MATRICES SUCCESSFULLY LOADED FROM STANDARD CONFIG / MANUAL BUILD.\n')
            # print(f'TEST DATA has {gs.get_shape("TEST_DATA", self.TEST_SWNL[0], self.data_run_orientation)[0]} examples.')

        elif self.test_build_select == 'O':    # 'overwrite TEST_SWNL(o)'
            print(f'\n*** 4/19/23 SWNL NO LONGER HAS SLOTS FOR TEST OBJECTS.  FIX THIS, OR REMOVE THIS OPTION.')
            # if vui.validate_user_str(
                # f'\n*** OVERWRITING SWNL TEST_MATRIX SPOTS WITH WIP TEST_MATRIX AND HEADER.  CANNOT BE UNDONE.  PROCEED? (y/n) ***', 'YN') == 'Y':
                # self.SUPER_WORKING_NUMPY_LIST[6] = deepcopy(self.TEST_SWNL[0])
                # self.SUPER_WORKING_NUMPY_LIST[7] = deepcopy(self.TEST_SWNL[1])
                # BIG MISTAKE, DIDNT PUT A SLOT IN IN SRNL TO INGEST TEST_TARGET & HEADER :(
                # print(f'\nTEST MATRIX & HEADER SUCCESSFULLY SAVED INTO WORKING DATA OBJECT.\n')

        elif self.test_build_select == 'N':  # 'None'
            self.TEST_SWNL = []
            self.CSUTM_DF = pd.DataFrame({})
            print(f'\n*** TEST OBJECTS AND PREVIOUS TEST RESULTS SUCCESSFULLY ERASED. ***\n')

        #    'accept & exit test objects build(a)'  IS EXTERNAL TO THIS BLOCK


    def sub_test_build_module(self):
        # return module with package-specific test matrix build commands, overwritten in child
        # overwritten in child
        pass


    def sub_test_build_cmds(self):
        # return list with package-specific test matrix build commands
        return {}    # SPECIFIED IN CHILDREN   CANT USE 'DRSPFTBONA'


    def test_build(self):

        print(f'\n*** BUILD TEST OBJECTS ***')
        print(f'*** IF PLANNING TO RUN K-FOLD CROSS VALIDATION, DO NOT NEED TO BUILD TEST OBJECTS HERE ***')
        print(f'*** DID YOU NEED TO RESET TRAIN OR DEV DATA BEFORE BUILDING TEST OBJECTS? ***')

        while True:
            print()
            self.test_build_select = dmp.DictMenuPrint(self.BASE_TEST_BUILD_CMDS | self.sub_test_build_cmds(),
                                                                disp_len=140).select('Select TEST MATRIX build option')

            if self.test_build_select == 'A':
                break

            self.base_test_build_module()
            self.sub_test_build_module()

    # END TEST BUILD DEFINITIONS ##################################################################################################

    # TRAIN CALC DEFINITIONS ###################################################################################################
    def core_training_code(self):
        # return unique core training algorithm for particular ML package
        pass

    def kfold_core_training_code(self):
        # return unique kfold core training algorithm for particular ML package
        pass

    def kfold_config(self):
        ########################################################################################################################
        # KFOLD CONFIG #########################################################################################################
        while True:
            # kfold config algorithm, configures number of regularization trials, factors, and early stopping
            # DOES NOT CONFIGURE PARTITIONS
            while True:
                num_rgzn_trials = vui.validate_user_int(f'\nEnter number of regularization trials > ', min=1)
                if vui.validate_user_str(f'User entered {num_rgzn_trials} regularization trials. Accept? (y/n) > ', 'YN') == 'Y':
                    break

            while True:
                RGLZTN_FACTORS = mnslf.manual_num_seq_list_fill('regularization', [], num_rgzn_trials, min=0,
                                                                max=float('inf'))
                print(f'USER SELECTED REGULARIZATION FACTORS: {RGLZTN_FACTORS}')
                if vui.validate_user_str(f'Accept? (y/n) > ', 'YN') == 'Y': break

            while True:
                if vui.validate_user_str(f'\nUse early-stopping? (y/n) > ', 'YN') == 'Y':
                    early_stop_interval = vui.validate_user_int(f'Enter early-stop dev validation interval > ', min=1)
                else: early_stop_interval = 1e12
                print(f'USER SELECTED{" NO" if early_stop_interval == 1e-12 else ""} EARLY STOPPING WITH INTERVAL OF {early_stop_interval: ,.0f}')
                if vui.validate_user_str(f'Accept? (y/n) > ', 'YN') == 'Y': break

            print()
            print(f'num_rgzn_trials = {num_rgzn_trials}')
            print(f'RGLZTN_FACTORS = {RGLZTN_FACTORS}')
            print(f'early_stop_interval = {early_stop_interval: ,.0f}')
            print()
            if vui.validate_user_str(f'Accept k-fold config? (y/n) > ', 'YN') == 'Y':
                return num_rgzn_trials, RGLZTN_FACTORS, early_stop_interval

        # END KFOLD CONFIG #########################################################################################################
        ############################################################################################################################

    # END TRAIN CALC DEFINITIONS ###################################################################################################



    # DEV CALC DEFINITIONS ##################################################################################################

    def rglztn_partition_iterator(self, number_of_partitions, draw_method='PARTITION'):
        # only used to save space in base_dev_calc_module()

        # CREATE COPIES OF TRAIN_SWNL & DEV_SWNL TO RETURN TO ORIGINAL STATE AFTER ITERATOR
        ORIG_RGLZTN_FCTR = deepcopy(self.rglztn_fctr)
        ORIG_TRAIN_SWNL = [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in  self.TRAIN_SWNL]
        ORIG_DEV_SWNL = [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in self.DEV_SWNL]

        while True:  # ALLOWS FOR SEND BACK UP TO CONFIG IF MLR OR GMLR TRAIN_XTX FAILS INV

            num_rgzn_trials, self.RGLZTN_FACTORS, self.early_stop_interval = self.kfold_config()

            # WANT TO ACCRETE ERROR FOR THE VARIOUS rglztn_fctrs AND partitions AS
            # [[error for various partitions in DATA PARTITIONS] for rglztn_fctrs in rglztn_TRIAL]
            self.DEV_ERROR = np.zeros((num_rgzn_trials, number_of_partitions), dtype=np.float64)

            _excepted = False
            for partition in range(number_of_partitions):

                # SET self.rglztn_fctr TO FIRST IN self.RGLZTN_FACTORS TO ENABLE TEST FOR XTX_INV OF TRAIN PARTITION IN
                # GMLRRun AND MLRegressionRun. THIS DOESNT MATTER FOR ANY OTHER ML PACKAGE
                self.rglztn_fctr = self.RGLZTN_FACTORS[0]

                # IF number_of_partitions == 1, USER IS USING STANDARD DEV LOOP (NOT K-FOLD), AND IF draw_method == PARTITION, ALLOW
                # USER TO SELECT WHICH PARTITION TO USE
                if number_of_partitions == 1 and draw_method == 'PARTITION':
                    partition = vui.validate_user_int(f'Select number of partition to use (of {number_of_partitions}) > ',
                                                      min=1, max=number_of_partitions) - 1  # <-- MAKE IT GO TO ZERO-BASED

                if draw_method == 'PARTITION':
                    # IF USER SENT TimeoutError BACK TO HERE, IT WAS BECAUSE INVERT OF XTX FOR MLR & GMLR FAILED AND CHOSE ABORT
                    # MODIFICATIONS TO self.TRAIN_SWNL ARE IMPLICIT
                    try:
                        print(f'\nDEV MATRICES BEING CREATED FROM PARTITION OF TRAINING DATA...\n')
                        self.DEV_SWNL = self.partition_dev_or_test_draw(number_of_partitions, partition)
                        print(f'DEV DATA has {gs.get_shape("DEV_DATA", self.DEV_SWNL[0], self.data_run_orientation)[0]} examples.')
                        print(f'\nDEV MATRICES SUCCESSFULLY LOADED FROM PARTITION OF TRAINING DATA.\n')
                    except TimeoutError: _excepted = True
                    except: raise

                elif draw_method == 'RANDOM':
                    # IF USER SENT TimeoutError BACK TO HERE, IT WAS BECAUSE INVERT OF XTX FOR MLR & GMLR FAILED AND CHOSE ABORT
                    # MODIFICATIONS TO self.TRAIN_SWNL ARE IMPLICIT
                    try:
                        print(f'\nDEV MATRICES BEING CREATED BY RANDOM DRAW OF TRAINING DATA...\n')
                        self.DEV_SWNL = self.random_dev_or_test_draw(gs.get_shape("DATA", self.SUPER_WORKING_NUMPY_LIST[0],
                                                                            self.data_run_orientation)[0] // number_of_partitions)
                        print(f'\nDEV MATRICES SUCCESSFULLY LOADED BY RANDOM DRAW OF TRAINING DATA.\n')
                        print(f'DEV DATA has {gs.get_shape("DEV_DATA", self.DEV_SWNL[0], self.data_run_orientation)[0]} examples.')
                    except TimeoutError: _excepted = True
                    except: raise

                else:
                    raise ValueError(f'\nINVALID draw_method in {self.this_module}.rglztn_partition_iterator(). ***\n')

                if _excepted: break

                try: self.rebuild_kernel = True  # ONLY FOR SVMRun & MLR (NOT GMLR) ... FOR MANAGING K-FOLD KERNEL REBUILDS
                except: pass

                for rglztn_trial in range(num_rgzn_trials):
                    self.rglztn_fctr = self.RGLZTN_FACTORS[rglztn_trial]

                    #####################################################################################################
                    print(f'\nPARTITION TRIAL {partition+1} OF {number_of_partitions}')
                    print(f'\nREGULARIZATION TRIAL {rglztn_trial + 1} OF {num_rgzn_trials}')
                    print(f'REGULARIZATION FACTOR: {self.RGLZTN_FACTORS[rglztn_trial]}')
                    _ = datetime.datetime.now().ctime()
                    print(f'\nCURRENT FOLD START DATE / TIME: {_[:11] + _[-4:] + _[10:-5]}')

                    self.kfold_core_training_code()

                    self.DEV_OUTPUT_VECTOR = self.dev_data_calc_algorithm()
                    self.dev_error = self.dev_data_error_algorithm()

                    self.DEV_ERROR[rglztn_trial][partition] = self.dev_error
                    #####################################################################################################

                    try: self.rebuild_kernel = False           # ONLY FOR SVMRun & MLR (NOT GMLR)... FOR MANAGING K-FOLD KERNEL REBUILDS
                    except: pass

                # RESET TRAIN & DEV OBJECTS TO ORIGINAL STATE FOR NEXT PARTITIONING OR EXIT
                self.TRAIN_SWNL = [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in ORIG_TRAIN_SWNL]
                self.DEV_SWNL = [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in ORIG_DEV_SWNL]

            if _excepted:
                self.TRAIN_SWNL = [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in ORIG_TRAIN_SWNL]
                self.DEV_SWNL = [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in ORIG_DEV_SWNL]
                __ = vui.validate_user_str(f'\nAbort during kfold, abort(a) try again(c) > ', 'AC')
                if __ == 'A': pass
                elif __ == 'C': continue

            del _excepted

            # RESTORE ORIGINAL STATE
            self.rglztn_fctr = deepcopy(ORIG_RGLZTN_FCTR)
            self.early_stop_interval = 1e12

            # 10/26/22 PRINT AT END OF RUN FOR MEMORY JOGGING AFTER OVERNIGHT RUNS
            print()
            self.hyperparameter_display_module()
            print()

            print(f'\n***** K-FOLD CROSS VALIDATION COMPLETE. ***** \n')
            for _ in range(3): wls.winlinsound(888, 500); time.sleep(1)

            if vui.validate_user_str(f'\nRun k-fold again(a) or exit(e) > ', 'AE') == 'E': break

        del ORIG_TRAIN_SWNL, ORIG_DEV_SWNL, ORIG_RGLZTN_FCTR


    def base_dev_calc_module(self):

        if self.dev_run_select == 'D':     # 'run truncated k-fold dev loop on one partition(d)'
            while True:
                draw_method = {'R':'RANDOM', 'P':'PARTITION', 'A':'A'}[
                    vui.validate_user_str(f'Randomly draw dev objects(r), draw a partition(p), or abort(a) > ', 'RPA')]
                if draw_method == 'A': break
                else: self.rglztn_partition_iterator(1, draw_method=draw_method)

                break

        elif self.dev_run_select == 'S':       # 'run standard dev loop on current dev objects(s)'

            while True:  # JUST TO ENABLE BREAK IF rglztn_type IS NOT SET

                if isinstance(self.rglztn_type, str): self.rglztn_type = self.rglztn_type.upper()

                if self.rglztn_type in ['NONE', None]:
                    print(f'\n*** REGULARIZATION TYPE MUST BE SET BEFORE RUNNING DEV LOOP ***\n')
                    break

                number_of_partitions = 1
                partition = 0

                # CREATE COPIES OF TRAIN_SWNL & DEV_SWNL TO RETURN TO ORIGINAL STATE AFTER ITERATOR
                ORIG_RGLZTN_FCTR = deepcopy(self.rglztn_fctr)

                num_rgzn_trials, self.RGLZTN_FACTORS, self.early_stop_interval = self.kfold_config()

                # WANT TO ACCRETE ERROR FOR THE VARIOUS rglztn_fctrs AND partitions AS
                # [[error for various partitions in DATA PARTITIONS] for rglztn_fctrs in rglztn_TRIAL]
                self.DEV_ERROR = np.zeros((num_rgzn_trials, number_of_partitions), dtype=np.float64)

                try: self.rebuild_kernel = True  # ONLY FOR SVMRun & MLR (NOT GMLR) ... FOR MANAGING K-FOLD KERNEL REBUILDS
                except: pass

                for rglztn_trial in range(num_rgzn_trials):
                    self.rglztn_fctr = self.RGLZTN_FACTORS[rglztn_trial]

                    print(f'REGULARIZATION TRIAL {rglztn_trial + 1} OF {num_rgzn_trials}')
                    print(f'REGULARIZATION FACTOR = {self.RGLZTN_FACTORS[rglztn_trial]}')
                    _ = datetime.datetime.now().ctime()
                    print(f'\nCURRENT FOLD START DATE / TIME: {_[:11] + _[-4:] + _[10:-5]}')

                    self.kfold_core_training_code()

                    self.DEV_OUTPUT_VECTOR = self.dev_data_calc_algorithm()
                    self.dev_error = self.dev_data_error_algorithm()

                    self.DEV_ERROR[rglztn_trial][partition] = self.dev_error

                    try: self.rebuild_kernel = False  # ONLY FOR SVMRun & MLR (NOT GMLR)... FOR MANAGING K-FOLD KERNEL REBUILDS
                    except: pass

                # RESTORE ORIGINAL RGLZTN FACTOR
                self.rglztn_fctr = deepcopy(ORIG_RGLZTN_FCTR)

                self.early_stop_interval = 1e12

                # 10/26/22 PRINT AT END OF RUN FOR MEMORY JOGGING AFTER OVERNIGHT RUNS
                print()
                self.hyperparameter_display_module()
                print()

                break

        elif self.dev_run_select == 'P':     # 'print dev results to screen(p)'
            self.print_dev_results_module()



        # 'accept & exit dev calc(a)'   IS EXTERNAL TO THIS LOOP


    def sub_dev_calc_module(self):
        # return module with package-specific commands to run current dev matrix
        # overwritten in child
        pass


    def sub_dev_calc_cmds(self):
        # return list with package - specific dev calc commands
        return {}    # SPECIFIED IN CHILDREN   CANT USE 'DSA'

    # END DEV CALC DEFINITIONS ##################################################################################################

    # TEST CALC DEFINITIONS ##################################################################################################
    def base_test_calc_module(self):

        if self.test_calc_select == 'N':    #'run one-off pass of test data(n)'
            self.CSUTM_DF, self.TEST_OUTPUT_VECTOR = self.output_calc(self.test_data_calc_algorithm(), self.TEST_SWNL,
                                                                      'TEST DATA')
            print(f'\n*** ONE-OFF PASS OF TEST DATA SUCCESSFULLY CALCULATED ***\n')

        elif self.test_calc_select == 'P':  # 'run test data perturber(p)'

            #####################################################################################################################
            #####################################################################################################################

            DATA_HEADER = self.WORKING_SUPOBJS[0][msod.QUICK_POSN_DICT()["HEADER"]].reshape((1,-1))

            PerturberClass = tp.TestPerturber(
                                                self.TEST_SWNL,
                                                self.WORKING_SUPOBJS[0],
                                                self.data_run_orientation,
                                                self.target_run_orientation,
                                                wb=self.wb,                 # BEAR MAY HAVE TO GIVE IT ITS OWN WB
                                                bypass_validation=self.bypass_validation
            )

            PerturberClass.menu()

            test_data_cols = gs.get_shape('TEST_DATA', self.TEST_SWNL[0], self.data_run_orientation)[1]

            # LOOKS LIKE ITS OK TO NOT UPDATE THIS SCOPES TEST_SWNL WITH PerturberClass TEST_SWNL BEFORE CALC OF
            # original_baseline_cost BECAUSE ON FIRST ENTRY PerturberClass TEST_SWNL SHOULD BE FULL TEST_SWNL (W/O PERTURB) AND
            # THIS SCOPES TEST_SWNL SHOULD ALSO BE FULL TEST_SWNL.... SEE A FEW LINES DOWN
            PerturberClass.original_baseline_cost = self.test_data_error_algorithm()

            while True:  #########################################################################################################

                # CALCULATE BASELINE COST
                # BUT... UPON LOOP BACK UP TO THE TOP ON SUBSEQUENT PASSES, NEED TO SET THIS SCOPES SWNL TO PerturberClass TEST_SWNL
                # TO CAPTURE ANY LOCKS, ETC THAT MAY HAPPENED IN menu() BY SETTING THIS SCOPES TEST_SWNL TO PClass TEST_SWNL
                self.TEST_SWNL[0] = PerturberClass.WipTestDataClass.OBJECT
                PerturberClass.baseline_cost = self.test_data_error_algorithm()

                for col_idx in range(test_data_cols):

                    print(f'\nRunning column {col_idx + 1} of {test_data_cols}...')

                    if col_idx in PerturberClass.COLUMNS_WITH_LOCKED_PERTURBATION:
                        PerturberClass.RESULTS_TABLE.loc[DATA_HEADER[0][col_idx], ('COST', '% CHANGE')] = ['-', '-']
                        continue

                    self.TEST_SWNL[0] = PerturberClass.perturber(col_idx)

                    # OUTPUT_VECTOR IS CALCULATED WITHIN test_data_error_algorithm
                    wip_cost = self.test_data_error_algorithm()

                    PerturberClass.update_results_table(col_idx, wip_cost)

                menu_select = PerturberClass.menu()

                if menu_select=='A': break
                elif menu_select=='T': continue

            # END while ###########################################################################################################

            self.PERTURBER_RESULTS = PerturberClass.RESULTS_TABLE

            self.TEST_SWNL[0] = PerturberClass.TEST_SWNL_BACKUP[0]

            del DATA_HEADER, PerturberClass, test_data_cols

            print(f'\n*** TEST PERTURBER COMPLETED SUCCESSFULLY ***\n')

            #####################################################################################################################
            #####################################################################################################################


        # 'accept & exit test calc(a)' IS EXTERNAL TO THIS LOOP


    def sub_test_calc_module(self):
        # return module with package-specific commands to run current test matrix
        # overwritten in child
        pass


    def sub_test_calc_cmds(self):
        # return list with package - specific test calc commands
        return {}    # SPECIFIED IN CHILDREN   CANT USE 'NA'

    # END TEST CALC DEFINITIONS ##################################################################################################


    # TRAIN DATA DISPLAY ##############################################################################################################

    def train_summary_statistics_module(self):
        # returns module for printing summary statistics of train data
        # OVERWRITTEN IN CHILD
        pass


    def print_train_results_module(self):
        # returns module for printing train results for particular ML package
        # OVERWRITTEN IN CHILD
        # REMEMBER!!! THAT TRAIN RESULTS ARE THE OUTPUT OF TRAINING, EG PARAMETERS IN NN MATRICES.  TEST RESULTS ARE THE
        # PREDICTIONS RETURNED WHEN RUNNING TEST DATA INTO THE TRAINED ALGORITHM!
        pass


    def train_filedump_module(self):
        # returns module for filedump of train results for particular ML package
        # overwritten in child
        pass

    # END TRAIN DATA DISPLAY ##############################################################################################################

    # DEV DATA DISPLAY ##############################A##############################A##############################A##############
    def dev_summary_statistics_module(self):
        # returns module for printing summary statistics of dev data
        # OVERWRITTEN IN CHILD
        pass


    def print_dev_results_module(self):
        if self.DEV_ERROR == [[]]: print(f'\nDEV CYCLE HAS NOT BEEN PERFORMED FOR THE CURRENT TRAIN / DEV / TEST SPLIT\n')
        else: krp.kfold_results_print(self.DEV_ERROR, self.RGLZTN_FACTORS)


                
        


    def dev_filedump_module(self):
        # returns module for filedump of dev results for all ML packages
        # In child, put package-specific summary dump module here, then put general dev dump module
        self.wb = gdrd.general_dev_results_dump(self.wb, self.DEV_ERROR, self.RGLZTN_FACTORS)


    # END DEV DATA DISPLAY ##############################A##############################A##############################A##############

    # CALC DATA DISPLAY ##############################A##############################A##############################A##############
    def test_summary_statistics_module(self):
        # returns module for printing summary statistics of test data
        # OVERWRITTEN IN CHILD
        pass


    def print_test_results_module(self):
        # returns module for printing test results for particular ML package
        # THINKING THIS SHOULD WORK FOR BOTH DEV & CALC, SINCE BOTH PRODUCE CSUTM_DF
        # REMEMBER!!! THAT TRAIN RESULTS ARE THE OUTPUT OF TRAINING, EG PARAMETERS IN NN MATRICES.  TEST RESULTS ARE THE
        # PREDICTIONS RETURNED WHEN RUNNING TEST DATA INTO THE TRAINED ALGORITHM!

        print(f'\nCONTEXT: \n')
        if len(self.WORKING_CONTEXT) == 0:
            print(f'None.')
        else:
            [print(_) for _ in self.WORKING_CONTEXT]

        print()

        print(f'\nFILTERING: ')

        # SCAN self.WORKING_FILTERING FOR ANY INPUT, IF YES, STORE obj_idx & col_idx
        filt_idx = msod.QUICK_POSN_DICT()["FILTERING"]
        FILTER_OBJECTS = (self.WORKING_DATA_SUPOBJS[filt_idx], self.WORKING_TARGET_SUPOBJS[filt_idx], self.WORKING_REFVECS_SUPOBJ[filt_idx])
        del filt_idx
        FILTERED_OBJS_AND_COLS = np.empty((2,0), dtype=np.int32)
        for obj_idx, OBJ_FILTERING in enumerate(FILTER_OBJECTS):
            for col_idx, COL_FILTERING in enumerate(OBJ_FILTERING):
                if len(COL_FILTERING) == 0: continue
                else: FILTERED_OBJS_AND_COLS = np.hstack((FILTERED_OBJS_AND_COLS, [[obj_idx], [col_idx]]))

        if len(FILTERED_OBJS_AND_COLS[0]) == 0:
            print(f'None.')
        else:
            OBJ_NAME_BEEN_PRINTED = []
            for obj_idx, col_idx in zip(*FILTERED_OBJS_AND_COLS):
                if obj_idx not in OBJ_NAME_BEEN_PRINTED:
                    OBJ_NAME_BEEN_PRINTED.append(obj_idx)
                    print(f'\n{self.SUPER_NUMPY_DICT[obj_idx]}:')
                else: print('')
                print(f'    {self.SUPER_WORKING_NUMPY_LIST[obj_idx+1][0][col_idx]}:')
                [print(f'        {_}') for _ in FILTER_OBJECTS[obj_idx][col_idx]]

        del FILTERED_OBJS_AND_COLS, OBJ_NAME_BEEN_PRINTED, FILTER_OBJECTS

        DF_COLUMNS = [[_ for _ in self.CSUTM_DF][__] for __ in self.DISPLAY_COLUMNS]

        if self.display_select in 'HB':
            print(f'\nTOP {self.display_rows} SCORES:')
            print(self.CSUTM_DF[DF_COLUMNS].head(self.display_rows))

        if self.display_select in 'B':
            print('')   # PUT A SPACE BETWEEN 'HEAD' & 'TAIL'

        if self.display_select in 'TB':
            print(f'BOTTOM {self.display_rows} SCORES:')
            print(self.CSUTM_DF[DF_COLUMNS].tail(self.display_rows))

        if self.display_select in 'A':
            print(f'\nALL {self.display_rows} SCORES:')
            print(self.CSUTM_DF[DF_COLUMNS])


    def test_filedump_module(self):
        # package-specific module for saving test results
        # In child, put package-specific summary dump module here, then put general test dump module
        # self.wb = gtrd.general_test_results_dump(self.wb, self.CSUTM_DF, self.DISPLAY_COLUMNS, self.display_select,
        #                                          self.display_rows)
        pass

    # END CALC DATA DISPLAY ##############################A##############################A##############################A##############


    def base_return_fxn(self):

        return self.TRAIN_SWNL, self.DEV_SWNL, self.TEST_SWNL, self.PERTURBER_RESULTS


    def return_fxn(self):
        # return *self.base_return_fxn() + whatever for specific ML package in child
        pass


    def run(self):  # MLRunTemplate loop

        while True:

            self.module_specific_main_menu_operations()

            if self.post_run_select == 'O':    #   'print hyperparameter settings to screen(O)'
                self.hyperparameter_display_module()

            elif self.post_run_select == 'L':    #   'display objects and columns(l)'

                margin = 10
                width = 25


                def get_header(OBJECT, SUP_OBJ, obj_idx, col_idx, obj_run_orient):
                    try:
                        # IF ERRORS ON THIS, RETURN EMPTY, ELSE RETURN ITS HEADER
                        # HAVE TO USE THIS TO ACCOUNT FOR COULD BE NP OR SD, ROW OR COLUMNS
                        TestClass = mlrco.MLRowColumnOperations(OBJECT[obj_idx], obj_run_orient, name='DUM', bypass_validation=self.bypass_validation)
                        TestClass.return_columns([col_idx], return_orientation='COLUMN', return_format='ARRAY')
                        del TestClass
                        return SUP_OBJ[obj_idx][msod.QUICK_POSN_DICT()["HEADER"]][col_idx][:width-2].ljust(width)
                    except:
                        try: del TestClass
                        except: pass
                        return " " * width


                # BUILD A GRID OF OBJECT LENGTHS, TO FIND THE MAX (CAN JUST MAP len HERE, COULD BE NO OBJECT AND WOULD EXCEPT)
                print(' '*margin + f'RAW'.ljust(width) + f'WORKING'.ljust(width) + f'TRAIN'.ljust(width)+ f'DEV'.ljust(width)+ f'TEST'.ljust(width))
                GRID = np.empty((3,5), dtype=np.int32)   # EACH INTERIOR [] HOLDS THE # COLUMNS FOR ALL DATAS, ALL TARGETS, & ALL REFVECS
                OBJECTS = [self.SUPER_RAW_NUMPY_LIST, self.SUPER_WORKING_NUMPY_LIST, self.TRAIN_SWNL, self.DEV_SWNL, self.TEST_SWNL]
                ORIENTS = [self.data_run_orientation, self.target_run_orientation, self.refvecs_run_orientation]

                for obj_idx in range(len(self.SUPER_WORKING_NUMPY_LIST)):  # ALL OF THESE MUST BE EQUAL LEN, SO COULD PICK ANY THAT ISNT EMPTY
                    for set_idx, SET in enumerate(OBJECTS):
                        try: GRID[obj_idx][set_idx] = gs.get_shape(self.SUPER_NUMPY_DICT[obj_idx], SET[obj_idx], ORIENTS[obj_idx])[1]
                        except: GRID[obj_idx][set_idx] = 0

                    print(self.SUPER_NUMPY_DICT[obj_idx].ljust(width))

                    for disp_idx, col_idx in enumerate(range(np.max(GRID[obj_idx]))):   # GET MAX LEN OF ALL RESPECTIVE DATA OBJECTS, TARGETS, ETC.
                        print(f'{str(disp_idx)}'.ljust(margin) + \
                              f'{get_header(self.SUPER_RAW_NUMPY_LIST, self.RAW_SUPOBJS, obj_idx, col_idx, ORIENTS[obj_idx])}' + \
                              f'{get_header(self.SUPER_WORKING_NUMPY_LIST, self.WORKING_SUPOBJS, obj_idx, col_idx, ORIENTS[obj_idx])}' + \
                              f'{get_header(self.TRAIN_SWNL, self.WORKING_SUPOBJS, obj_idx, col_idx, ORIENTS[obj_idx])}' + \
                              f'{get_header(self.DEV_SWNL, self.WORKING_SUPOBJS, obj_idx, col_idx, ORIENTS[obj_idx])}' + \
                              f'{get_header(self.TEST_SWNL, self.WORKING_SUPOBJS, obj_idx, col_idx, ORIENTS[obj_idx])}')

                    print()
                del get_header, GRID, OBJECTS, ORIENTS

            elif self.post_run_select == 'X':   #   'print frequencies(x)'
                SETS = [self.SUPER_RAW_NUMPY_LIST, self.SUPER_WORKING_NUMPY_LIST, self.TRAIN_SWNL, self.DEV_SWNL, self.TEST_SWNL]
                SET_NAMES = ['SUPER_RAW_NUMPY_LIST', 'SUPER_WORKING_NUMPY_LIST', 'TRAIN_SWNL', 'DEV_SWNL', 'TEST_SWNL']
                ORIENTS = [self.data_run_orientation, self.target_run_orientation, self.refvecs_run_orientation]
                ACTV_IDXS, ACTV_SETS, ACTV_NAMES = [],[],[]
                for idx, (OBJ, name) in enumerate(zip(SETS, SET_NAMES)):
                    if not np.array_equiv(OBJ, []) and not OBJ is None:
                        ACTV_IDXS.append(idx)
                        ACTV_SETS.append(OBJ)
                        ACTV_NAMES.append(name)

                op_select = 'Z'  # FORCES FULL MENU
                while True:
                    if op_select == 'E':
                        del SETS, SET_NAMES, ORIENTS, ACTV_IDXS, ACTV_SETS, ACTV_NAMES
                        break
                    if op_select in 'ZA':
                        set_idx = ACTV_IDXS[ls.list_single_select(ACTV_NAMES, f'\nSelect set of objects for frequency analysis', 'idx')[0]]
                    if op_select in 'ZAB':
                        obj_idx = ls.list_single_select([*self.SUPER_NUMPY_DICT.values()], f'\nSelect object', 'idx')[0]
                    if op_select in 'ZABC':
                        if SET_NAMES[set_idx] == 'SUPER_RAW_NUMPY_LIST': ACTV_SUPOBJS = self.RAW_SUPOBJS
                        else: ACTV_SUPOBJS = self.WORKING_SUPOBJS
                        col_idx = ls.list_single_select(ACTV_SUPOBJS[obj_idx][msod.QUICK_POSN_DICT()["HEADER"]], f'Select column', 'idx')[0]

                    # UNIQUES, COUNTS = np.unique(xx[obj_idx][col_idx].astype(object), return_counts=True)
                    # 12-23-21 CANT USE np.unique HERE IT CANT HANDLE STR & FLOAT MIXED TOGETHER, EVEN WHEN astype(object)
                    # UNIQUES = ru.return_uniques(SETS[set_idx][obj_idx][col_idx], [], 'STR', suppress_print='Y')[0]

                    ObjClass = mlo.MLObject(SETS[set_idx][obj_idx],
                                            ORIENTS[obj_idx],
                                            name=SET_NAMES[set_idx],
                                            return_orientation='AS_GIVEN',
                                            return_format='AS_GIVEN',
                                            bypass_validation=self.bypass_validation,
                                            calling_module=self.this_module,
                                            calling_fxn=inspect.stack()[0][3])

                    ACTV_COLUMN = ObjClass.return_columns([col_idx], return_orientation='COLUMN', return_format='ARRAY').astype(str)
                    UNIQUES, COUNTS = np.unique(ACTV_COLUMN, return_counts=True)
                    del ObjClass

                    MASK = np.flip(np.argsort(COUNTS))  # DESCENDING
                    UNIQUES = UNIQUES[..., MASK]
                    COUNTS = COUNTS[..., MASK]
                    del MASK

                    print(f'\nFREQUENCIES IN {SET_NAMES[set_idx]} - {self.SUPER_NUMPY_DICT[obj_idx]} - {ACTV_SUPOBJS[obj_idx][msod.QUICK_POSN_DICT()["HEADER"]][col_idx]}:')
                    del ACTV_SUPOBJS

                    # GET MAX LEN OUT OF UNIQUES
                    freq_max_len = max(len('ENTRY '), max(map(len, UNIQUES)) + 3)
                    print(f'RANK'.ljust(6) + f'VALUE'.ljust(freq_max_len) + f'CT')
                    [print(f'{_ + 1})'.ljust(6) + f'{UNIQUES[_].ljust(freq_max_len)}' + f'{COUNTS[_]}') for _ in range(len(UNIQUES))]

                    del UNIQUES, COUNTS, freq_max_len

                    op_select = vui.validate_user_str(f'\nSelect new set(a), select new object from current set(b), '
                                                      f'select new column from current object(c), exit(e) > ', 'ABCE')

            elif self.post_run_select == 'N':   #  'dump WORKING DATA objects to file(n)'
                oed.ObjectsExcelDump(self.SUPER_WORKING_NUMPY_LIST, 'WORKING').dump()

            elif self.post_run_select == 'B':   #  'run train without reconfig(b)'
                if not np.array_equiv(self.DEV_SWNL, []):
                    if vui.validate_user_str(f'\nUse early-stopping? (y/n) > ', 'YN') == 'Y':
                        self.early_stop_interval = vui.validate_user_int( f'Enter early-stop dev validation interval > ', min=1)
                _ = datetime.datetime.now().ctime()
                print(f'\nCURRENT TRAIN START DATE / TIME: {_[:11] + _[-4:] + _[10:-5]}')

                try: self.rebuild_kernel = True           # ONLY FOR SVMRun & MLR (NOT GMLR) ... FOR MANAGING K-FOLD KERNEL REBUILDS
                except: pass

                self.core_training_code()
                self.early_stop_interval = 1e12

                self.train_has_been_run = True

                # 9/27/22 PRINT AT END OF RUN FOR MEMORY JOGGING AFTER OVERNIGHT RUNS
                print()
                self.hyperparameter_display_module()
                print()

            elif self.post_run_select == 'V':   #   'print train target & output vectors(v)'
                # TRAIN_TARGET_VECTOR, TRAIN_TARGET_VECTOR_HEADER, ROWID_VECTOR, TRAIN_OUTPUT_VECTOR

                ROW_ID_VECTOR = mlo.MLObject(self.TRAIN_SWNL[2], self.refvecs_run_orientation, 'REFVECS',
                    return_orientation='COLUMN', return_format='ARRAY', bypass_validation=self.bypass_validation,
                    calling_module=self.this_module, calling_fxn=inspect.stack()[0][3]).return_columns([0],
                                                                    return_orientation='COLUMN', return_format='ARRAY')

                if len(self.TRAIN_OUTPUT_VECTOR) == len(self.TRAIN_SWNL[1]):
                    output_orientation = self.target_run_orientation
                else:
                    output_orientation = {'ROW':'COLUMN','COLUMN':'ROW'}[self.target_run_orientation]

                pv.print_vectors(self.WORKING_SUPOBJS[1][msod.QUICK_POSN_DICT()["HEADER"]], self.TRAIN_SWNL[1],
                                 self.target_run_orientation, self.TRAIN_OUTPUT_VECTOR, output_orientation,
                                 ROW_ID_VECTOR)

                del ROW_ID_VECTOR, output_orientation

            elif self.post_run_select == 'S':    #  'print train summary stats(s)'
                self.train_summary_statistics_module()

            elif self.post_run_select == 'P':  # 'print train results to screen(p)'
                self.print_train_results_module()

            elif self.post_run_select == 'F':  # 'dump train results to file(f)'
                try:
                    self.wb = Workbook()
                    self.train_filedump_module()
                    self.filedump_path()
                    self.wb = Workbook()
                    print(f'\nTrain results saved to {self.full_path} successfully.\n')
                except:
                    print(f'\n*** EXCEPTION TRYING TO DUMP TRAIN RESULTS TO FILE ***\n')

            elif self.post_run_select == 'I':   # 'generate \ reset train objects(i)'
                self.reset_train_data()


            elif self.post_run_select == 'G':   #   'run dev(g)'

                while True:
                    print()
                    self.dev_run_select = dmp.DictMenuPrint(self.BASE_DEV_CALC_CMDS | self.sub_dev_calc_cmds(),
                                                                    disp_len=140).select(f'Select dev calc option')

                    if self.dev_run_select == 'A':
                        break

                    self.base_dev_calc_module()
                    self.sub_dev_calc_module()

            elif self.post_run_select == 'J':    #  'print dev results to screen(j)'
                self.print_dev_results_module()

            elif self.post_run_select == 'K':    # 'dump dev results to file(k)'
                self.wb = Workbook()
                self.filedump_general_ml_setup_module()
                self.filedump_package_specific_setup_module()
                self.dev_filedump_module()
                self.filedump_path()
                self.wb = Workbook()

            elif self.post_run_select == 'H':   # 'generate / reset dev objects(h)'
                self.dev_build()
                self.dev_error = 0
                # self.DEV_ERROR = [[]]
                self.RGLZTN_FACTORS = []

            elif self.post_run_select == 'R':   # 'calc test data(r)'

                while True:

                    if not self.train_has_been_run:
                       print(f'\n*** TRAIN HAS NOT BEEN RUN, MUST TRAIN PARAMETERS BEFORE RUNNING TEST ***\n')
                       break

                    print()
                    self.test_calc_select = dmp.DictMenuPrint(self.BASE_TEST_CALC_CMDS,
                                                              disp_len=140).select(f'Select test calc option')

                    if self.test_calc_select == 'A':
                        break

                    self.base_test_calc_module()
                    self.sub_test_calc_module()

            elif self.post_run_select == 'E':    # 'print test target & output vectors(e)'
                # TEST_TARGET_VECTOR, TEST_TARGET_VECTOR_HEADER, TEST_ROWID_VECTOR, TEST_OUTPUT_VECTOR

                ROW_ID_VECTOR = mlo.MLObject(self.TEST_SWNL[2], self.refvecs_run_orientation, 'REFVECS',
                    return_orientation='COLUMN', return_format='ARRAY', bypass_validation=self.bypass_validation,
                    calling_module=self.this_module, calling_fxn=inspect.stack()[0][3]).return_columns([0],
                                                                    return_orientation='COLUMN', return_format='ARRAY')

                if len(self.TEST_OUTPUT_VECTOR) == len(self.TEST_SWNL[1]): output_orientation = self.target_run_orientation
                else: output_orientation = {'ROW':'COLUMN','COLUMN':'ROW'}[self.target_run_orientation]

                pv.print_vectors(self.WORKING_SUPOBJS[1][msod.QUICK_POSN_DICT()["HEADER"]], self.TEST_SWNL[1],
                                 self.target_run_orientation, self.TEST_OUTPUT_VECTOR, output_orientation,
                                 ROW_ID_VECTOR)

                del ROW_ID_VECTOR, output_orientation


            elif self.post_run_select == 'U':     # 'print test summary stats(u)'
                self.test_summary_statistics_module()

            elif self.post_run_select == 'C':     # 'print test results to screen(c)'
                while True:
                    self.row_column_display_select()
                    self.print_test_results_module()
                    if vui.validate_user_str(f'\nTry again(t) or return to run menu(r) > ', 'TR') == 'R':
                        break

            elif self.post_run_select == 'D':     # 'dump test results to file(d)'
                while True:
                    self.wb = Workbook()
                    self.filedump_general_ml_setup_module()
                    self.filedump_package_specific_setup_module()
                    self.row_column_display_select()
                    self.test_filedump_module()

                    __ = vui.validate_user_str(f'Accept(a), try again(t) or abort(x) > ', 'ATX')
                    if __ == 'A': pass
                    elif __ == 'T': continue
                    elif __ == 'X': break

                    self.filedump_path()
                    self.wb = Workbook()
                    print(f'\nTest results successfully dumped to {self.full_path}.\n')
                    break

            elif self.post_run_select == 'T':   # 'generate / reset test objects(t)'
                self.test_build()

            elif self.post_run_select == 'W':  # 'run k-fold cross validation(w)'

                while True:
                    if vui.validate_user_str(f'\n*** REALLY PROCEED WITH K-FOLD CROSS VALIDATION? (y/n) *** > ', 'YN') == 'N':
                        break

                    if isinstance(self.rglztn_type, str): self.rglztn_type = self.rglztn_type.upper()

                    if self.rglztn_type in ['NONE', None]:
                        print(f'\n*** REGULARIZATION TYPE MUST BE SET BEFORE RUNNING K-FOLD ***\n')
                        break

                    # BEAR TOOK THIS OUT 6/25/23... WANT TEST SPLIT TO STAY SEPARATE AND NOT GET OVERWRIT
                    '''
                    # RESET TRAIN / DEV OBJECTS (ANY PREVIOUS TRAIN/DEV SPLIT WOULD BE CARRIED INTO rglztn_partition_iterator)
                    self.TRAIN_SWNL = np.array(
                        [_.copy() if isinstance(_, np.ndarray) else deepcopy(_) for _ in self.SUPER_WORKING_NUMPY_LIST], dtype=object)
                    self.DEV_SWNL = []
                    self.TEST_SWNL = []
                    self.CSUTM_DF = pd.DataFrame({})
                    self.DEV_ERROR = [[]]
                    '''

                    self.DEV_SWNL = []
                    self.DEV_ERROR = [[]]

                    while True:
                        number_of_partitions = vui.validate_user_int(f'Enter number of partitions > ', min=2,
                            max=gs.get_shape("DATA", self.SUPER_WORKING_NUMPY_LIST[0], self.data_run_orientation)[0]
                        )
                        if vui.validate_user_str(f'User entered {number_of_partitions} partitions... Accept? (y/n) > ', 'YN') == 'Y':
                            break

                    self.rglztn_partition_iterator(number_of_partitions, draw_method='PARTITION')

                    break

            elif self.post_run_select == 'Z':   # 'dump all results to file (z)'

                self.wb = Workbook()
                self.filedump_general_ml_setup_module()
                self.train_filedump_module()

                if not self.DEV_ERROR == [[]]:
                    self.dev_filedump_module()

                if not self.CSUTM_DF.equals(pd.DataFrame({})):
                    self.row_column_display_select()
                    self.test_filedump_module()

                if not self.PERTURBER_RESULTS.equals(pd.DataFrame({})):
                    sheetname = 'PERTURBER RESULTS TABLE'
                    self.wb.create_sheet(sheetname)

                    start_row, start_col = 2, 2

                    # MAKE HEADER
                    for col_idx, hdr in enumerate(self.PERTURBER_RESULTS.keys(), start_col+1):
                        ow.openpyxl_write(self.wb, sheetname, start_row, col_idx, hdr, horiz='center', vert='center', bold=True)
                    # MAKE INDEX
                    for row_idx, column_name in enumerate(self.PERTURBER_RESULTS.index, start_row+1):
                        ow.openpyxl_write(self.wb, sheetname, row_idx, start_col, column_name, horiz='left', vert='center', bold=True)
                    # FILL DATA
                    for row_idx in range(self.PERTURBER_RESULTS.shape[0]):
                        for col_idx in range(self.PERTURBER_RESULTS.shape[1]):
                            ow.openpyxl_write(self.wb, sheetname,
                                              row_idx + start_row + 1,
                                              col_idx + start_col + 1,
                                              self.PERTURBER_RESULTS.iloc[row_idx, col_idx],
                                              horiz='center',
                                              vert='center',
                                              bold=False
                            )

                self.filedump_path()
                self.wb = Workbook()
                print(f'\n*** FILE DUMP OF ALL RESULTS TO FILE SUCCESSFUL. ***\n')

            elif self.post_run_select == 'A':   # 'accept, return to config-run menu(e)'
                break

            elif self.post_run_select == 'Q':    #  'quit'
                sys.exit(f'User terminated.')

            ############################################################################################################
            # STUFF USED TO MANAGE MAIN MENU OPTIONS DISPLAY #######################################################

            disallowed = self.base_disallowed

            if len(self.TRAIN_SWNL)==0: disallowed += 'BPFWHTZ'

            if len(self.TRAIN_OUTPUT_VECTOR)==0: disallowed += 'VS'

            if len(self.DEV_SWNL)==0: disallowed += 'G'

            if np.array_equiv(self.DEV_ERROR, [[]]): disallowed += 'JK'

            if len(self.TEST_SWNL)==0: disallowed += 'R'

            # IF TEST RESULTS HAVE NOT BEEN GENERATED (IE, STILL AN EMPTY DATAFRAME)
            if self.CSUTM_DF.equals(pd.DataFrame({})): disallowed += 'EUCD'

            # END STUFF USED TO MANAGE MAIN MENU OPTIONS DISPLAY ###################################################
            ########################################################################################################

            print()
            self.post_run_select = dmp.DictMenuPrint(self.BASE_MAIN_MENU_CMDS | self.module_specific_main_menu_cmds(),
                 disp_len=140, disallowed=disallowed).select(f'Select letter')


        return self.return_fxn()





























if __name__ == '__main__':

    from MLObjects.TestObjectCreators.SXNL import CreateSXNL as csxnl

    DATA = pd.read_csv(r'C:\Users\Bill\Documents\WORK STUFF\RESUME\1 - OTHER\SRP\beer_reviews.csv',
                      nrows=100,
                      header=0).dropna(axis=0)

    DATA = DATA[DATA.keys()[[3, 4, 5, 7, 8, 9, 11]]]

    TARGET = DATA['review_overall'].to_numpy()
    TARGET.resize(1, len(TARGET))
    TARGET_HEADER = [['review_overall']]
    target_given_orientation = 'COLUMN'

    DATA = DATA.drop(columns=['review_overall'])

    # KEEP THIS FOR SRNL
    RAW_DATA = DATA.copy()
    RAW_DATA_HEADER = np.fromiter(RAW_DATA.keys(), dtype='<U50').reshape((1,-1))
    RAW_DATA = RAW_DATA.to_numpy().transpose()
    data_given_orientation = 'COLUMN'

    REF_VEC = np.fromiter(range(len(RAW_DATA[0])), dtype=int).reshape((1,-1))
    REF_VEC_HEADER = [['ROW_ID']]
    refvecs_given_orientation = 'COLUMN'



    data_return_orientation = 'COLUMN'
    target_return_orientation = 'COLUMN'
    refvecs_return_orientation = 'COLUMN'



    SXNLClass = csxnl.CreateSXNL(rows=None,
                                 bypass_validation=True,
                                 data_return_format='ARRAY',
                                 data_return_orientation=data_return_orientation,
                                 DATA_OBJECT=RAW_DATA,
                                 DATA_OBJECT_HEADER=RAW_DATA_HEADER,
                                 DATA_FULL_SUPOBJ_OR_SINGLE_MDTYPES=None,
                                 data_override_sup_obj=False,
                                 data_given_orientation=data_given_orientation,
                                 data_columns=None,
                                 DATA_BUILD_FROM_MOD_DTYPES=None,
                                 DATA_NUMBER_OF_CATEGORIES=None,
                                 DATA_MIN_VALUES=None,
                                 DATA_MAX_VALUES=None,
                                 DATA_SPARSITIES=None,
                                 DATA_WORD_COUNT=None,
                                 DATA_POOL_SIZE=None,
                                 target_return_format='ARRAY',
                                 target_return_orientation=target_return_orientation,
                                 TARGET_OBJECT=TARGET,
                                 TARGET_OBJECT_HEADER=TARGET_HEADER,
                                 TARGET_FULL_SUPOBJ_OR_SINGLE_MDTYPES=None,
                                 target_type='FLOAT',
                                 target_override_sup_obj=False,
                                 target_given_orientation=target_given_orientation,
                                 target_sparsity=None,
                                 target_build_from_mod_dtype=None,
                                 target_min_value=None,
                                 target_max_value=None,
                                 target_number_of_categories=None,
                                 refvecs_return_format='ARRAY',
                                 refvecs_return_orientation=refvecs_return_orientation,
                                 REFVECS_OBJECT=REF_VEC,
                                 REFVECS_OBJECT_HEADER=REF_VEC_HEADER,
                                 REFVECS_FULL_SUPOBJ_OR_SINGLE_MDTYPES=None,
                                 REFVECS_BUILD_FROM_MOD_DTYPES=None,
                                 refvecs_override_sup_obj=False,
                                 refvecs_given_orientation=refvecs_given_orientation,
                                 refvecs_columns=None,
                                 REFVECS_NUMBER_OF_CATEGORIES=None,
                                 REFVECS_MIN_VALUES=None,
                                 REFVECS_MAX_VALUES=None,
                                 REFVECS_SPARSITIES=None,
                                 REFVECS_WORD_COUNT=None,
                                 REFVECS_POOL_SIZE=None
                                 )

    SRNL = SXNLClass.SXNL.copy()
    SRNL_SUPOBJS = SXNLClass.SXNL_SUPPORT_OBJECTS.copy()

    # expand #########################################################################################################
    SXNLClass.expand_data(expand_as_sparse_dict=False, auto_drop_rightmost_column=False)

    SWNL = SXNLClass.SXNL
    SWNL_SUPOBJS = SXNLClass.SXNL_SUPPORT_OBJECTS

    standard_config = 'None'
    sub_config = 'None'
    split_method = 'None'
    LABEL_RULES = []
    number_of_labels = 1
    event_value = ''
    negative_value = ''
    conv_kill = 1000
    pct_change = 0.1
    conv_end_method = 'KILL'
    rglztn_type = 'L2'
    rglztn_fctr = 1
    bypass_validation = False
    module = __name__

    TestClass = MLRunTemplate(
                                standard_config,
                                sub_config,
                                SRNL,
                                SRNL_SUPOBJS,
                                SWNL,
                                SWNL_SUPOBJS,
                                data_return_orientation,
                                target_return_orientation,
                                refvecs_return_orientation,
                                [],    # CONTEXT
                                [],    # KEEP
                                [],    # TRAIN_SWNL
                                [],    # DEV_SWNL
                                [],    # TEST_SWNL
                                split_method,
                                LABEL_RULES,
                                number_of_labels,
                                event_value,
                                negative_value,
                                conv_kill,
                                pct_change,
                                conv_end_method,
                                rglztn_type,
                                rglztn_fctr,
                                bypass_validation,
                                module
    ).run()

























