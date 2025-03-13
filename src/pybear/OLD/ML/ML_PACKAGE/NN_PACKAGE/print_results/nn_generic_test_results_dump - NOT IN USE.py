from openpyxl import Workbook
from ML_PACKAGE.openpyxl_shorthand import openpyxl_write as ow
from read_write_file.generate_full_filename import base_path_select as bps, filename_enter as fe

# NOT IN USE AS OF 4-1-22
# USING GENERIC_PRINT.general_test_results_dump IN NNRun


def nn_generic_test_results_dump(standard_config, gd_method, conv_method, lr_method, LEARNING_RATE, METADATA_KEEP, START_DATE,
             END_DATE, FILTER_MATRIX, interaction, int_cutoff, intercept, SELECT_LINK_FXN, non_neg_coeffs, rglztn_fctr,
             ROWID_VECTOR, TARGET_VECTOR, OUTPUT_VECTOR, ARRAY_OF_NODES, RETAINED_ATTRIBUTES, BIG_MATRIX, COLIN_CHOPPED,
             CSUTM_DF, display_select, display_rows, things_in_intercept):

    # CURRENT *params GOING IN nn_generic_test_results_dump IN NNRun
    '''
    self.standard_config, self.gd_method,
    self.SUPER_WORKING_NUMPY_LIST, self.WORKING_VALIDATED_DATATYPES, self.WORKING_MODIFIED_DATATYPES,
    self.WORKING_FILTERING, self.WORKING_MIN_CUTOFFS, self.WORKING_USE_OTHER, self.WORKING_CONTEXT,
    self.WORKING_KEEP, self.split_method, self.LABEL_RULES, self.number_of_labels, self.event_value,
    self.negative_value, self.gd_method, self.conv_method, self.lr_method, self.LEARNING_RATE,
    self.SELECT_LINK_FXN, self.non_neg_coeffs, self.ROWID_VECTOR, self.OUTPUT_VECTOR, self.ARRAY_OF_NODES,
    self.CSUTM_DF, self.display_select, self.display_rows'''

    if len(ROWID_VECTOR) != len(BIG_MATRIX[0]):
        raise ValueError(f'len(ROWID_VECTOR) != len(DATA[0) in nn_generic_test_results_dump().')


    #ESTABLISH FILE PATH
    base_path = bps.base_path_select()
    file_name = fe.filename_wo_extension()
    full_path = base_path + file_name + '.xlsx'


    #CREATE FILE & SETUP SHEET
    wb = Workbook()
    wb.active.title = 'SETUP'


    def custom_write(sheet, row, column, value, horiz, vert, bold):
        ow.openpyxl_write(wb, sheet, row, column, value, horiz=horiz, vert=vert, bold=bold)


    custom_write('SETUP', 1, 1, 'SETUP', 'left', 'center', True)

    row_counter = 2



    #BUILD RESULTS SHEET
    wb.create_sheet('RESULT')

    custom_write('RESULTS', 1, 1, "ROW ID", 'center', 'center', True)
    for label_idx in range(len(TARGET_VECTOR)):
        custom_write('RESULTS', 1, 2+label_idx, f"TARGET{label_idx+1}", 'center', 'center', True)
    for output_idx in range(len(OUTPUT_VECTOR)):
        custom_write('RESULTS', 1, 2+len(TARGET_VECTOR)+output_idx, f"OUTPUT{output_idx+1}", 'center', 'center', True)

    TARGET_VECTOR_HIT = []
    OUTPUT_VECTOR_HIT = []
    for example in range(0, len(TARGET_VECTOR[0])):
        # DISPLAY EXAMPLE NUMBER
        custom_write('RESULTS', 2+example, 1, ROWID_VECTOR[example], 'center', 'center', True)
        wb['RESULT'].cell(2 + example, 1).number_format = '0'

        # BUILD TARGET & OUTPUT VECTORS THAT CONTAINS ONLY THE "HITS"... NEEDED FOR SOFTMAX
        # IF TARGET LEN>1 -> SOFTMAX, HIT MUST BE 1
        [TARGET_VECTOR_HIT.append(TARGET_VECTOR[0][example]) if len(TARGET_VECTOR) == 1 else TARGET_VECTOR_HIT.append(1)]

        [OUTPUT_VECTOR_HIT.append(OUTPUT_VECTOR[0][example]) if len(OUTPUT_VECTOR) == 1 else '']

        for label_idx in range(len(TARGET_VECTOR)):
            if len(OUTPUT_VECTOR) > 1:
                if TARGET_VECTOR[label_idx][example] == 1:
                    OUTPUT_VECTOR_HIT.append(OUTPUT_VECTOR[label_idx][example])

            # DISPLAY TARGET_VECTOR
            custom_write('RESULTS', 2+example, 2+label_idx, TARGET_VECTOR[example], 'center', 'center', False)
            wb['RESULT'].cell(2+example, 2+label_idx).number_format = '0'

            # DISPLAY OUTPUT_VECTOR
        for output_idx in range(len(OUTPUT_VECTOR)):
            custom_write('RESULTS', 2+example, 1+len(TARGET_VECTOR)+output_idx+1, OUTPUT_VECTOR[example], 'center', 'center', False)
            if OUTPUT_VECTOR[example] == 0:
                wb['RESULT'].cell(2+example, 1+len(TARGET_VECTOR)+output_idx+1).number_format = '0'
            else: wb['RESULT'].cell(2 + example, 1+len(TARGET_VECTOR)+output_idx+1).number_format = '0.00000000'


    # CREATE CHART ON RESULTS SHEET
    from openpyxl.chart import ScatterChart, Reference, Series

    chart = ScatterChart()
    chart.title = 'TARGET_VECTOR  vs.  OUTPUT_VECTOR'
    chart.style = 13
    chart.x_axis.title = 'Row ID'
    chart.y_axis.title = 'Target Probability'
    # xvalues = Reference(wb['RESULT'], min_col=1, min_row=1, max_row=len(TARGET_VECTOR)+2)
    xvalues = Reference(ROWID_VECTOR)

    for i in range(2):
        # values = Reference(wb['RESULT'], min_col=i, min_row=1, max_row=len(TARGET_VECTOR)+2)
        # series = Series(values, xvalues, title_from_data=True)
        # chart.series.append(series)
        DATA_HOLDER = [TARGET_VECTOR_HIT, OUTPUT_VECTOR_HIT]
        values = Reference(DATA_HOLDER[i])
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)

    wb['RESULT'].add_chart(chart, 'N2')










    wb.save(full_path)

    print('File dump complete.')





    # BUILD RESULTS SHEET
    wb.create_sheet('RESULTS')

    custom_write('RESULTS', 1, 1, "ROW ID", 'center', 'center', True)
    TARGET_VECTOR = SUPER_WORKING_NUMPY_LIST[2]
    OUTPUT_VECTOR = [[n.random.random() for _ in range(len(TARGET_VECTOR[0]))]]
    for label_idx in range(len(TARGET_VECTOR)):
        custom_write('RESULTS', 1, 2 + label_idx, f"TARGET{label_idx + 1}", 'center', 'center', True)
    for output_idx in range(len(OUTPUT_VECTOR)):
        custom_write('RESULTS', 1, 2 + len(TARGET_VECTOR) + output_idx, f"OUTPUT{output_idx + 1}", 'center', 'center', True)

    TARGET_VECTOR_HIT = []
    OUTPUT_VECTOR_HIT = []
    for example in range(0, len(TARGET_VECTOR[0])):
        # DISPLAY EXAMPLE NUMBER
        custom_write('RESULTS', 2 + example, 1, ROWID_VECTOR[example], 'center', 'center', True)
        wb['RESULTS'].cell(2 + example, 1).number_format = '0'

        # BUILD TARGET & OUTPUT VECTORS THAT CONTAINS ONLY THE "HITS"... NEEDED FOR SOFTMAX
        # IF TARGET LEN>1 -> SOFTMAX, HIT MUST BE 1
        [TARGET_VECTOR_HIT.append(TARGET_VECTOR[0][example]) if len(TARGET_VECTOR) == 1 else TARGET_VECTOR_HIT.append(1)]

        [OUTPUT_VECTOR_HIT.append(OUTPUT_VECTOR[0][example]) if len(OUTPUT_VECTOR) == 1 else '']

        for label_idx in range(len(TARGET_VECTOR)):
            if len(OUTPUT_VECTOR) > 1:
                if TARGET_VECTOR[label_idx][example] == 1:
                    OUTPUT_VECTOR_HIT.append(OUTPUT_VECTOR[label_idx][example])

            # DISPLAY TARGET_VECTOR
            custom_write('RESULTS', 2 + example, 2 + label_idx, TARGET_VECTOR[label_idx][example], 'center', 'center',
                         False)
            wb['RESULTS'].cell(2 + example, 2 + label_idx).number_format = '0'

            # DISPLAY OUTPUT_VECTOR
        for output_idx in range(len(OUTPUT_VECTOR)):
            custom_write('RESULTS', 2 + example, 1 + len(TARGET_VECTOR) + output_idx + 1,
                         OUTPUT_VECTOR[output_idx][example], 'center', 'center', False)
            if OUTPUT_VECTOR[output_idx][example] == 0:
                wb['RESULTS'].cell(2 + example, 1 + len(TARGET_VECTOR) + output_idx + 1).number_format = '0'
            else:
                wb['RESULTS'].cell(2 + example, 1 + len(TARGET_VECTOR) + output_idx + 1).number_format = '0.00000000'

    # CREATE CHART ON RESULTS SHEET
    from openpyxl.chart import ScatterChart, Reference, Series

    chart = ScatterChart()
    chart.title = 'TARGET_VECTOR  vs.  OUTPUT_VECTOR'
    chart.style = 13
    chart.x_axis.title = 'Row ID'
    chart.y_axis.title = 'Target Probability'
    xvalues = Reference(wb['RESULTS'], min_col=2, min_row=1, max_row=len(TARGET_VECTOR) + 2)
    # xvalues = ROWID_VECTOR

    DATA_HOLDER = [TARGET_VECTOR_HIT, OUTPUT_VECTOR_HIT]
    for i in range(1, 3):
        values = Reference(wb['RESULTS'], min_col=i, min_row=1, max_row=len(TARGET_VECTOR) + 2)
        series = Series(values, xvalues, title_from_data=False)
        chart.series.append(series)
        # values = DATA_HOLDER[i] # Reference(DATA_HOLDER[i])
        # series = Series(values, xvalues, title_from_data=False)
        # chart.series.append(series)

    wb['RESULTS'].add_chart(chart, 'N2')

    # BUILD BIG_MATRIX SHEET
    wb.create_sheet(f'DATA')

    # HEADER
    custom_write('DATA', 2, 1, 'ROW ID', 'center', 'center', True)

    # LABEL ROWS W ROW_ID
    for column in range(len(SUPER_WORKING_NUMPY_LIST[0][0])):
        custom_write('DATA', column + 3, 1, ROWID_VECTOR[column], 'center', 'center', True)

    # LABEL COLUMNS W ATTRIBUTE & CATEGORY
    for row in range(len(SUPER_WORKING_NUMPY_LIST[1][0])):
        custom_write('DATA', 1, row + 2, SUPER_WORKING_NUMPY_LIST[1][0][row], 'center', 'center', True)
        # custom_write('DATA', 2, row+2, SUPER_WORKING_NUMPY_LIST[1][0][row], 'center', 'center', True)

    # FILL SPREADSHEET W TRANSPOSED DATA
    for row in range(len(SUPER_WORKING_NUMPY_LIST[0])):
        for column in range(len(SUPER_WORKING_NUMPY_LIST[0][0])):
            custom_write('DATA', column + 3, row + 2, SUPER_WORKING_NUMPY_LIST[0][row][column], 'center', 'center', False)





if __name__ == '__main__':
    standard_config = 'AA'
    gd_method = 'G'
    conv_method = 'G'
    lr_method = 'C'
    learning_rate = [.01]
    METADATA_KEEP = []
    START_DATE = ''
    END_DATE = ''
    FILTER_MATRIX = []
    interaction = 'N'
    int_cutoff = 0
    intercept = 'N'
    SELECT_LINK_FXN = ['Logistic', 'Logistic']
    non_neg_coeffs = 'N'
    rglztn_fctr = 0.5
    ROWID_VECTOR = [0,1,2,3]
    TARGET_VECTOR = [0,1,0,1]
    OUTPUT_VECTOR = [0.2,0.9,0.1,0.8]
    ARRAY_OF_NODES = [[],[]]
    RETAINED_ATTRIBUTES = [[]]
    BIG_MATRIX = [[]]
    COLIN_CHOPPED = [[]]
    CSUTM_DF = []
    display_select = 'B'
    display_rows = 20
    things_in_intercept = []


    nn_generic_test_results_dump(standard_config, gd_method, conv_method, lr_method, learning_rate, METADATA_KEEP,
                                 START_DATE, END_DATE, FILTER_MATRIX, interaction, int_cutoff, intercept, SELECT_LINK_FXN,
                                 non_neg_coeffs, rglztn_fctr, ROWID_VECTOR, TARGET_VECTOR, OUTPUT_VECTOR, ARRAY_OF_NODES,
                                 RETAINED_ATTRIBUTES, BIG_MATRIX, COLIN_CHOPPED, CSUTM_DF, display_select, display_rows,
                                 things_in_intercept)










