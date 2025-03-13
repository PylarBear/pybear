from general_list_ops import list_select as ls
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from read_write_file.generate_full_filename import base_path_select as bps, filename_enter as fe


#CALLED BY NNRun
def AON_excel_dump(ARRAY_OF_NODES):

    wb = Workbook()

    wb.active.title = 'NODE 0'
    for row_idx in range(len(ARRAY_OF_NODES[0])):
        for elmt_idx in range(len(ARRAY_OF_NODES[0][row_idx])):
            wb['NODE 0'].cell(row_idx+1, elmt_idx+1).value = \
                ARRAY_OF_NODES[0][row_idx][elmt_idx]


    # 2-2-22 SEEM TO REMEMBER HAD TO BREAK THESE OUT INTO SEPARATE if STATEMENTS, COULDNT USE A for LOOP TO
    # GENERATE NEW WORKSHEETS, FOR SOME REASONS wb.create_sheet() WOULDNT WORK UNDER THE for LOOP (??)
    if len(ARRAY_OF_NODES) >= 2:
        wb.create_sheet('NODE 1')
        for row_idx in range(len(ARRAY_OF_NODES[1])):
            for elmt_idx in range(len(ARRAY_OF_NODES[1][row_idx])):
                wb['NODE 1'].cell(row_idx+1, elmt_idx+1).value = ARRAY_OF_NODES[1][row_idx][elmt_idx]

    if len(ARRAY_OF_NODES) >= 3:
        wb.create_sheet('NODE 2')
        for row_idx in range(len(ARRAY_OF_NODES[2])):
            for elmt_idx in range(len(ARRAY_OF_NODES[2][row_idx])):
                wb['NODE 2'].cell(row_idx+1, elmt_idx+1).value = ARRAY_OF_NODES[2][row_idx][elmt_idx]

    if len(ARRAY_OF_NODES) >= 4:
        wb.create_sheet('NODE 3')
        for row_idx in ARRAY_OF_NODES[3]:
            for elmt_idx in ARRAY_OF_NODES[3][row_idx]:
                wb['NODE 3'].cell(row_idx+1, elmt_idx+1).value = ARRAY_OF_NODES[3][row_idx][elmt_idx]

    if len(ARRAY_OF_NODES) >= 5:
        wb.create_sheet('NODE 4')
        for row_idx in ARRAY_OF_NODES[4]:
            for elmt_idx in ARRAY_OF_NODES[4][row_idx]:
                wb['NODE 4'].cell(row_idx+1, elmt_idx+1).value = ARRAY_OF_NODES[4][row_idx][elmt_idx]


    # wb['NODE 0'].cell(1, 1).value = 'SETUP'
    # wb['SETUP'].cell(1, 1).alignment = Alignment(horizontal='left', vertical='center')
    # wb['SETUP'].cell(1, 1).font = Font(bold='True')


    # CREATE FILE & SETUP SHEET
    while True:
        try:
            base_path = bps.base_path_select()
            file_name = fe.filename_wo_extension()
            full_path = base_path + file_name + '.xlsx'
            print(f'\nEntered path is > {full_path}\n')
            wb.save(full_path)
            print(f'\nARRAY_OF_NODES elements saved to {full_path} successfully.\n')
            break
        except:
            print(f'\n*** INVALID PATH & FILENAME, TRY AGAIN ***')
            continue


















