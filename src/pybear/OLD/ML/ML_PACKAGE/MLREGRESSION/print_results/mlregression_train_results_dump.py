from data_validation import validate_user_input as vui
from ML_PACKAGE.openpyxl_shorthand import openpyxl_write as ow
from openpyxl.utils.dataframe import dataframe_to_rows


def mlregression_train_results_dump(wb, TRAIN_RESULTS, rglztn_fctr):
    # TRAIN_RESULTS IS DF

    SORT_DICT = {'P': 'p VALUE', 'C': 'COEFFS', 'A': 'ABSOLUTE'} if rglztn_fctr==0 else {'C': 'COEFFS', 'A': 'ABSOLUTE'}
    allowed = "".join(list(SORT_DICT.keys()))
    _text = f'\nSort by{" p-values(p), " if rglztn_fctr==0 else " "}coeffs(c), or absolute value of coeffs(a)? > '

    sort_column = SORT_DICT[vui.validate_user_str(_text, allowed)]
    sort_order = {'A': True, 'D': False}[vui.validate_user_str(f'\nSort ascending(a) or descending(d)? > ', 'AD')]

    del SORT_DICT, allowed, _text

    OVERALL_RESULTS_HOLDER = TRAIN_RESULTS.iloc[0: -4:]

    if sort_column in ['p VALUE', 'COEFFS']:
        SORTED_TR = TRAIN_RESULTS.sort_values(by=[('      ', sort_column)], ascending=sort_order)
    elif sort_column == 'ABSOLUTE':
        SORTED_TR = TRAIN_RESULTS.sort_values(by=[('      ', 'COEFFS')], key=abs, ascending=sort_order)

    SORTED_TR.iloc[0: -4:] = OVERALL_RESULTS_HOLDER

    del sort_column, sort_order, OVERALL_RESULTS_HOLDER


    wb.create_sheet('MLRegression TRAIN RESULTS')

    row_counter = 1

    ow.openpyxl_write(wb, 'MLRegression TRAIN RESULTS', row_counter, 1, 'MLRegression TRAIN RESULTS', 'left', 'center', True)

    row_counter += 2

    # FILL IN DATA
    for grid_row_idx, row in enumerate(dataframe_to_rows(SORTED_TR, index=False, header=True), row_counter):

        for grid_col_idx, value in enumerate(row, 3):
            # if grid_col_idx == 3 and grid_row_idx != row_counter: align = 'left'
            # else: align = 'center'
            if grid_row_idx-row_counter <= 1: bold = True
            else: bold = False

            ow.openpyxl_write(wb, 'MLRegression TRAIN RESULTS', grid_row_idx, grid_col_idx, value, 'center', 'center', bold)

    for grid_row_idx, hdr in enumerate(TRAIN_RESULTS.index, row_counter+2):
        ow.openpyxl_write(wb, 'MLRegression TRAIN RESULTS', grid_row_idx, 2, hdr, 'left', 'center', False)


    del SORTED_TR




    '''   NEWEST STUFF TAKEN FROM MLRegressionCoreRunCode
    ###################################################################################################################
    # FUNCTION FOR HANDLING DUMP OF FULL OR GOOD RESULTS TO EXCEL #####################################################

    sheet_name = 'TRAIN_RESULTS'

    base_path = bps.base_path_select()
    file_name = fe.filename_wo_extension()

    full_path = base_path + file_name + ".xlsx"
    print(f'\nSaving file to {full_path}....')

    with pd.ExcelWriter(full_path) as writer:
        # index must be True, NotImplementedError: Writing to Excel with MultiIndex columns and no index ('index'=False) is not yet implemented.
        self.TRAIN_RESULTS.style.set_properties(**{'text-align': 'center'}).to_excel(
            excel_writer=writer,
            sheet_name=sheet_name, float_format='%.4f', startrow=1, startcol=1, merge_cells=False,
            index=True, na_rep='NaN')

    del PATH_LIST, file_name, full_path, base_path, sheet_name
    print('Done.')
    # END FUNCTION FOR HANDLING DUMP OF FULL OR GOOD RESULTS TO EXCEL #####################################################
    ###################################################################################################################
    '''

    return wb






if __name__ == '__main__':

    import numpy as np
    from openpyxl import Workbook
    from MLObjects.TestObjectCreators import test_header as th
    from read_write_file.generate_full_filename import base_path_select as bps
    from ML_PACKAGE.MLREGRESSION import build_empty_mlr_train_results as bemtr

    wb = Workbook()

    _rows = 5
    DATA_HEADER = th.test_header(_rows)
    DF = bemtr.build_empty_mlr_train_results(DATA_HEADER.reshape((1,-1)))

    TEST_DATA = np.random.uniform(0,1,(_rows,2))
    DF.iloc[:, :2] = TEST_DATA


    print(DF)
    __ = input(f'\nDISPLAY DF > ')

    wb.active.title = 'BLANK'
    wb['BLANK'].cell(1,1).value = 'INTENTIONALLY LEFT BLANK'

    rglztn_fctr = 0

    wb = mlregression_train_results_dump(wb, DF, rglztn_fctr)

    basepath = bps.base_path_select()

    wb.save(basepath + r'MLR_train_file_dump_test.xlsx')
