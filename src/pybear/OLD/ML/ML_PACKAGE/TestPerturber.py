import sys, inspect, warnings
import numpy as np, pandas as pd
from copy import deepcopy
from debug import get_module_name as gmn
from general_data_ops import get_shape as gs
from general_list_ops import list_select as ls
from data_validation import validate_user_input as vui, arg_kwarg_validater as akv
from ML_PACKAGE._data_validation import list_dict_validater as ldv
from openpyxl import Workbook
from MLObjects.SupportObjects import master_support_object_dict as msod
from MLObjects import MLObject as mlo
from ML_PACKAGE.GENERIC_PRINT import DictMenuPrint as dmp
from read_write_file.generate_full_filename import base_path_select as bps, filename_enter as fe



# init HOLDS ORIGINAL TEST_SWNL. CALL perturber TO RETURN SWNL WITH SPECIFIED col_idx PERTURBED.


# menu_options()
# dump_to_file()
# menu()
# set_perturbation_type()
# display_results_table()
# update_results_table()
# sort_results_table()
# build_empty_results_table()
# reset_test_data_to_original()
# lock_perturbation_on_columns()    # 'lock individual columns'
# unlock_perturbation_on_columns()  # 'unlock individual columns'
# unlock_all_columns()              # 'unlock all columns'
# perturber()                       # use this wrapper on core_perturber() to conceal all the yield/next jive that is happening in there
# core_perturber()                  # returns TEST_SWNL with col_idx replaced with a perturbed column, either generated on the fly here or passed.




class TestPerturber:

    def __init__(self, TEST_SWNL, DATA_HEADER_OR_FULL_DATA_SUPOBJ, data_run_orientation, target_run_orientation,
                 wb=None, bypass_validation=None):

        self.this_module = gmn.get_module_name(str(sys.modules[__name__]))
        fxn = inspect.stack()[0][3]

        self.bypass_validation = akv.arg_kwarg_validater(bypass_validation, 'bypass_validation', [True, False, None],
                                                         self.this_module, fxn, return_if_none=False)

        self.TEST_SWNL = list(TEST_SWNL)

        try:
            self.TEST_SWNL_BACKUP = [deepcopy(_) if isinstance(_, dict) else _.copy() for _ in self.TEST_SWNL]
        except:
            print(f'\n\033[91m')
            print(f'\n*** MAKE TEST_SWNL_BACKUP EXCEPTED ***')
            print(f'\nTEST DATA:'); print(self.TEST_SWNL[0])
            print(f'\nTEST_TARGET:'); print(self.TEST_SWNL[1])
            print(f'\nTEST_REFVECS:'); print(self.TEST_SWNL[2])

            raise Exception(f'*** MAKE TEST_SWNL_BACKUP EXCEPTED ***')

        DATA_HEADER_OR_FULL_DATA_SUPOBJ = \
            ldv.list_dict_validater(DATA_HEADER_OR_FULL_DATA_SUPOBJ, 'DATA_HEADER')[1]

        supobj_len = len(DATA_HEADER_OR_FULL_DATA_SUPOBJ)
        if supobj_len == 1:  self.DATA_HEADER = DATA_HEADER_OR_FULL_DATA_SUPOBJ   # IS HEADER ALREADY
        elif supobj_len == len(msod.master_support_object_dict()):   # IS FULL SUPOBJ
            self.DATA_HEADER = DATA_HEADER_OR_FULL_DATA_SUPOBJ[msod.QUICK_POSN_DICT()["HEADER"]].reshape((1,-1))
        else: # IS INVALIDLY SIZED SUPOBJ
            raise Exception(f'*** TestPerturber.init() >>> DATA_HEADER_OR_FULL_DATA_SUPOBJ HAS INVALID NUM ROWS ({supobj_len}). '
                            f'Must be 1 or {len(msod.master_support_object_dict())}')


        self.data_run_orientation = akv.arg_kwarg_validater(data_run_orientation, 'data_run_orientation', ['ROW', 'COLUMN'],
                                                            self.this_module, fxn)

        self.target_run_orientation = akv.arg_kwarg_validater(target_run_orientation, 'target_run_orientation', ['ROW', 'COLUMN'],
                                                            self.this_module, fxn)

        self.test_data_rows, self.test_data_cols = gs.get_shape('TEST_DATA', self.TEST_SWNL[0], self.data_run_orientation)



        self.WipTestDataClass = mlo.MLObject(
                                             self.TEST_SWNL[0],
                                             self.data_run_orientation,
                                             name='WIP_TEST_DATA',
                                             return_orientation='AS_GIVEN',
                                             return_format='AS_GIVEN',
                                             bypass_validation=self.bypass_validation,
                                             calling_module=self.this_module,
                                             calling_fxn=fxn
        )



        # BEAR
        # PERTURBATION METHOD
        #  I) ALL ZEROS
        # II) RANDOM NUMBERS
            # A) PERCENT TO PERTURB
            # B) AMOUNT TO PERTURB

        self.COLUMNS_WITH_LOCKED_PERTURBATION = []

        self.INSERT_COLUMN_DICT = {
            'Z': np.zeros(self.test_data_rows, dtype=np.int8)
        }

        self.original_baseline_cost = 0
        self.baseline_cost = 0

        self.perturbation_type = 'Z'

        self.RESULTS_TABLE = None
        self.build_empty_results_table()

        self.WB = Workbook() if wb is None else wb
        self.MENU_OPTIONS = None
        self.menu_options()


    # END init ##############################################################################################################
    #########################################################################################################################
    #########################################################################################################################


    def menu_options(self):
        self.MENU_OPTIONS = {

            'a': 'exit perturber',
            'b': 'build empty RESULTS TABLE',
            'c': 'set perturbation type',
            'd': 'display RESULTS TABLE',
            'f': 'dump RESULTS TABLE to file',
            'l': 'lock individual columns',
            'p': 'unlock individual columns',
            'q': 'unlock all columns',
            'r': 'reset TEST DATA to original',
            's': 'sort RESULTS TABLE',
            't': 'run perturber'
        }




    def dump_to_file(self):

        if 'TEST_PERTURBER' in self.WB: pass
        else: self.WB.active.title = 'TEST_PERTURBER'

        basepath = bps.base_path_select()
        filename = fe.filename_wo_extension()
        ext = '.xlsx'

        # MODIFY RESULTS_TABLE TO HAVE BASELINE COST IN TOP ROW
        # CANNOT INSERT ROWS, SO DO GYMNASTICS TO INSERT BASELINE COSTS AS COLUMNS

        if not (self.baseline_cost==0 and self.original_baseline_cost==0):
            # IF baseline_cost & original_baseline_cost ARE AVAILABLE, PUT THEM AT THE TOP OF RESULTS FOR FILE DUMP ONLY
            # THESE ARE NOT GENERATED INSIDE THIS CLASS AND MUST BE SET BY ASSIGNMENT TO THESE ATTRS OF AN INSTANCE OF THIS CLASS
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    percent_change = round(100 * (self.baseline_cost - self.original_baseline_cost) / self.original_baseline_cost, 7)
            except:
                percent_change = '-'

            self.RESULTS_TABLE = self.RESULTS_TABLE.T
            self.RESULTS_TABLE.insert(0, 'CURRENT BASELINE COST', ['-', '-', self.baseline_cost, percent_change])
            self.RESULTS_TABLE.insert(0, 'ORIGINAL BASELINE COST', ['-', '-', self.original_baseline_cost, 0])
            self.RESULTS_TABLE = self.RESULTS_TABLE.T

            del percent_change

        self.RESULTS_TABLE.to_excel(
                                    excel_writer=basepath+filename+ext,
                                    sheet_name='PERTURBER RESULTS TABLE',
                                    header=True,
                                    index=True,
                                    float_format = '%.5f',
                                    startrow = 1,
                                    startcol = 1,
                                    merge_cells = False
                                    )

        # REMOVE BASELINE ROWS AFTER DUMP
        self.RESULTS_TABLE = self.RESULTS_TABLE.iloc[2:, :]

        del basepath, filename, ext

        self.WB = Workbook()


    def menu(self, allowed=None, disallowed=None, max_width=None):

        max_width = 140 if max_width is None else max_width

        allowed = '' if allowed is None else allowed
        disallowed = '' if disallowed is None else disallowed

        while True:

            if len(self.COLUMNS_WITH_LOCKED_PERTURBATION) == 0:
                disallowed += 'P' if 'P' not in disallowed else ''
                disallowed += 'Q' if 'Q' not in disallowed else ''
            else:
                disallowed = "".join([_ for _ in disallowed if _ not in 'PQ'])

            print()
            MenuClass = dmp.DictMenuPrint(self.MENU_OPTIONS, disp_len=max_width, allowed=allowed, disallowed=disallowed)
            menu_select = MenuClass.select(f'Select letter')

            if menu_select == 'A':             # 'accept & exit menu'
                return menu_select
            elif menu_select == 'B':           # 'build empty RESULTS TABLE'
                self.build_empty_results_table()
            elif menu_select == 'C':           # 'set perturbation type'
                self.set_perturbation_type()
            elif menu_select == 'D':           # 'display RESULTS TABLE'
                self.display_results_table()
            elif menu_select == 'F':
                self.dump_to_file()
            elif menu_select == 'L':           # 'lock individual columns'
                self.lock_perturbation_on_columns()
            elif menu_select == 'P':           # 'unlock individual columns'
                self.unlock_perturbation_on_columns()
            elif menu_select == 'Q':           # 'unlock all columns'
                self.unlock_all_columns()
            elif menu_select == 'R':           # 'reset TEST DATA to original'
                self.reset_test_data_to_original()
            elif menu_select == 'S':           # 'sort RESULTS TABLE'
                self.sort_results_table()
            elif menu_select == 'T':           # 'run perturber'
                return menu_select

    def set_perturbation_type(self):

        PERTURBATION_TYPES = {
                                'z': 'all zeros'
        }

        OPTIONS = list(PERTURBATION_TYPES.values())
        KEYS = "".join(list(PERTURBATION_TYPES.keys())).upper()

        self.perturbation_type = KEYS[ls.list_single_select(OPTIONS, f'Select perturbation method > ', 'idx')[0]]

        del PERTURBATION_TYPES, OPTIONS, KEYS


    def display_results_table(self):
        print(self.RESULTS_TABLE)
        print(f'\nOriginal baseline cost = {self.original_baseline_cost}')
        print(f'\nCurrent baseline cost = {self.baseline_cost}\n')


    def update_results_table(self, col_idx, wip_cost):

        column_name = self.DATA_HEADER[0][col_idx]

        self.RESULTS_TABLE.loc[column_name, 'COST'] = wip_cost

        if self.baseline_cost == 0: pass
        else:
            self.RESULTS_TABLE.loc[column_name, '% CHANGE'] = \
                100 * round((wip_cost - self.baseline_cost) / self.baseline_cost, 7)

        del column_name


    def sort_results_table(self, sort_values=None, ascending=None):

        fxn = inspect.stack()[0][3]

        if (not sort_values is None) + (not ascending is None) not in [0, 2]:
            raise Exception(f'{self.this_module}.{fxn}() >>> EITHER NO OR ALL KWARGS MUST BE PASSED')

        if not sort_values is None:
            if not isinstance(sort_values, str):
                raise Exception(f'{self.this_module}.{fxn}() >>> sort_values MUST BE "C", "I", OR "S"')
            else: sort_values = sort_values.upper()

        if not ascending is None and not isinstance(ascending, bool):
            raise Exception(f'{self.this_module}.{fxn}() >>> ascending MUST BE BOOL')

        # IF NO KWARGS WERE PASSED, GIVE USER MENUS
        if not (sort_values is None and ascending is None):
            _column = sort_values; _order = ascending
        else:
            _column = vui.validate_user_str(f'Sort by column name(c) column index (i) or cost(s) > ', 'CIS')
            _order = {'A': True, 'D': False}[vui.validate_user_str(f'Sort ascending(a) or descending(d) > ', 'AD')]

        # 6/25/23 "-" CAN BE IN DF AND WILL BE IN DF IF ANY COLUMNS WERE LOCKED IN PERTURBED STATE FOR PERTURBER TEST.
        # CAUSES sort_values() TO EXCEPT, REGARDLESS OF HOW dtype OR astype FOR THE WHOLE DF OR COLUMN IS. TO SORT THIS
        # WITHOUT MODIFYING ACTUAL DF, PULL OUT PERTINENT COLUMNS, CONVERT W to_numeric(). TO ALWAYS KEEP NaNs AT THE
        # BOTTOM OF THE TABLE REGARDLESS OF SORT ORDER, USE fillna() TO FORCE IN VALUES TO MANIPULATE SORT. USE
        # np.argsort() TO SORT THE CONVERTED/MANIPULATED COLUMNS INTO A MASK, THEN MASK THE ORIGINAL DF WITH IT

        if _column == 'S': ACTV_COL = self.RESULTS_TABLE['COST']
        elif _column == 'C': self.RESULTS_TABLE.sort_index(ascending=_order, inplace=True)     # COLUMN NAME IS DF INDEX
        elif _column == 'I': ACTV_COL = self.RESULTS_TABLE['COL IDX']

        if _column in ['S', 'I']:
            ARGSORT = np.argsort(pd.to_numeric(ACTV_COL, errors="coerce").fillna(-np.inf if _order is False else np.inf))
            if _order is False: ARGSORT = np.flip(ARGSORT)
            self.RESULTS_TABLE = self.RESULTS_TABLE.iloc[ARGSORT, :]
            del ACTV_COL, ARGSORT

        del _column, _order


    def build_empty_results_table(self):

        pd.set_option(
                      'display.multi_sparse', False,
                      'display.colheader_justify', 'center',
                      'display.max_columns', None,
                      'display.max_rows', None,
                      'display.width', 140,
                      'display.max_colwidth', 50
                      )

        pd.options.display.float_format = '{:,.5f}'.format

        RESULTS_HEADER = ['COL IDX', 'STATE', 'COST', '% CHANGE']

        self.RESULTS_TABLE = pd.DataFrame(columns=RESULTS_HEADER, index=self.DATA_HEADER[0], dtype=np.float64).fillna('-')

        self.RESULTS_TABLE['COL IDX'] = np.arange(0, len(self.RESULTS_TABLE), dtype=np.int32)

        del RESULTS_HEADER

        return self.RESULTS_TABLE


    def reset_test_data_to_original(self):

        _ = self.TEST_SWNL_BACKUP[0]
        self.WipTestDataClass.OBJECT = deepcopy(_) if isinstance(_, dict) else _.copy()

        del _


    def lock_perturbation_on_columns(self, IDXS_TO_LOCK=None):                 # 'lock individual columns'

        while True:
            # IF IDXS_TO_LOCK WAS PASSED ###################################################################################
            if not IDXS_TO_LOCK is None:
                self.COLUMNS_WITH_LOCKED_PERTURBATION = list(set(IDXS_TO_LOCK + self.COLUMNS_WITH_LOCKED_PERTURBATION))
                del IDXS_TO_LOCK

                self.RESULTS_TABLE.loc[self.DATA_HEADER[0][self.COLUMNS_WITH_LOCKED_PERTURBATION], 'STATE'] = 'LOCKED ON'

                for lock_idx in self.COLUMNS_WITH_LOCKED_PERTURBATION:
                    self.WipTestDataClass.delete_columns([lock_idx])
                    self.WipTestDataClass.insert_column(lock_idx, self.INSERT_COLUMN_DICT[self.perturbation_type],
                                                        insert_orientation='COLUMN')

                break
            # END IF IDXS_TO_LOCK WAS PASSED ################################################################################

            # vvvv IF IDXS_TO_LOCK WAS NOT PASSED, USER MUST SELECT COLUMNS TO LOCK vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
            print(self.RESULTS_TABLE)

            # BUILD A MODIFIED LIST OF COLUMNS AND INFO FOR DISPLAY IN list_select() #########################################
            DUMMY_HEADER = self.RESULTS_TABLE.index.to_numpy().reshape((1, -1))[0]
            DISPLAY = []
            for idx in range(len(self.RESULTS_TABLE)):
                hdr_text = DUMMY_HEADER[idx][:48].ljust(50)
                # IF iloc VALUE IS NOT A NUM (LIKE "-") WILL BLOWUP, SO IF NOT A NUM, JUST LEAVE AS IS UNFORMATTED
                try: score_text = f'{self.RESULTS_TABLE.iloc[idx,2]: ,.5f}'
                except: score_text = self.RESULTS_TABLE.iloc[idx,2]
                try: percent_text = f'{self.RESULTS_TABLE.iloc[idx,3]: ,.5f}'
                except: percent_text = self.RESULTS_TABLE.iloc[idx,3]
                status_text = self.RESULTS_TABLE.iloc[idx, 1]
                DISPLAY.append(f'{hdr_text.ljust(50)}{status_text.ljust(20)}{score_text.ljust(15)}{percent_text.ljust(15)}')

            del DUMMY_HEADER, hdr_text, score_text, percent_text
            # END BUILD A MODIFIED LIST OF COLUMNS AND INFO FOR DISPLAY IN list_select() ######################################

            if vui.validate_user_str(f'\nCustom select columns(c) or select all columns after a certain column(s) > ', 'CS') == 'C':
                RESULTS_TABLE_IDX_HOLDER = ls.list_custom_select(DISPLAY, 'idx')
            else:
                [print(f'{_}) {__}') for _,__ in enumerate(DISPLAY, 1)]
                cutoff = vui.validate_user_int(f'\nEnter column number after which to chop (not inclusive) > ', min=1, max=self.test_data_cols)
                RESULTS_TABLE_IDX_HOLDER = list(range(cutoff, self.test_data_cols))

            # TRANSLATE THE idxs IN HOLDER (WHICH ARE idxed TO THE ORDER IN RESULT_TABLE) TO THE CORRESPONDING INDEX IN
            # DATA_HEADER, WHICH IS ACTUAL ORDER OF THE COLUMNS IN TRAIN/DEV/TEST DATA
            # 1) GET COLUMN NAMES OUT OF RESULT_TABLE index USING HOLDER INTO A NP ARRAY
            # 2) CREATE A DICT OF DATA_HEADER AND RESPECTIVE idx POSNS
            # 3) MAP DICT INTO COLUMN NAMES VECTOR TO GET ORIGINAL idx

            COLUMN_NAMES = self.RESULTS_TABLE.index.to_numpy().reshape((1,-1))[0]
            ORIG_IDX_DICT = dict((zip(self.DATA_HEADER[0], range(len(self.DATA_HEADER[0])))))

            ORIG_IDX_HOLDER = list(map(lambda x: ORIG_IDX_DICT[x], COLUMN_NAMES[RESULTS_TABLE_IDX_HOLDER]))

            del COLUMN_NAMES, ORIG_IDX_DICT

            print(f'\nUser selected to lock perturbation ON for the following columns:')
            [print(_) for _ in self.DATA_HEADER[0][ORIG_IDX_HOLDER]]
            print()

            del DISPLAY

            __ =  vui.validate_user_str(f'Accept (y/n) Abort (a) > ', 'YNA')

            if __ == 'Y':

                self.COLUMNS_WITH_LOCKED_PERTURBATION = list(set(self.COLUMNS_WITH_LOCKED_PERTURBATION + ORIG_IDX_HOLDER))
                del ORIG_IDX_HOLDER

                self.RESULTS_TABLE.iloc[RESULTS_TABLE_IDX_HOLDER, 1] = 'LOCKED ON'; del RESULTS_TABLE_IDX_HOLDER

                for lock_idx in self.COLUMNS_WITH_LOCKED_PERTURBATION:
                    self.WipTestDataClass.delete_columns([lock_idx])
                    self.WipTestDataClass.insert_column(lock_idx, self.INSERT_COLUMN_DICT[self.perturbation_type],
                                                        insert_orientation='COLUMN')
                break

            elif __ == 'N': continue

            elif __ == 'A': break

            # ^^^^ IF IDXS_TO_LOCK WAS NOT PASSED, USER MUST SELECT COLUMNS TO LOCK ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^


    def unlock_perturbation_on_columns(self):           # 'unlock individual columns'

        IDXS_TO_UNLOCK = ls.list_custom_select(self.DATA_HEADER[0][self.COLUMNS_WITH_LOCKED_PERTURBATION], 'idx')
        IDXS_TO_UNLOCK = np.array(self.COLUMNS_WITH_LOCKED_PERTURBATION)[..., IDXS_TO_UNLOCK]

        self.COLUMNS_WITH_LOCKED_PERTURBATION = [_ for _ in self.COLUMNS_WITH_LOCKED_PERTURBATION if _ not in IDXS_TO_UNLOCK]

        self.RESULTS_TABLE['STATE'] = '-'

        self.reset_test_data_to_original()

        self.lock_perturbation_on_columns(IDXS_TO_LOCK=self.COLUMNS_WITH_LOCKED_PERTURBATION)



    def unlock_all_columns(self):                   # 'unlock all columns'
        self.reset_test_data_to_original()
        self.COLUMNS_WITH_LOCKED_PERTURBATION = []
        self.RESULTS_TABLE['STATE'] = '-'


    def perturber(self, col_idx, INSERT_COLUMN=None):
        # USE THIS WRAPPER ON core_perturber() TO CONCEAL ALL THE yield/next JIVE THAT IS HAPPENING IN THERE

        PERTUBER_INSTANCE = self.core_perturber(col_idx, INSERT_COLUMN=INSERT_COLUMN)

        WIP_TEST_DATA = next(PERTUBER_INSTANCE)   # CREATE A BACKUP OF ORIGINAL COLUMN AND SWAP IN PERTURBED COLUMN IN TEST_DATA AND RETURN
        next(PERTUBER_INSTANCE)   # SWAP OUT PERTURBED COLUMN AND PUT ORIGINAL COLUMN BACK IN
        del PERTUBER_INSTANCE
        return WIP_TEST_DATA


    def core_perturber(self, col_idx, INSERT_COLUMN=None):

        # RETURNS TEST_SWNL WITH col_idx REPLACED WITH A PERTURBED COLUMN, EITHER GENERATED ON THE FLY HERE OR PASSED.

        fxn = inspect.stack()[0][3]

        BACKUP_COLUMN = self.WipTestDataClass.return_columns([col_idx], return_orientation='AS_GIVEN', return_format='AS_GIVEN')

        self.WipTestDataClass.delete_columns([col_idx])

        # BEAR HAVE TO FIGURE OUT HOW TO DO DIFFERENT PERTURBATION METHODS
        if not INSERT_COLUMN is None: pass     # IF INSERT_COLUMN IS PASSED AS KWARG
        else: INSERT_COLUMN = self.INSERT_COLUMN_DICT[self.perturbation_type]

        self.WipTestDataClass.insert_column(col_idx, INSERT_COLUMN, 'COLUMN')

        yield self.WipTestDataClass.OBJECT
        self.WipTestDataClass.delete_columns([col_idx])

        self.WipTestDataClass.insert_column(col_idx, BACKUP_COLUMN, insert_orientation=self.data_run_orientation)

        del BACKUP_COLUMN

        yield None


























if __name__ == '__main__':

    from MLObjects.TestObjectCreators.SXNL import CreateSXNL as cs
    from general_data_ops import TrainDevTestSplit as tdts


    DATA = pd.read_csv(r'C:\Users\Bill\Documents\WORK STUFF\RESUME\1 - OTHER\SRP\beer_reviews.csv',
                       nrows=25000,
                       header=0
    ).dropna(axis=0)

    DATA = DATA[DATA.keys()[[3 ,4 ,5, 7, 8, 9, 11]]]    # ,7


    TARGET = DATA['review_overall']
    TARGET_HEADER = [['review_overall']]
    TARGET = TARGET.to_numpy().reshape((1,-1))


    pd.set_option(
                  'display.multi_sparse', False,
                  'display.colheader_justify', 'center',
                  'display.max_columns', None,
                  'display.max_rows', None,
                  'display.width', 140,
                  'display.max_colwidth', 50
                  )


    DATA = DATA.drop(columns=['review_overall'])

    RAW_DATA = DATA.copy()
    RAW_DATA_HEADER = np.fromiter(RAW_DATA.keys(), dtype='<U50').reshape((1,-1))
    RAW_DATA = RAW_DATA.to_numpy()

    # FORMATS & ORIENTS TO RETURN FROM CreateSXNL
    _format = 'ARRAY'
    _orient = 'COLUMN'


    SXNLClass = cs.CreateSXNL(
                                rows=None,
                                bypass_validation=True,
                                ####################################################################################
                                # DATA #############################################################################
                                data_return_format=_format,
                                data_return_orientation=_orient,
                                DATA_OBJECT=RAW_DATA,
                                DATA_OBJECT_HEADER=RAW_DATA_HEADER,
                                DATA_FULL_SUPOBJ_OR_SINGLE_MDTYPES=None,
                                data_override_sup_obj=False,
                                # CREATE FROM GIVEN ONLY ###############################################
                                data_given_orientation='ROW',
                                # END CREATE FROM GIVEN ONLY #############################################
                                # CREATE FROM SCRATCH_ONLY ################################
                                data_columns=None,
                                DATA_BUILD_FROM_MOD_DTYPES=None,
                                DATA_NUMBER_OF_CATEGORIES=None,
                                DATA_MIN_VALUES=None,
                                DATA_MAX_VALUES=None,
                                DATA_SPARSITIES=None,
                                DATA_WORD_COUNT=None,
                                DATA_POOL_SIZE=None,
                                # END DATA ##############################################################
                                #########################################################################

                                #########################################################################
                                # TARGET ################################################################
                                target_return_format=_format,
                                target_return_orientation=_orient,
                                TARGET_OBJECT=TARGET,
                                TARGET_OBJECT_HEADER=TARGET_HEADER,
                                TARGET_FULL_SUPOBJ_OR_SINGLE_MDTYPES=None,
                                target_type='FLOAT',  # MUST BE 'BINARY','FLOAT', OR 'SOFTMAX'
                                target_override_sup_obj=False,
                                target_given_orientation='COLUMN',
                                # END CORE TARGET_ARGS ########################################################
                                # FLOAT AND BINARY
                                target_sparsity=None,
                                # FLOAT ONLY
                                target_build_from_mod_dtype=None,  # COULD BE FLOAT OR INT
                                target_min_value=None,
                                target_max_value=None,
                                # SOFTMAX ONLY
                                target_number_of_categories=None,
                                # END TARGET ##############################################################
                                ###########################################################################

                                ###########################################################################
                                # REFVECS #################################################################
                                refvecs_return_format=_format,  # IS ALWAYS ARRAY (WAS, CHANGED THIS 4/6/23)
                                refvecs_return_orientation=_orient,
                                REFVECS_OBJECT=None,
                                REFVECS_OBJECT_HEADER=None,
                                REFVECS_FULL_SUPOBJ_OR_SINGLE_MDTYPES=None,
                                REFVECS_BUILD_FROM_MOD_DTYPES='STR',
                                refvecs_override_sup_obj=False,
                                refvecs_given_orientation=None,
                                refvecs_columns=5,
                                REFVECS_NUMBER_OF_CATEGORIES=10,
                                REFVECS_MIN_VALUES=None,
                                REFVECS_MAX_VALUES=None,
                                REFVECS_SPARSITIES=None,
                                REFVECS_WORD_COUNT=None,
                                REFVECS_POOL_SIZE=None
                                # END REFVECS ##############################################################
                                ############################################################################
    )

    # BUILD SRNL ##################################################################################################################

    # SRNL = SXNLClass.SXNL.copy()
    # RAW_SUPOBJS = SXNLClass.SXNL_SUPPORT_OBJECTS.copy()
    # WORKING_KEEP = SXNLClass.SXNL[1][0].copy()

    # END BUILD SRNL ##################################################################################################################

    # EXPAND SRNL TO SWNL #############################################################################################################
    SXNLClass.expand_data(expand_as_sparse_dict=isinstance(SXNLClass.DATA, dict), auto_drop_rightmost_column=False)

    # END EXPAND SRNL TO SWNL #############################################################################################################

    # BUILD SWNLs ##################################################################################################################
    DATA_SWNL = SXNLClass.SXNL
    WORKING_SUPOBJS = SXNLClass.SXNL_SUPPORT_OBJECTS
    data_given_orientation = SXNLClass.data_current_orientation
    target_given_orientation = SXNLClass.target_current_orientation
    refvecs_given_orientation = SXNLClass.refvecs_current_orientation


    # DELETE LOWLIFE COLUMNS W FEW ENTRIES THAT R SCREWING UP XTX_INV #######################################################
    DataClass = mlo.MLObject(
                             DATA_SWNL[0],
                             _orient,
                             name='DATA',
                             return_orientation='AS_GIVEN',
                             return_format='AS_GIVEN',
                             bypass_validation=True,
                             calling_module='TestPertuber',
                             calling_fxn='tests'
    )

    print(f'\nDELETING COLUMNS WITH LESS THAN 10 ENTRIES...')
    for col_idx in range(DataClass._cols-1, -1, -1):
        if np.sum(DataClass.return_columns([col_idx], return_format='ARRAY', return_orientation='COLUMN')) < 10:
            DataClass.delete_columns([col_idx])
            print(f'DELETED {WORKING_SUPOBJS[0][0][col_idx]}')
            WORKING_SUPOBJS[0] = np.delete(WORKING_SUPOBJS[0], col_idx, axis=1)

    DATA_SWNL[0] = DataClass.OBJECT

    del DataClass
    # END DELETE LOWLIFE COLUMNS W FEW ENTRIES THAT R SCREWING UP XTX_INV #######################################################

    # END BUILD SWNLs ##################################################################################################################

    DATA_HEADER = WORKING_SUPOBJS[0][msod.QUICK_POSN_DICT()["HEADER"]].reshape((1,-1))


    # BUILD TRAIN / DEV / TEST SPLITS #################################################################################################
    ctr = 0
    while True:
        ctr += 1

        print(f'\nTrying to get invertible train split... try #{ctr}...')

        # SXNLClass.train_dev_test_split()
        TDTClass = tdts.TrainDevTestSplit(
                                          DATA=DATA_SWNL[0],
                                          TARGET=DATA_SWNL[1],
                                          REFVECS=DATA_SWNL[2],
                                          data_given_orientation=data_given_orientation,
                                          target_given_orientation=target_given_orientation,
                                          refvecs_given_orientation=refvecs_given_orientation,
                                          bypass_validation=True
        )

        TRAIN_SWNL, DEV_SWNL, TEST_SWNL = TDTClass.random(
                                                          dev_count=None,
                                                          dev_percent=20,
                                                          test_count=None,
                                                          test_percent=20
        )

        del TDTClass


        print(f'Trying to invert TRAIN DATA...')

        ObjectClass = mlo.MLObject(
                                     TRAIN_SWNL[0],
                                     _orient,
                                     name='TRAIN DATA',
                                     return_orientation='AS_GIVEN',
                                     return_format='AS_GIVEN',
                                     bypass_validation=True,
                                     calling_module='TestPertuber',
                                     calling_fxn='tests'
        )

        try: ObjectClass.return_XTX_INV(return_format='ARRAY'); print(f'Success.'); break
        except: print(f'Fail.')

        # END BUILD TRAIN / DEV / TEST SPLITS #################################################################################################

    del SXNLClass



    '''
    # DUMP DATA TO FILE #################################################################################################################


    for itr, name in enumerate(('TRAIN', "DEV", "TEST")):

        if name=='TRAIN': data = np.hstack((TRAIN_SWNL[0].transpose(), TRAIN_SWNL[1].transpose()))
        elif name == 'DEV': data = np.hstack((DEV_SWNL[0].transpose(), DEV_SWNL[1].transpose()))
        elif name == 'TEST': data = np.hstack((TEST_SWNL[0].transpose(), TEST_SWNL[1].transpose()))

        DF = pd.DataFrame(data=data, columns=np.hstack((DATA_HEADER[0], ['TARGET']))
                          )

        from pathlib import Path
        desktop_path = Path.home() / "Desktop"
        filename = desktop_path / f"{name}.xlsx"

        DF.to_excel(
            excel_writer=filename,
            sheet_name=name,
            header=True,
            index=False
        )

    del DF

    # END DUMP DATA TO FILE #################################################################################################################
    '''


    def train_coeffs(TRAIN_DATA, TRAIN_TARGET, data_orient):

        # FUDGE FOR OUTPUT / COST
        if data_orient=='ROW': XTX = np.matmul(TRAIN_DATA.transpose().astype(np.float64), TRAIN_DATA.astype(np.float64))
        elif data_orient == 'COLUMN': XTX = np.matmul(TRAIN_DATA.astype(np.float64), TRAIN_DATA.transpose().astype(np.float64))

        # BEAR PUT A LITTLE RIDGE INTO IT TO OVERCOME MULTICOLIN
        # XTX = XTX + 0 * np.identity(len(XTX))

        XTX_INV = np.linalg.inv(XTX); del XTX

        if data_orient=='ROW': XTX_AT = np.matmul(XTX_INV.astype(np.float64), TRAIN_DATA.transpose().astype(np.float64))
        elif data_orient == 'COLUMN': XTX_AT = np.matmul(XTX_INV.astype(np.float64), TRAIN_DATA.astype(np.float64))

        del XTX_INV

        WIP_TARGET = TRAIN_TARGET.reshape((-1, 1))

        COEFFS = np.matmul(XTX_AT.astype(np.float64), WIP_TARGET.astype(np.float64))


        return COEFFS


    def test_output_calc_algorithm(TEST_DATA, data_orient, COEFFS):


        if data_orient=='ROW': TEST_OUTPUT_VECTOR = np.matmul(TEST_DATA.astype(np.float64), COEFFS.astype(np.float64))
        elif data_orient == 'COLUMN': TEST_OUTPUT_VECTOR = np.matmul(TEST_DATA.transpose().astype(np.float64),
                                                                COEFFS.astype(np.float64))

        return TEST_OUTPUT_VECTOR


    def test_data_error_algorithm(TEST_DATA, TEST_TARGET, data_orient, COEFFS):

        TEST_TARGET = TEST_TARGET.reshape((1,-1))

        TEST_OUTPUT_VECTOR = test_output_calc_algorithm(TEST_DATA, data_orient, COEFFS).reshape((1,-1))

        error = np.sum(np.power((TEST_TARGET - TEST_OUTPUT_VECTOR), 2))

        return error


    wb = Workbook()





    #####################################################################################################################
    #####################################################################################################################
    # START STUFF THAT WOULD GO INTO MLRunTemplate ######################################################################

    # GET COEFFS FROM TRAIN_DATA
    COEFFS = train_coeffs(*TRAIN_SWNL[:2], _orient)


    PerturberClass = TestPerturber(
                                    TEST_SWNL,
                                    WORKING_SUPOBJS[0],
                                    _orient,
                                    _orient,
                                    wb=wb,
                                    bypass_validation=False
    )

    test_data_cols = gs.get_shape('TEST_DATA', TEST_SWNL[0], _orient)[1]

    PerturberClass.original_baseline_cost = test_data_error_algorithm(
                                                                        PerturberClass.WipTestDataClass.OBJECT,
                                                                        PerturberClass.TEST_SWNL[1],
                                                                        _orient,
                                                                        COEFFS
    )

    while True: #########################################################################################################

        # CALCULATE BASELINE ERROR
        PerturberClass.baseline_cost = test_data_error_algorithm(
                                                                 PerturberClass.WipTestDataClass.OBJECT,
                                                                 PerturberClass.TEST_SWNL[1],
                                                                 _orient,
                                                                 COEFFS
        )

        printer = {'Y': True, 'N': False}[vui.validate_user_str(f'PRINT TEST_DF DURING RUN TO WATCH THE COLUMNS CHANGE? (y/n) > ', 'YN')]

        for col_idx in range(test_data_cols):

            print(f'\nRunning column {col_idx + 1} of {test_data_cols}...')

            if col_idx in PerturberClass.COLUMNS_WITH_LOCKED_PERTURBATION:
                PerturberClass.RESULTS_TABLE.loc[DATA_HEADER[0][col_idx], ('COST', '% CHANGE')] = ['-', '-']
                continue

            WIP_TEST_DATA = PerturberClass.perturber(col_idx)

            if printer:
                print(pd.DataFrame(data=WIP_TEST_DATA.transpose(), columns=DATA_HEADER[0]))
                __ = input(f'PAUSED TO LOOK AT TEST DATA, HIT ENTER > ')


            # OUTPUT_VECTOR IS CALCULATED WITHIN test_data_error_algorithm
            wip_cost = test_data_error_algorithm(WIP_TEST_DATA, TEST_SWNL[1], _orient, COEFFS)

            PerturberClass.update_results_table(col_idx, wip_cost)

        print(PerturberClass.RESULTS_TABLE)
        print()
        print(f'\nOriginal baseline cost = {PerturberClass.original_baseline_cost}')
        print(f'\nCurrent baseline cost = {PerturberClass.baseline_cost}')

        menu_select = PerturberClass.menu()

        if menu_select == 'A': break
        elif menu_select == 'T': continue

    # END while ###########################################################################################################

    RESULTS_TABLE = PerturberClass.RESULTS_TABLE

    del PerturberClass, WIP_TEST_DATA

    # END STUFF THAT WOULD GO INTO MLRunTemplate ######################################################################
    #####################################################################################################################
    #####################################################################################################################


























