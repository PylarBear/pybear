from ML_PACKAGE.openpyxl_shorthand import openpyxl_write as ow
from openpyxl.utils.dataframe import dataframe_to_rows



def gmlr_train_results_dump(wb, TRAIN_RESULTS):
    # TRAIN_RESULTS IS DF


    def custom_write(sheet, row, column, value, horiz, vert, bold):
        ow.openpyxl_write(wb, sheet, row, column, value, horiz=horiz, vert=vert, bold=bold)


    wb.create_sheet('GMLR TRAIN RESULTS')

    row_counter = 1

    custom_write('GMLR TRAIN RESULTS', row_counter, 1, 'GMLR TRAIN RESULTS', 'left', 'center', True)

    row_counter += 2

    # FILL IN DATA
    for grid_row_idx, row in enumerate(dataframe_to_rows(TRAIN_RESULTS, index=False), row_counter):
        for grid_col_idx, value in enumerate(row, 3):
            if grid_row_idx - row_counter <= 1: bold = True
            else: bold = False
            custom_write('GMLR TRAIN RESULTS', grid_row_idx, grid_col_idx, value, 'center', 'center', bold)

    for grid_row_idx, value in enumerate(TRAIN_RESULTS.index, row_counter+2):
        custom_write('GMLR TRAIN RESULTS', grid_row_idx, 2, value, 'left', 'center', False)



    return wb






if __name__ == '__main__':

    import numpy as n, pandas as p
    from openpyxl import Workbook
    from read_write_file.generate_full_filename import base_path_select as bps

    wb = Workbook()

    TEST_DATA = n.random.randint(0,10,(3,10))
    DF = p.DataFrame(data=TEST_DATA.transpose(), columns=['A','B','C'])


    wb.active.title = 'BLANK'
    wb['BLANK'].cell(1,1).value = 'INTENTIONALLY LEFT BLANK'

    wb = gmlr_train_results_dump(wb, TEST_DATA)
    base_path = bps.base_path_select()
    wb.save(base_path + r'test_file_dump_xyz.xlsx')
