from ML_PACKAGE.openpyxl_shorthand import openpyxl_write as ow
from openpyxl.utils.dataframe import dataframe_to_rows



def mi_train_results_dump(wb, TRAIN_RESULTS):
    # TRAIN_RESULTS IS DF


    def custom_write(sheet, row, column, value, horiz, vert, bold):
        ow.openpyxl_write(wb, sheet, row, column, value, horiz=horiz, vert=vert, bold=bold)


    wb.create_sheet('MI TRAIN RESULTS')

    row_counter = 1

    custom_write('MI TRAIN RESULTS', row_counter, 1, 'MI TRAIN RESULTS', 'left', 'center', True)

    row_counter += 2

    # FILL IN DATA
    for grid_row_idx, row in enumerate(dataframe_to_rows(TRAIN_RESULTS, index=False), row_counter):
        for grid_col_idx, value in enumerate(row, 3):
            if grid_row_idx - row_counter <= 1: bold = True
            else: bold = False

            custom_write('MI TRAIN RESULTS', grid_row_idx, grid_col_idx, value, 'center', 'center', bold)

    for grid_row_idx, value in enumerate(TRAIN_RESULTS.index, row_counter+2):
        custom_write('MI TRAIN RESULTS', grid_row_idx, 2, value, 'left', 'center', False)

    return wb



if __name__ == '__main__':
    import numpy as n
    from openpyxl import Workbook
    from read_write_file.generate_full_filename import base_path_select as bps
    from ML_PACKAGE.MUTUAL_INFORMATION import build_empty_mi_train_results as bemtr
    from MLObjects.TestObjectCreators import test_header as th

    _cols = 10


    wb = Workbook()

    DATA = n.random.uniform(0,10,(_cols,5))
    DATA_HEADER = th.test_header(_cols)
    DF = bemtr.build_empty_mi_train_results(DATA_HEADER[0])

    DF.iloc[:, :5] = DATA
    print(DF)

    wb.active.title = 'BLANK'
    wb['BLANK'].cell(1,1).value = 'INTENTIONALLY LEFT BLANK'


    # rows = dataframe_to_rows(DF, index=False)
    # for row_idx, row in enumerate(rows, 1):
    #     for col_idx, value in enumerate(row, 1):
    #         wb['TRAIN_RESULTS'].cell(row=row_idx, column=col_idx, value=value)
    #         if row_idx == 1:
    #             custom_write('TRAIN_RESULTS', row_idx, col_idx, value, 'center', 'center', True)
    #         else:
    #             custom_write('TRAIN_RESULTS', row_idx, col_idx, value, 'center', 'center', False)
    #         # wb['TRAIN_RESULTS'].cell(row=row_idx, column=col_idx, value=value)


    wb = mi_train_results_dump(wb, DF)
    basepath = bps.base_path_select()
    wb.save(basepath + r'mi_train_results_dump_test.xlsx')
    print(f'Done.')


