from general_list_ops import list_select as ls
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from read_write_file.generate_full_filename import base_path_select as bps, filename_enter as fe


#CALLED BY SVMRun
def boundary_excel_dump(SUPPORT_VECTORS, SUPPORT_TARGETS, SUPPORT_ALPHAS, b):

    # CREATE FILE & SETUP SHEET
    base_path = bps.base_path_select()

    file_name = fe.filename_wo_extension()

    full_path = base_path + file_name + '.xlsx'

    wb = Workbook()

    wb.active.title = 'SUPPORT VECTORS'
    wb.create_sheet('SUPPORT TARGETS')
    wb.create_sheet('SUPPORT ALPHAS')
    for row_idx in range(len(SUPPORT_VECTORS)):
        for col_idx in range(len(SUPPORT_VECTORS[row_idx])):
            wb['SUPPORT VECTORS'].cell(row_idx+1, col_idx+1).value = SUPPORT_VECTORS[row_idx][col_idx]
            wb['SUPPORT TARGETS'].cell(row_idx + 1, 1).value = SUPPORT_TARGETS[row_idx]
            wb['SUPPORT ALPHAS'].cell(row_idx+1, 1).value = SUPPORT_ALPHAS[row_idx]

    wb.create_sheet('b')
    wb['b'].cell(1,1).value = b

    wb.save(full_path)

    print(f'\nALPHAS saved to {base_path+file_name} successfully.\n')













