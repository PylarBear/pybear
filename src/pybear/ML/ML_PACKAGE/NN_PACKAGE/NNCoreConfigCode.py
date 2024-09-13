import sys, inspect
import numpy as np
import sparse_dict as sd
import openpyxl as xl
from debug import get_module_name as gmn
from data_validation import validate_user_input as vui, arg_kwarg_validater as akv
from general_data_ops import get_shape as gs
from general_list_ops import list_select as ls, manual_num_seq_list_fill as mnslf
from ML_PACKAGE.NN_PACKAGE import SetLearningRate as lrs
from ML_PACKAGE.NN_PACKAGE.gd_run import error_calc as ec, output_vector_calc as ovc
from ML_PACKAGE.GENERIC_PRINT import print_object_create_success as pocs
from ML_PACKAGE.NN_PACKAGE.link_functions import link_fxns as lf, select_link_fxn as slf
from ML_PACKAGE.NN_PACKAGE.print_results import print_nn_config as pnnc

from read_write_file.generate_full_filename import base_path_select as bps, filename_enter as fe



# return_fxn()                            module return
# activation_constant_fxn()               set number to append to each node's output vector
# batch_size_vector_config()              set up BATCH_SIZE
# gd_convergence_methods()                convergence methods for gradient descent methods
# gd_convergence_str()                    allowed menu selections for gd convergence
# gd_momentum()                           set momentum
# print_learning_rates()                  display learning rates during config
# array_of_nodes_build()                  build array of nodes after all build params are set
# node_seed_fxn()                         set node seeds
# config()                                exe


class NNCoreConfigCode:
    def __init__(self, nn_config, DATA, TARGET_VECTOR, data_run_orientation, target_run_orientation, ARRAY_OF_NODES,
                 NEURONS, nodes, node_seed, activation_constant, aon_base_path, aon_filename, cost_fxn, SELECT_LINK_FXN,
                 LIST_OF_NN_ELEMENTS, OUTPUT_VECTOR, batch_method, BATCH_SIZE, gd_method, conv_method, lr_method,
                 LEARNING_RATE, momentum_weight, rglztn_type, rglztn_fctr, conv_kill, pct_change, conv_end_method,
                 gd_iterations, non_neg_coeffs, allow_summary_print, summary_print_interval, iteration):

        self.this_module = gmn.get_module_name(str(sys.modules[__name__]))
        fxn = inspect.stack()[0][3]


        self.nn_config = nn_config
        self.DATA = DATA
        self.TARGET_VECTOR = TARGET_VECTOR
        self.data_run_orientation = akv.arg_kwarg_validater(data_run_orientation, 'data_run_orientation', ['ROW', 'COLUMN'], self.this_module, fxn)
        self.target_run_orientation = akv.arg_kwarg_validater(target_run_orientation, 'target_run_orientation', ['ROW', 'COLUMN'], self.this_module, fxn)
        self.ARRAY_OF_NODES = ARRAY_OF_NODES
        self.NEURONS = NEURONS
        self.nodes = nodes
        self.node_seed = node_seed
        self.activation_constant = activation_constant
        self.aon_base_path = aon_base_path
        self.aon_filename = aon_filename
        self.cost_fxn = cost_fxn.upper()
        self.SELECT_LINK_FXN = SELECT_LINK_FXN
        self.LIST_OF_NN_ELEMENTS = LIST_OF_NN_ELEMENTS
        self.OUTPUT_VECTOR = OUTPUT_VECTOR
        self.batch_method = batch_method
        self.BATCH_SIZE = BATCH_SIZE
        self.gd_method = gd_method
        self.conv_method = conv_method
        self.lr_method = lr_method
        self.LEARNING_RATE = LEARNING_RATE
        self.momentum_weight = momentum_weight
        self.rglztn_type = rglztn_type
        self.rglztn_fctr = rglztn_fctr
        self.conv_kill = conv_kill
        self.pct_change = pct_change
        self.conv_end_method = conv_end_method
        self.gd_iterations = gd_iterations
        self.non_neg_coeffs = non_neg_coeffs
        self.allow_summary_print = allow_summary_print
        self.summary_print_interval = summary_print_interval
        self.iteration = iteration

        # BEAR MAYBE THESE COME IN AS ARGS ONE DAY
        self.data_rows, self.data_cols = gs.get_shape('DATA', self.DATA, self.data_run_orientation)


        self.number_of_labels = gs.get_shape('TARGET', self.TARGET_VECTOR, self.target_run_orientation)[1]

        # HOLDERS
        self.new_error_start = float('inf')


    def return_fxn(self):
        return self.ARRAY_OF_NODES, self.NEURONS, self.nodes, self.node_seed, self.activation_constant, \
                self.aon_base_path, self.aon_filename, self.cost_fxn, self.SELECT_LINK_FXN, self.LIST_OF_NN_ELEMENTS, \
                self.batch_method, self.BATCH_SIZE, self.gd_method, self.conv_method, self.lr_method, self.LEARNING_RATE, \
                self.momentum_weight, self.rglztn_type, self.rglztn_fctr, self.conv_kill, self.pct_change, self.conv_end_method, \
                self.gd_iterations, self.non_neg_coeffs, self.allow_summary_print, self.summary_print_interval, self.iteration


    def activation_constant_fxn(self):
        # use a constant on each layers' output --- except examples going in (handled by BIG MATRIX config) and final output
        # append a "1" to each INTERMEDIATE_RESULT (ie, feed wTx+b into next layer, not (THETA)Tx
        return vui.validate_user_int(f'\nEnter activation constant (1 if used, else 0) > ', min=0, max=1)


    def gd_convergence_methods(self):
        # gd_convergence_methods()                convergence methods for gradient descent methods
        return ['GD(g)', 'RMSProp(r)', 'ADAM(a)', 'Newton(n)']


    def gd_convergence_str(self):
        # gd_convergence_str()                    allowed menu selections for gd convergence
        return 'GRAN'


    def batch_size_vector_config(self):

        print(f'\nLast BATCH_SIZE[:10] was {self.BATCH_SIZE[:10]}\n')


        if self.SELECT_LINK_FXN[-1].upper() == 'MULTI-OUT':
            print(f'Multi-out is being used, batch method must be BATCH')
            self.batch_method = 'B'
            self.BATCH_SIZE = np.full(self.gd_iterations, self.data_rows, dtype=np.int32)

        else:
            while True:
                self.batch_method = vui.validate_user_str('Select batch method - batch(b) mini-batch(m) stochastic(s) > ', 'BMS')

                if self.batch_method == 'B':      #     batch_method = 'BATCH'
                    print('BATCH method selected')
                    self.BATCH_SIZE = np.full(self.gd_iterations, self.data_rows, dtype=np.int32)

                elif self.batch_method == 'M':     #     batch_method = 'MINI-BATCH'
                    print(f'MINI-BATCH method selected')

                    # bsm = batch_size_method
                    bsm = vui.validate_user_str(f'Use constant(c), fxn(f), or variable(v) mini-batch size? > ', 'CFV')
                    if bsm == 'C':
                        batch_size = vui.validate_user_int(f'Enter mini-batch size (max is {self.data_rows}) > ', min=1, max=self.data_rows)
                        self.BATCH_SIZE = [batch_size for _ in range(self.gd_iterations)]

                    elif bsm == 'F':
                        print(f'\nFormula format is A * iteration# + C. All batches must be integers greater than zero.')
                        a = vui.validate_user_float(f'Enter A > ')
                        c = vui.validate_user_float(f'Enter C > ')
                        # PROCESS OUTPUT TO BE INTEGER, AT LEAST 1, AND AT MOST FULL BATCH SIZE
                        self.BATCH_SIZE = [int(max(1, min(a * x + c, self.data_rows))) for x in range(self.gd_iterations)]

                    elif bsm == 'V':
                        self.BATCH_SIZE = mnslf.manual_num_seq_list_fill('batch size', self.BATCH_SIZE, self.gd_iterations)

                    del bsm

                elif self.batch_method == 'S':    #  batch_method = 'STOCHASTIC'
                    print('STOCHASTIC method selected')
                    self.BATCH_SIZE = [int(1) for _ in range(self.gd_iterations)]

                # if self.batch_method in 'MS':
                print(f'\nBATCH_SIZE[:10] looks like {self.BATCH_SIZE[:10]}\n')

                if vui.validate_user_str('Accept batch size config? (y/n) > ', 'YN') == 'Y':
                    break


    def gd_momentum(self):

        print(f'\nMomentum formula is (1 - weight) * new gradient + weight * old gradient = momentum gradient\n')
        self.momentum_weight = vui.validate_user_float(f'Enter weight applied to previous gradient (0 < w < 1) > ', min=0, max=1)



    def print_learning_rates(self):
        LR_DUM = [[f'{_:.5g}' for _ in __] for __ in self.LEARNING_RATE]  # DUMMY FOR DISPLAY
        for node in range(len(self.LEARNING_RATE)):
            print(f'NODE {node}) {LR_DUM[node][:10]}')


    def array_of_nodes_build(self):

        ####################################################################################################################
        ####################################################################################################################
        # ARRAY OF NODES BUILD #############################################################################################

        ''' ********************************KEEP THIS FOR REFERENCE***********************************************************
        ARRAY_OF_NODES_ROWS = [    neurons1      ,   neurons2,      1     ]
        ARRAY_OF_NODES_COLS = [len(BIG_MATRIX[0]),   neurons1,   neurons2  ]

        CREATE NODES FILLED W SEED (CAN'T START WITH ALL ZEROS ANY SINGLE
        LEARNING RATE GETS MULTIPLIED OUT TO ZERO CUZ THE REMAINING NODES ARE ALL ZERO
        ****************************************************************************** '''

        # CREATE A NUMBER THAT MODIFIES THE NUMBER OF COLUMNS IN EACH NODE (EXCEPT FIRST & LAST) TO ACCOMMODATE
        # activation_constant, IF BEING USED
        if self.activation_constant == 0: col_adj = 0
        else: col_adj = 1

        # AT THIS POINT node_seed IS JUST A PLACEHOLDER, REAL node_seeds ARE FILLED AFTER

        self.ARRAY_OF_NODES.clear()
        for node in range(self.nodes):
            if node == 0 and self.nodes != 1:
                self.ARRAY_OF_NODES.append(
                    np.full(shape=[self.NEURONS[node], self.data_cols], fill_value=self.node_seed, dtype=float)
                )

            if node != 0 and node != self.nodes - 1:
                self.ARRAY_OF_NODES.append(
                    np.full(shape=[self.NEURONS[node], self.NEURONS[node - 1] + col_adj], fill_value=self.node_seed, dtype=float)
                )

            if node == self.nodes - 1 and self.nodes != 1:
                self.ARRAY_OF_NODES.append(
                    np.full(shape=[self.number_of_labels, self.NEURONS[node - 1] + col_adj], fill_value=self.node_seed, dtype=float)
                )
            if node == 0 and self.nodes == 1:
                self.ARRAY_OF_NODES.append(
                    np.full(shape=[self.number_of_labels, self.data_cols], fill_value=self.node_seed, dtype=float)
                )

        self.LIST_OF_NN_ELEMENTS = []
        for node in range(0, len(self.ARRAY_OF_NODES)):
            for list_item in range(len(self.ARRAY_OF_NODES[node])):
                for element in range(len(self.ARRAY_OF_NODES[node][list_item])):
                    self.LIST_OF_NN_ELEMENTS.append([node, list_item, element])

        # FILL AON WITH NODE SEED
        self.node_seed_fxn()

        pocs.print_object_create_success('', 'ARRAY OF NODES')

        self.iteration = 0

        return self.ARRAY_OF_NODES, self.LIST_OF_NN_ELEMENTS, self.iteration

        # END ARRAY OF NODES BUILD #########################################################################################
        ####################################################################################################################
        ####################################################################################################################


    def node_seed_fxn(self):

        SEED_OPTIONS = [f'Seed with a constant(c)', f'load from file(f)',
                        f'random numbers from standard normal dist(r)', 'Xavier-He method(x)']

        seed_str = 'CFRX'

        while True:

            seed_method = seed_str[ls.list_single_select(SEED_OPTIONS, f'Select NN seed method', 'idx')[0]]

            # THIS IS JUST FOR CARRYING THE FILENAME (OR LACK THEREOF) AROUND.... #THIS MUST STAY SEPARATE FROM THE
            # IF-ELIF CHAIN THAT COMES AFTER
            if seed_method != 'S':
                self.aon_base_path = 'None - '
                self.aon_filename = 'values not loaded from file.'

            if seed_method == 'F':

                while True:
                    try:
                        self.aon_base_path = bps.base_path_select()
                        self.aon_filename = fe.filename_w_extension()

                        wb = xl.load_workbook(f"{self.aon_base_path + self.aon_filename}")

                        reading_blanks = 'N'
                        for array_idx in range(len(self.ARRAY_OF_NODES)):
                            for row_idx in range(len(self.ARRAY_OF_NODES[array_idx])):
                                for elmt_idx in range(len(self.ARRAY_OF_NODES[array_idx][row_idx])):
                                    # py INDEXED TO 0, EXCEL (openpyxl) ROWS & COLUMNS INDEXED TO 1, THEREFORE THE "+1"s BELOW
                                    self.ARRAY_OF_NODES[array_idx][row_idx][elmt_idx] = \
                                        wb[f'NODE {array_idx}'].cell(row_idx + 1, elmt_idx + 1).value
                                    if wb[f'NODE {array_idx}'].cell(row_idx + 1, elmt_idx + 1).value == '':
                                        reading_blanks = 'Y'

                        if reading_blanks == 'Y': print(
                            f'\n*** ATTEMPTING TO READ BLANK CELLS IN SOURCE FILE.  CHECK CONFIG. ***\n')
                        else: print(f'\nNODE SEED FILE READ SUCCESSFUL.\n')

                        break

                    except:
                        if vui.validate_user_str(f'\nError trying to read {self.aon_base_path + self.aon_filename}' + \
                                                    f'\nTry again(t) or abort file read(a)? > ', 'AT') == 'A': break


            elif seed_method == 'C':
                if vui.validate_user_str('Estimate seed value? (y/n) > ', 'YN') == 'Y':
                    print('No estimating capability at this time.')

                self.node_seed = vui.validate_user_float('Enter node seed > ')

                for nle in self.LIST_OF_NN_ELEMENTS:
                    self.ARRAY_OF_NODES[nle[0]][nle[1]][nle[2]] = self.node_seed


            elif seed_method == 'R':
                dist_mean = vui.validate_user_float(f'\nEnter mean of seed distribution > ')
                dist_stdev = vui.validate_user_float(f'Enter stdev of seed distribution > ', min=1e-99)
                for nle in self.LIST_OF_NN_ELEMENTS:
                    self.ARRAY_OF_NODES[nle[0]][nle[1]][nle[2]] = np.random.normal(dist_mean, dist_stdev)


            elif seed_method == 'X':
                # XAVIER-HE
                # N(0, sqrt(2 / (NEURONS IN CURRENT NODE + NEURONS IN PREVIOUS NODE)))
                for nle in self.LIST_OF_NN_ELEMENTS:
                    if nle[0] == 0:
                        # NO PREVIOUS NODE FOR FIRST NODE (REF MATERIAL DOESN'T SAY HOW TO HANDLE THIS)
                        _ = np.random.normal(0, np.sqrt(2 / (len(self.ARRAY_OF_NODES[nle[0]]))), 1)
                    else:
                        _ = np.random.normal(0,
                                            np.sqrt(2 / (len(self.ARRAY_OF_NODES[nle[0]]) + len(self.ARRAY_OF_NODES[nle[0] - 1]))),
                                            1)

                    self.ARRAY_OF_NODES[nle[0]][nle[1]][nle[2]] = _

            rows = 5
            print(
                f'\n[:{len(self.ARRAY_OF_NODES[0][:rows])}][:{len(self.ARRAY_OF_NODES[0][0][:rows])}] of first node looks like:')
            [print(self.ARRAY_OF_NODES[0][_][:rows]) for _ in range(len(self.ARRAY_OF_NODES[0][:rows]))]

            if vui.validate_user_str('Accept node seed? (y/n) > ', 'YN') == 'Y':
                break


    def config(self):

        # configure BATCH_SIZE vector specifying mini-batch size for each iteration
        if self.nn_config == 'B':  # batch method(b)
            self.batch_size_vector_config()

        # change convergence method(c)
        if self.nn_config == 'C':
            if self.gd_method == 'C':
                print(f'Current method is COORDINATE, dont need to change gradient descent convergence method.')
            else:
                print(f'Current convergence method is {self.conv_method}.')
                self.conv_method = \
                    self.gd_convergence_str()[ls.list_single_select(self.gd_convergence_methods(), 'Select gradient descent convergence method', 'idx')[0]]

                if self.conv_method == 'G':
                    if vui.validate_user_str(f'\nUse momentum for gradient descent? (y/n) > ', 'YN') == 'Y':
                        self.gd_momentum()
                    else: self.momentum_weight = 0
                elif self.conv_method == 'A':
                    print(f'\nSet momentum weight for ADAM > ')
                    self.gd_momentum()


        # activation constant(d)
        if self.nn_config == 'D':
            self.activation_constant_fxn()

        # change cost function(f)
        if self.nn_config == 'F':
            print(f'\nFinal link in NN is {self.SELECT_LINK_FXN[-1]}')
            self.cost_fxn = list(ec.cost_functions().keys())[ls.list_single_select(list(ec.cost_functions().values()),
                                                                            f'Select cost function', 'idx')[0]].upper()

            # CHECK IF TRYING TO RUN LEAST SQUARES BUT ACCIDENTALLY PUT -LOG LIKELIHOOD
            if self.SELECT_LINK_FXN[-1].upper() == 'NONE' and self.cost_fxn in ['minus log-likelihood(l)', 'l', 'L']:
                print(f'BEAR WANTED TO SEE IN NNCoreConfigCOde what self.cost_fxn is: {self.cost_fxn}')
                print(f'\n*** YOU HAVE SELECTED FINAL LINK FUNCTION "None" WITH MINUS LOG-LIKELIHOOD AS COST FUNCTION ***\n')
                _ = vui.validate_user_str(f'Change final link(l), change cost function(c), ignore(i), quit(q) > ', 'LCIQ')

                if _ == 'Q': raise Exception(f'\nPROGRAM TERMINATED BY USER.\n')
                elif _ == 'I': pass
                elif _ == 'L':
                    self.SELECT_LINK_FXN[-1] = ls.list_single_select(lf.define_links(), f'Select final link function > ', 'value')
                elif _ == 'C':
                    self.cost_fxn = list(ec.cost_functions().keys())[ls.list_single_select(list(ec.cost_functions().keys()),
                                                                                           f'Select cost function', 'idx')[0]].upper()

            # CHECK IF TRYING TO RUN LEAST SQUARES BUT ACCIDENTALLY PUT -LOG LIKELIHOOD
            if self.cost_fxn in ['minus log-likelihood(l)', 'l', 'L'] and (np.min(self.TARGET_VECTOR)!=0 and np.max(self.TARGET_VECTOR!=1)):
                print(f'BEAR WANTED TO SEE IN NNCoreConfigCOde what self.cost_fxn is: {self.cost_fxn}')
                print(f'\n*** YOU HAVE SELECTED MINUS LOG-LIKELIHOOD AS COST FUNCTION BUT TARGET IS NOT BINARY ***\n')
                _ = vui.validate_user_str(f'Change cost function(c), ignore(i), quit(q) > ', 'LCIQ')

                if _ == 'Q': raise Exception(f'\nPROGRAM TERMINATED BY USER.\n')
                elif _ == 'I':
                    self.cost_fxn = list(ec.cost_functions().keys())[
                        ls.list_single_select(list(ec.cost_functions().keys()), f'Select cost function', 'idx')[0]].upper()
                elif _ == 'C':
                    pass


        # config gd_method(g)
        if self.nn_config == 'G':
            self.gd_method = vui.validate_user_str('\nMethod: gradient(g) coordinate(c) > ', 'CG')

            self.gd_iterations = vui.validate_user_int('\nEnter number of iterations > ', min=1)

            if self.gd_method == 'G':
                self.conv_method = \
                    self.gd_convergence_str()[ls.list_single_select(self.gd_convergence_methods(), 'Select gradient descent convergence method', 'idx')[0]]

                if self.conv_method == 'G':
                    if vui.validate_user_str(f'\nUse momentum for gradient descent? (y/n) > ', 'YN') == 'Y':
                        self.gd_momentum()
                    else:
                        self.momentum_weight = 0
                elif self.conv_method == 'A':
                    print(f'\nSet momentum weight for ADAM > ')
                    self.gd_momentum()

            elif self.gd_method == 'C':
                print('    ' + f'\nCURRENT METHOD is COORDINATE so LEARNING RATES must be (re)filled.')
                print('    ' + f'Last LEARNING RATE LOOKED LIKE:')
                self.print_learning_rates()
                self.LEARNING_RATE = lrs.SetLearningRate(self.LEARNING_RATE, self.ARRAY_OF_NODES,
                                                         self.SELECT_LINK_FXN, self.lr_method, self.gd_iterations).learning_rate_run()

            self.batch_size_vector_config()

        #change lr_method(h)
        if self.nn_config == 'H':
            if self.gd_method == 'C' or self.conv_method == 'N':
                print(f'\nMETHOD IS COORDINATE OR GRADIENT DESCENT CONVERGENCE METHOD IS COORDINATE, DONT NEED TO CHANGE LR METHOD.')
            else:
                self.lr_method = vui.validate_user_str('\nLearning rate method: user-specified(c) or Cauchy(s) > ', 'CS')
                if self.lr_method == 'C':
                    self.LEARNING_RATE = lrs.SetLearningRate(self.LEARNING_RATE, self.ARRAY_OF_NODES,
                                        self.SELECT_LINK_FXN, self.lr_method, self.gd_iterations).learning_rate_run()

                elif self.lr_method == 'S':
                    print('\nCauchy selected.\n')  # THIS WILL BE HANDLED IN NNRun

        #change # of iterations(v)
        if self.nn_config == 'V':
            print(f'\nChanging iterations')
            BATCH_DICT = {'B':'BATCH', 'M': 'MINI-BATCH', 'S': 'STOCHASTIC'}
            print(f'Last batch_method was {BATCH_DICT[self.batch_method]} and BATCH_SIZE[:10] (# of examples per iteration) was \n{self.BATCH_SIZE[:10]}\n')
            self.gd_iterations = vui.validate_user_int('Enter number of iterations > ', min=1)

            #IF ITERATIONS IS CHANGED, BATCH_SIZE VECTOR MUST ALSO BE RE-FILLED
            print(f'\nIterations has changed so BATCH_SIZE vector must be (re)filled.\n')
            self.batch_size_vector_config()

            #IF ITERATIONS IS CHANGED, LEARNING RATE VECTOR(S) MUST ALSO BE RE-FILLED
            print(f'\nIterations has changed so LEARNING RATE must be (re)filled.\n')
            self.print_learning_rates()
            self.LEARNING_RATE = lrs.SetLearningRate(self.LEARNING_RATE, self.ARRAY_OF_NODES, self.SELECT_LINK_FXN,
                self.lr_method, self.gd_iterations).learning_rate_run()

        # adjust learning rates(j)
        if self.nn_config == 'J':
            self.print_learning_rates()
            self.LEARNING_RATE = lrs.SetLearningRate(self.LEARNING_RATE, self.ARRAY_OF_NODES, self.SELECT_LINK_FXN,
                self.lr_method, self.gd_iterations).learning_rate_run()

        # adjust convergence kill(k)
        if self.nn_config == 'K':
            print(f'\nCurrent iterations with no improvement to kill loop: {self.conv_kill}')
            self.conv_kill = vui.validate_user_int('Enter iterations until kill > ', min=1, max=10000)

            print(f'\nCurrent min % change to avert kill: {self.pct_change}')
            self.pct_change = vui.validate_user_float(f'Enter min % change required to avert kill > ', min=0, max=1000)

            print(f'\nCurrent convergence end method : {self.conv_end_method}')
            self.conv_end_method = dict({'K':'KILL', 'P':'PROMPT'})[
                vui.validate_user_str(f'Enter new convergence end method - kill(k) or prompt(p) > ', 'KP')]

        # change link function(l)
        if self.nn_config == 'L':
            if len(self.SELECT_LINK_FXN) > 0:
                pass
                # pnnc.print_nn_config(SELECT_LINK_FXN, NEURONS)   CHOOSING PRINT COMMAND LOCATION IN select_link_fxn
            else:
                print('Link functions empty.')
            self.SELECT_LINK_FXN = slf.select_link_fxn(self.ARRAY_OF_NODES, lf.define_links(), self.SELECT_LINK_FXN, self.NEURONS)


        # momentum for gradient descent(m)
        if self.nn_config == 'M':
            if self.gd_method == 'N' or self.conv_method == 'C':
                print(f'\nMETHOD IS COORDINATE OR GRADIENT DESCENT CONVERGENCE METHOD IS COORDINATE, MOMENTUM IS NOT NEEDED.')
            else:
                if self.conv_method == 'G':
                    print(f'\nSet momentum weight for conventional gradient descent > ')
                    self.gd_momentum()
                elif self.conv_method == 'A':
                    print(f'\nSet momentum weight for ADAM > ')
                    self.gd_momentum()


        #non-neg coeffs(n)
        if self.nn_config == 'N':
            self.non_neg_coeffs = vui.validate_user_str('\nNon-negative NN parameters? (y/n) > ', 'YN')


        # initialize nodes(r)
        if self.nn_config == 'R':
            # SET PARAMETERS FOR THE NEURAL NETWORK (NODES, NEURONS, NODE SEED)


            ####################################################################################################################
            ####################################################################################################################
            # ARRAY OF NODES BUILD CONFIG ######################################################################################
            while True:
                # NUMBER OF NODES
                self.nodes = vui.validate_user_int(f'\nSelect number of nodes (NN Matrices) > ', min=1, max=10)

                # IF NODES = 1, "activation_constant" is handled by adding "intercept" to BIG_MATRIX
                if self.nodes >= 2:
                    self.activation_constant_fxn()


                # LINK FUNCTIONS, NUMBER OF NEURONS
                print('')
                self.SELECT_LINK_FXN.clear()
                self.NEURONS.clear()
                for node in range(self.nodes):
                    if node == self.nodes - 1:  # IF ON THE LAST NODE
                        use_multi_out = 'N'
                        if self.number_of_labels > 1:
                            # IF TARGET HAS MORE THAN 1 VECTOR (LABEL), SOFTMAX OR MULTI-OUT MUST BE USED ON FINAL NODE
                            print(
                                f'TARGET HAS MORE THAN 1 VECTOR, {lf.define_multi_out_links()} MUST BE USED ON FINAL NODE')
                            selected_fxn = ls.list_single_select(lf.define_multi_out_links(),
                                                                 'Select final link',
                                                                 'value'
                                                                 )[0]

                            # IF SOFTMAX IN USE JUST SKIP OUT AND FINISH "SELECT_LINK_FXN"
                            # OTHERWISE, CREATE AN INDICATOR THAT MULTI-OUT IS IN USE
                            if selected_fxn == 'Multi-out':
                                use_multi_out = 'Y'
                                selected_fxn = ls.list_single_select(
                                    lf.define_single_out_links(),
                                    'Select final link function',
                                    'value'
                                )[0]

                        elif self.number_of_labels == 1:
                            # IF TARGET HAS 1 VECTOR (LABEL), FINAL NEURON MUST BE 1 AND SOFTMAX CAN'T BE USED
                            print('TARGET HAS 1 VECTOR, FINAL NODE MUST HAVE 1 NEURON AND CANT BE SOFTMAX')

                            selected_fxn = ls.list_single_select(
                                [*lf.define_single_out_links()],
                                f'Select link function for node {node}',
                                'value'
                            )[0]

                        self.SELECT_LINK_FXN.append(selected_fxn)
                        # APPEND "MULTI-OUT" TO SELECT_LINK_FXN TO INDICATE IN OUTPUT_VECTOR_CALC TO DO FINAL TRANSPOSE
                        if use_multi_out == 'Y': self.SELECT_LINK_FXN.append('Multi-out')
                        self.NEURONS.append(self.number_of_labels)  # NEURONS FOR LAST NODE ALWAYS = # OF LABELS

                        print(f'{selected_fxn}, {self.number_of_labels} neurons(s) selected for node {node}.\n')

                        break

                    else:  # IF NOT ON THE LAST NODE
                        selected_fxn = \
                            ls.list_single_select([x for x in lf.define_links() if x != 'Multi-out'],
                                                  f'Select link function for node {node}',
                                                  'value'
                                                  )[0]

                        self.SELECT_LINK_FXN.append(selected_fxn)
                        self.NEURONS.append(
                            vui.validate_user_int(f'\nEnter # of neurons for node {node}, {selected_fxn} > ',
                                                  min=1)
                        )

                        print(f'{selected_fxn}, {self.NEURONS[-1]} neuron(s) selected for node {node}.\n')

                if len(self.NEURONS) != len([x for x in self.SELECT_LINK_FXN if x != 'Multi-out']):
                    raise Exception(f'#NEURONS != #LINK, MISSION ABORT.')

                if len(self.NEURONS) != self.nodes or len([x for x in self.SELECT_LINK_FXN if x != 'Multi-out']) != self.nodes:
                    raise Exception(f'MISMATCH IN NEURON OR LINK VECTOR AGAINST # OF NODES, MISSION ABORT.')

                pnnc.print_nn_config(self.SELECT_LINK_FXN, self.NEURONS)

                if vui.validate_user_str('Accept node / link configuration? (y/n) > ', 'YN') == 'Y':
                    # self.array_of_nodes_build()   # BEAR TOOK THIS OUT 5/18/23 PROMPTING FOR NODE SEED SEQUENCE 2X....
                    # THINK THAT 'BR' CONFIG SEQUENCE IS HITTING THIS THEN THE "r" DOWN THE LINE IN THIS LOOP.
                    # IF EVERYTHING IS OK DELETE THESE NOTES
                    break
            # END ARRAY OF NODES BUILD CONFIG ##################################################################################
            ####################################################################################################################
            ####################################################################################################################


        # reset seed(p)
        if self.nn_config == 'P':
            self.iteration = 0
            self.node_seed_fxn()


        # (q)
        if self.nn_config == 'Q':
            # PLACEHOLDER
            pass


        #secret fxn to build AON after loading config via standard config(r)  --- cannot be accessed thru the menu 2-9-22
        if self.nn_config == 'R':
            self.array_of_nodes_build()


        #adjust regularization(s)
        if self.nn_config == 'S':
            print(f'\nCurrent regularization type: {self.rglztn_type}')
            self.rglztn_type = {'1':'L1', '2':'L2'}[vui.validate_user_str(f'Enter regularization type, L1(1) or L2(2) > ', '12')]
            print(f'\nCurrent regularization factor: {self.rglztn_fctr}')
            self.rglztn_fctr = vui.validate_user_float(f'Enter regularization factor (enter 0 for no regularization) > ', min=0, max=float('inf'))


        # allow summary print / interval(t)'
        if self.nn_config == 'T':
            self.allow_summary_print = vui.validate_user_str(f'\nAllow summary print? (y/n) > ', 'YN')
            if self.allow_summary_print == 'N': self.summary_print_interval = 1e12
            else: self.summary_print_interval = vui.validate_user_int(f'Enter print interval > ', min=1, max=float('inf'))


        #restart iterations(w)
        if self.nn_config == 'W':
            print('\nIteration counter set to 0')
            self.iteration = 0


        # accept gd setup / continue(a)
        if self.nn_config == 'A':
            self.gd_iterations = vui.validate_user_str('Allow statistics print for each iteration? (y/n) > ', 'YN')
            # OUTPUT & ERROR HERE FOR STARTING CONDITION IN
            try:
                self.OUTPUT_VECTOR = ovc.output_vector_calc(self.DATA, self.ARRAY_OF_NODES,
                                               self.SELECT_LINK_FXN, self.OUTPUT_VECTOR, self.activation_constant)

                print(f'TEST OUTPUT_VECTOR[:{min(10, len(self.OUTPUT_VECTOR))}][:{min(10, len(self.OUTPUT_VECTOR[0]))}] = ')
                [print([f'{y + 1})'.rjust(3, ' ')] + [f'{x:4.2f}'.rjust(6, " ") for x in self.OUTPUT_VECTOR[y][:10]]) for y in
                 range(10)]

                self.new_error_start = ec.error_calc(self.ARRAY_OF_NODES, self.TARGET_VECTOR, self.OUTPUT_VECTOR, self.cost_fxn, self.new_error_start,
                                                self.SELECT_LINK_FXN, self.rglztn_type, self.rglztn_fctr)
            except:
                print(sys.exc_info()[0])
                print(f'\n*** Crash trying to calculate test OUTPUT VECTOR leaving NNCoreConfigCode ***\n')

        return self.return_fxn()



if __name__ == '__main__':
    pass




















