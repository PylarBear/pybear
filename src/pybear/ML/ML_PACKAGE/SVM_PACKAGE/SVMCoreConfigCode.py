import openpyxl as xl
from data_validation import validate_user_input as vui
from general_list_ops import list_select as ls
from ML_PACKAGE.SVM_PACKAGE import svm_kernels as sk, svm_error_calc as sec
from read_write_file.generate_full_filename import base_path_select as bps, filename_enter as fe


class SVMCoreConfigCode:

    def __init__(self, svm_config, margin_type, C, cost_fxn, kernel_fxn, constant, exponent, sigma, alpha_seed,
                 alpha_selection_alg, max_passes, tol, SMO_a2_selection_method, conv_kill, pct_change, conv_end_method):

        self.svm_config = svm_config

        self.margin_type = margin_type
        self.C = C
        self.cost_fxn = cost_fxn

        self.kernel_fxn = kernel_fxn
        self.constant = constant
        self.exponent = exponent
        self.sigma = sigma

        self.alpha_seed = alpha_seed
        self.alpha_selection_alg = alpha_selection_alg
        self.max_passes = max_passes
        self.tol = tol
        self.SMO_a2_selection_method = SMO_a2_selection_method
        self.conv_kill = conv_kill
        self.pct_change = pct_change
        self.conv_end_method = conv_end_method

        self.SUPPORT_VECTORS = []
        self.SUPPORT_TARGETS = []
        self.SUPPORT_ALPHAS = []
        self.b = 0

        self.boundary_base_path = ''
        self.boundary_filename = ''


    def return_fxn(self):
        return self.SUPPORT_VECTORS, self.SUPPORT_TARGETS, self.SUPPORT_ALPHAS, self.b, self.margin_type, self.C, \
               self.cost_fxn, self.kernel_fxn, self.constant, self.exponent, self.sigma, \
               self.alpha_seed, self.alpha_selection_alg, self.max_passes, self.tol, self.SMO_a2_selection_method, \
               self.conv_kill, self.pct_change, self.conv_end_method


    def max_passes_(self):
        self.max_passes = vui.validate_user_int(f'Enter maximum number of passes > ', min=1)


    def tol_(self):
        self.tol = vui.validate_user_float(f'Enter KKT tolerance (usually around 0.001) > ', min=1e-8)


    def a2_selection_method(self):
        # FOR SMO ONLY
        self.SMO_a2_selection_method = {'R':'RANDOM', 'P':'PLATT'}[
                            vui.validate_user_str(f'Enter a2 selection method for SMO --- Random(r) Platt(p) > ', 'RP')]


    def constant_(self):
        self.constant = vui.validate_user_float(f'Enter constant [c] > ', min=float('-inf'), max=float('inf'))


    def exponent_(self):
        self.exponent = vui.validate_user_float(f'Enter exponent [d] > ', min=float('-inf'), max=float('inf'))


    def sigma_(self):
        self.sigma = vui.validate_user_float(f'\nEnter sigma > ', min=float('-inf'), max=float('inf'))


    def convergence_kill(self):
        print(f'\nCurrent full-epoch passes with no KKT violator improvement to kill loop: {self.conv_kill}')
        self.conv_kill = vui.validate_user_int('Enter full-epoch passes until kill > ', min=1, max=10000)

        print(f'\nCurrent min % change to avert kill: {self.pct_change}')
        self.pct_change = vui.validate_user_float(f'Enter min % change required to avert kill > ', min=0, max=1000)

        print(f'\nCurrent convergence end method : {self.conv_end_method}')
        self.conv_end_method = dict({'K': 'KILL', 'P': 'PROMPT'})[
            vui.validate_user_str(f'Enter new convergence end method - kill(k) or prompt(p) > ', 'KP')]


    def config(self):

        # set margin type(j)
        if self.svm_config == 'J':
            self.margin_type = {'S':'SOFT', 'H':'HARD'}[vui.validate_user_str('\nSelect margin type SOFT(s) HARD(h) > ', 'SH')]


        # load boundary info from file(b)
        if self.svm_config == 'B':

            while True:
                try:
                    self.boundary_base_path = bps.base_path_select()
                    self.boundary_filename = fe.filename_w_extension()

                    wb = xl.load_workbook(f"{self.boundary_base_path + self.boundary_filename}")

                    if False in map(lambda x: x in wb, ('SUPPORT VECTORS', 'SUPPORT TARGETS', 'SUPPORT ALPHAS', 'b')):
                        print(f'\n*** BOUNDARY EXCEL FILE MUST CONSTAIN 4 TABS NAMED "SUPPORT VECTORS", "SUPPORT TARGETS", ' + \
                              f'"SUPPORT ALPHAS", AND "b" ***')
                        raise Exception

                    self.SUPPORT_VECTORS = []
                    self.SUPPORT_TARGETS = []
                    self.SUPPORT_ALPHAS = []
                    self.b = 0

                    reading_blanks = False
                    row = 0
                    while wb['SUPPORT VECTORS'].cell(row, 1).value != '':
                        self.SUPPORT_VECTORS.append([])
                        col = 0
                        row += 1
                        while wb['SUPPORT VECTORS'].cell(row, col).value != '':
                            col += 1
                            self.SUPPORT_VECTORS[-1].append(wb['SUPPORT VECTORS'].cell(row, col).value)


                        if wb['SUPPORT TARGETS'].cell(row,1).value == '' or wb['SUPPORT ALPHAS'].cell(row,1).value == '':
                            reading_blanks = True

                        self.SUPPORT_TARGETS.append(wb['SUPPORT TARGETS'].cell(row,1).value)
                        self.SUPPORT_ALPHAS.append(wb['SUPPORT ALPHAS'].cell(row,1).value)

                        if wb['b'].cell(1,1).value == '': reading_blanks = True
                        self.b = wb[f'b'].cell(1, 1).value

                    if reading_blanks: print(
                        f'\n*** ATTEMPTING TO READ BLANK CELLS IN SOURCE FILE.  CHECK CONFIG. ***\n')
                    else: print(f'\nBOUNDARY FILE READ SUCCESSFUL.\n')

                    break

                except:
                    if vui.validate_user_str(f'\nError trying to read {self.boundary_base_path + self.boundary_filename}' + \
                                                f'\nTry again(t) or abort file read(a)? > ', 'AT') == 'A': break
                    else: break


        # set regularization constant(c):
        if self.svm_config == 'C':
            if self.margin_type == 'SOFT':
                self.C = vui.validate_user_float('Enter regularization constant C (upper bound on alpha) > ', min=0, max=float('inf'))
            else:
                print(f'Hard margin is being used therefore C is infinity and cannot be changed.')
                self.C = float('inf')


        # change cost function(f)
        if self.svm_config == 'F':
            _ = ls.list_single_select(list(sec.cost_functions().values()), f'Select cost function', 'value')[0]
            self.cost_fxn = {v:k for k,v in sec.cost_functions().items()}[_]


        # select kernel(g) ###############################################################################################
        if self.svm_config == 'G':
            self.kernel_fxn = ls.list_single_select(sk.kernel_function_list(), f'\nSelect kernel function', 'value')[0]

            if self.kernel_fxn == 'LINEAR': pass
            elif self.kernel_fxn == 'POLYNOMIAL':
                print(f'\nPolynomial format is (DOT + c)^d')
                self.constant_()
                self.exponent_()
            elif self.kernel_fxn == 'GAUSSIAN':
                self.sigma_()
        # END select kernel(k) ###############################################################################################


        # set alpha seed(l)
        if self.svm_config == 'L':
            self.alpha_seed = vui.validate_user_float('\nEnter seed for alpha vector > ')


        # alpha selection algorithm(h) #######################################################################################
        if self.svm_config == 'H':
            while True:
                __ = vui.validate_user_str(f'Enter alpha optimization algorithm SMO(s) Chunk(c) Osuna(o) > ', 'SCO')
                self.alpha_selection_alg = {'S': 'SMO', 'C': 'CHUNK', 'O': 'OSUNA'}[__]
                if self.alpha_selection_alg.upper() == 'SMO':
                    self.a2_selection_method()
                    self.max_passes_()
                    self.tol_()
                    break

                elif self.alpha_selection_alg.upper() == 'CHUNK':
                    print(f'\n *** CHUNK NOT AVAILABLE YET *** \n')
                    continue

                elif self.alpha_selection_alg.upper() == 'OSUNA':
                    print(f'\n *** OSUNA NOT AVAILABLE YET *** \n')
                    continue
        # END CONFIG ALPHA SELECTION ALGORITHM #######################################################################################

        # set a2 selection method(d)
        if self.svm_config == 'D':
            self.a2_selection_method()

        # set max passes(k)
        if self.svm_config == 'K':
            self.max_passes_()

        # set tolerance(m)
        if self.svm_config == 'M':
            self.tol_()

        # set polynomial constant(n)
        if self.svm_config == 'N':
            self.constant_()

        # set polynomial exponent(p)
        if self.svm_config == 'P':
            self.exponent_()

        # set gaussian sigma(r)
        if self.svm_config == 'R':
            self.sigma_()

        # adjust passes to convergence kill(s)
        if self.svm_config == 'S':
            self.convergence_kill()


        return self.return_fxn()
















